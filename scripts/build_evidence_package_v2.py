"""
갑우그룹 경영증거패키지 v2 (2026-04-20)

v1 대비 변경점:
  1) 동행복권(8668700833) 지입(아방) 확정 반영
  2) 세금계산서 3개년(23·24·25) + 사업자번호 기반 지입 태깅
  3) 지입 거래처 분리 시트 신규 (일반 vs 지입 수익구조 대비)
  4) 갑우 수익성 3개년 확장 시트 신규 (viewGabwoo_마감 체인 23·24·25)

출처:
  - 세금계산서: /Users/jack/dev/gabwoo/23,24년 제품매출현황(갑우,비피,더원).xls
                /Users/jack/dev/gabwoo/25년 제품매출현황(갑우,비피,더원).xls
  - ERP (SELECT only): SNOTES MSSQL, viewGabwoo_마감 + PUR_POH → PRT_SOH 체인

출력:
  AX 전환 계획/2026-04-20/갑우그룹_경영증거패키지_v2_20260420.xlsx
"""
import sys
import re
from pathlib import Path
from datetime import datetime

try:
    import pymssql
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"pip3 install pymssql pandas openpyxl xlrd: {e}"); sys.exit(1)

ROOT = Path("/Users/jack/dev/gabwoo")
ENV_FILE = ROOT / "견적계산기/.env.local"
XLS_2324 = ROOT / "23,24년 제품매출현황(갑우,비피,더원).xls"
XLS_25 = ROOT / "25년 제품매출현황(갑우,비피,더원).xls"
OUT_DIR = ROOT / "AX 전환 계획" / "2026-04-20"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_XLSX = OUT_DIR / f"갑우그룹_경영증거패키지_v2_{datetime.now():%Y%m%d}.xlsx"

# 지입(=아방) 거래처 — 사업자번호 기반 (대시 제거한 10자리)
# 출처: memory/reference_avang_jiip.md + Jack 2026-04-20 (동행복권 확정)
JIIP_CUST_BIZ = {
    "1078140772": "(주)갑우문화사(관계)",
    "1418115138": "(주)비피앤피(관계)",
    "4968602009": "(주)교원",
    "6678702103": "(주)교원구몬",
    "8668700833": "(주)동행복권",   # 2026-04-20 Jack 확인
}
# 이름 키워드 백업 (사업자번호 없는 로우에만 제한 적용)
# ⚠️ Jack 2026-04-20: "동행복권이 다. 더 없다" → 이투스·에듀윌은 지입 확정치 아님
# 이름 키워드는 사업자번호로 매칭된 지입 거래처의 '지점/계열' 변형만 커버
JIIP_NAME_KEYWORDS = ["교원구몬", "교원 ", "동행복권"]  # 사업자번호 없는 행 백업

# 간접비 배분 — 2025 감사보고서 기준 (갑우 매출 대비 간접비율)
# v1 gabwoo_profitability.json 재계산: 52.9% (판관비+제조간접비 - 용지매입)
# 한계: 23·24년 별도 비율 데이터 없어 일괄 적용. 보고서 주석 명시.
INDIRECT_RATIO = 0.529

# 분석 제외 거래처명 키워드 (특수 케이스)
# 프린트뱅크 이관으로 매출·원가가 전사 실적에서 왜곡되는 거래처 (본부장 2026-04-16 확인)
EXCLUDE_NAME_KEYWORDS = ["지에스리테일"]


def is_excluded(name):
    nm = name or ""
    return any(kw in nm for kw in EXCLUDE_NAME_KEYWORDS)

# 1억+ 임계 (v1 동일)
CUST_THRESHOLD = 100_000_000

# 스타일
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
NOTE_FONT = Font(italic=True, size=10, color="555555")
BORDER = Border(left=Side(style="thin", color="BFBFBF"),
                right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="thin", color="BFBFBF"))
JIIP_FILL = PatternFill("solid", fgColor="FFF2CC")  # 연노랑


# ─────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────
def load_env():
    env = {}
    for line in ENV_FILE.read_text().splitlines():
        if "=" in line and not line.strip().startswith("#"):
            k, v = line.split("=", 1); env[k.strip()] = v.strip()
    return env


def normalize_name(name):
    if not isinstance(name, str): return ""
    s = name.replace("(주)", "").replace("(사)", "").replace("주식회사", "")
    s = re.sub(r"\s+", "", s).strip()
    return s


def normalize_biz(biz):
    if not isinstance(biz, str): return ""
    return re.sub(r"[^0-9]", "", biz)


def is_jiip(biz_raw, name_raw):
    biz = normalize_biz(biz_raw)
    if biz and biz in JIIP_CUST_BIZ:
        return True
    nm = name_raw or ""
    for kw in JIIP_NAME_KEYWORDS:
        if kw in nm:
            return True
    return False


# ─────────────────────────────────────────
# 1. 세금계산서 xls 파싱
# ─────────────────────────────────────────
def parse_tax_invoice_xls():
    rows = []
    files = [
        (XLS_2324, {"갑우23,24": "갑우", "비피23,24": "비피", "더원23,24": "더원"}),
        (XLS_25,   {"갑우": "갑우", "비피": "비피", "더원": "더원"}),
    ]
    for f, sheets in files:
        print(f"  📥 {f.name}")
        xl = pd.ExcelFile(f)
        for sn, firm in sheets.items():
            df = pd.read_excel(xl, sheet_name=sn)
            # 거래처명·날짜 모두 있는 행만 (월계/누계 제외)
            df = df[df["거래처명"].notna() & df["날짜"].notna()].copy()
            df["회사"] = firm
            df["연도"] = df["날짜"].astype(str).str[:4]
            df["거래처명"] = df["거래처명"].astype(str).str.strip()
            df["사업자번호"] = df["사업자번호"].astype(str).apply(normalize_biz)
            df["대변"] = pd.to_numeric(df["대변"], errors="coerce").fillna(0)
            df["key"] = df["거래처명"].apply(normalize_name)
            rows.append(df[["회사","연도","거래처명","사업자번호","key","대변"]])
    all_df = pd.concat(rows, ignore_index=True)
    # 연도 필터 2023~2025
    all_df = all_df[all_df["연도"].isin(["2023","2024","2025"])].copy()
    return all_df


def aggregate_by_customer(raw):
    # 거래처명 정규화된 key + 회사 + 연도로 집계
    # 사업자번호는 각 거래처의 대표값(가장 자주 쓰인) 채택
    biz_map = (raw[raw["사업자번호"] != ""]
               .groupby(["회사", "key"])["사업자번호"]
               .agg(lambda s: s.mode().iloc[0] if not s.mode().empty else "")
               .to_dict())
    name_map = (raw.groupby(["회사", "key"])["거래처명"]
                .agg(lambda s: s.mode().iloc[0] if not s.mode().empty else s.iloc[0])
                .to_dict())
    piv = (raw.groupby(["회사", "key", "연도"])["대변"].sum()
           .unstack(fill_value=0).reset_index())
    for y in ["2023", "2024", "2025"]:
        if y not in piv.columns: piv[y] = 0
    piv["거래처명"] = piv.apply(lambda r: name_map.get((r["회사"], r["key"]), r["key"]), axis=1)
    piv["사업자번호"] = piv.apply(lambda r: biz_map.get((r["회사"], r["key"]), ""), axis=1)
    piv["지입여부"] = piv.apply(lambda r: "지입" if is_jiip(r["사업자번호"], r["거래처명"]) else "일반", axis=1)
    piv["제외여부"] = piv["거래처명"].apply(lambda n: "제외(프린트뱅크 이관)" if is_excluded(n) else "")
    piv["max_year"] = piv[["2023", "2024", "2025"]].max(axis=1)
    piv = piv[piv["max_year"] >= CUST_THRESHOLD].copy()
    piv["변화(25-23)"] = piv["2025"] - piv["2023"]
    def pct(a, b):
        if b == 0: return None
        return (a - b) / b
    piv["YoY_24"] = piv.apply(lambda r: pct(r["2024"], r["2023"]), axis=1)
    piv["YoY_25"] = piv.apply(lambda r: pct(r["2025"], r["2024"]), axis=1)
    piv["변화율_23→25"] = piv.apply(lambda r: pct(r["2025"], r["2023"]), axis=1)

    def pattern(r):
        a, b, c = r["2023"], r["2024"], r["2025"]
        if a == 0 and b == 0 and c > 0: return "🆕 신규(25~)"
        if a > 0 and c == 0: return "🚨 이탈(23→25)"
        if a == 0 and c > 0: return "🆕 신규"
        chg = r["변화율_23→25"]
        if chg is None: return "—"
        if chg <= -0.2: return "📉 축소"
        if chg >= 0.2: return "📈 성장"
        return "➡️ 유지"
    piv["패턴"] = piv.apply(pattern, axis=1)
    piv = piv.sort_values("2025", ascending=False).reset_index(drop=True)
    return piv[["회사","거래처명","사업자번호","key","2023","2024","2025",
                "변화(25-23)","변화율_23→25","YoY_24","YoY_25","패턴","지입여부","제외여부"]]


# ─────────────────────────────────────────
# 2. 갑우 용지원가 체인 3개년 (ERP)
# ─────────────────────────────────────────
def fetch_gabwoo_cost_3y(conn):
    print("  📥 viewGabwoo_마감 3개년 체인 (23·24·25)...")
    # ⚠️ NO_LINE 매칭 필수 — 없으면 같은 용지원가가 여러 WO에 중복 집계됨 (2026-04-17 세션 노트)
    sql = """
        SELECT YEAR(v.일자) AS 연도, soh.CD_PARTNER AS 거래처코드,
               c.NM_CUST AS 거래처명, SUM(v.공급가액) AS 용지원가
        FROM dbo.viewGabwoo_마감 v
        JOIN PUR_POH ph ON v.CustKey = ph.NO_PO AND ph.CD_FIRM='7000'
        JOIN PUR_POL pl ON ph.CD_FIRM = pl.CD_FIRM
                       AND ph.NO_PO = pl.NO_PO
                       AND pl.NO_LINE = v.NO_LINE
        JOIN PRT_WO wo ON pl.NO_WO = wo.NO_WO AND wo.CD_FIRM='7000'
        JOIN PRT_SOH soh ON wo.NO_SO = soh.NO_SO AND soh.CD_FIRM='7000'
        LEFT JOIN MAS_CUST c ON soh.CD_FIRM=c.CD_FIRM AND soh.CD_PARTNER=c.CD_CUST
        WHERE v.일자 >= '2023-01-01' AND v.일자 <= '2025-12-31'
          AND v.공급가액 > 0
          AND soh.CD_CUST_OWN = '10000'
        GROUP BY YEAR(v.일자), soh.CD_PARTNER, c.NM_CUST
    """
    df = pd.read_sql(sql, conn)
    df["연도"] = df["연도"].astype(str)
    df["key"] = df["거래처명"].apply(normalize_name)
    piv = df.pivot_table(index=["거래처코드", "key"], columns="연도",
                         values="용지원가", aggfunc="sum").fillna(0).reset_index()
    for y in ["2023", "2024", "2025"]:
        if y not in piv.columns: piv[y] = 0
    piv = piv.rename(columns={"2023": "용지원가_2023",
                              "2024": "용지원가_2024",
                              "2025": "용지원가_2025"})
    return piv[["거래처코드", "key", "용지원가_2023", "용지원가_2024", "용지원가_2025"]]


def build_gabwoo_profitability_3y(agg_df, cost_df):
    # 갑우 소속 + 일반·지입. 제외 거래처는 분석에서 배제
    g = agg_df[(agg_df["회사"] == "갑우") & (agg_df["제외여부"] == "")].copy()
    merged = g.merge(cost_df, on="key", how="left")
    for y in ["2023", "2024", "2025"]:
        merged[f"용지원가_{y}"] = merged[f"용지원가_{y}"].fillna(0)

    def compute(row, y):
        매출 = row[y]
        용지 = row[f"용지원가_{y}"] if row["지입여부"] == "일반" else 0
        간접 = 매출 * INDIRECT_RATIO
        마진 = 매출 - 용지 - 간접
        return pd.Series({
            f"매출_{y}": 매출,
            f"용지원가_{y}": 용지,
            f"용지원가_실측_{y}": row[f"용지원가_{y}"],
            f"간접비_{y}": 간접,
            f"추정마진_{y}": 마진,
            f"추정마진율_{y}": (마진 / 매출) if 매출 > 0 else None,
            f"용지원가율_{y}": (용지 / 매출) if 매출 > 0 else None,
        })

    out_rows = []
    for _, row in merged.iterrows():
        rec = {
            "거래처명": row["거래처명"],
            "사업자번호": row["사업자번호"],
            "거래처코드": row.get("거래처코드", ""),
            "지입여부": row["지입여부"],
        }
        for y in ["2023", "2024", "2025"]:
            rec.update(compute(row, y).to_dict())
        out_rows.append(rec)
    return pd.DataFrame(out_rows).sort_values("매출_2025", ascending=False).reset_index(drop=True)


# ─────────────────────────────────────────
# 3. 엑셀 빌드
# ─────────────────────────────────────────
def autosize(ws, max_width=50):
    for col in ws.columns:
        max_len = 0; col_letter = None
        for cell in col:
            if col_letter is None: col_letter = get_column_letter(cell.column)
            try:
                v = str(cell.value) if cell.value is not None else ""
                ln = sum(2 if ord(c) > 127 else 1 for c in v)
                if ln > max_len: max_len = ln
            except Exception: pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def write_cover(wb):
    ws = wb.create_sheet("표지")
    ws["A1"] = "갑우그룹 경영증거패키지 v2"
    ws["A1"].font = Font(bold=True, size=18, color="1F3864")
    ws.merge_cells("A1:F1")
    ws["A3"] = "세금계산서 기준 3개년(2023·2024·2025) 거래처별 매출 · 지입 분리 · 갑우 수익성 확장"
    ws.merge_cells("A3:F3")
    ws["A5"] = f"작성일: {datetime.now():%Y-%m-%d}  /  작성: AX전환팀"
    ws.merge_cells("A5:F5")
    ws["A7"] = "📌 v1 대비 변경점"
    ws["A7"].font = SECTION_FONT; ws["A7"].fill = SECTION_FILL
    ws.merge_cells("A7:F7")
    changes = [
        "1. 동행복권(8668700833) 지입(아방) 확정 반영 (Jack 2026-04-20)",
        "2. 사업자번호 기반 지입 태깅 — 거래처명 정규화 부정확성 보완",
        "3. [6_지입거래처_분리] 시트 신규 — 일반 vs 지입 수익구조 대비",
        "4. [7_갑우_수익성_3개년] 시트 신규 — viewGabwoo_마감 체인 23·24·25",
        "5. 나머지 시트(매출 추이·패턴분류)는 v1 구조 유지, 지입여부 컬럼만 추가",
    ]
    for i, t in enumerate(changes, start=8):
        ws[f"A{i}"] = t; ws[f"A{i}"].font = Font(size=11)
        ws.merge_cells(f"A{i}:F{i}")
    ws["A14"] = "⚠️ 이 보고서는 매출 숫자는 '세금계산서(확정치)' 기준, 수익성 숫자는 'ERP 역산+간접비 배분(추정)' 기준입니다."
    ws["A14"].font = Font(italic=True, color="C00000", size=10)
    ws.merge_cells("A14:F14")


def write_executive(wb, agg):
    ws = wb.create_sheet("1_경영진_1페이지")
    ws["A1"] = "갑우그룹 3개년 매출 추이 (세금계산서 기준)"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:H1")

    ws["A3"] = "■ 3사 연간 매출 (지입 포함 전체)"
    ws["A3"].font = Font(bold=True, size=11)
    hdr = ["회사", "2023", "2024", "2025", "변화(25-23)", "변화율", "비고"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
    r = 5
    for firm in ["갑우", "비피", "더원"]:
        sub = agg[agg["회사"] == firm]
        s23, s24, s25 = sub["2023"].sum(), sub["2024"].sum(), sub["2025"].sum()
        chg = s25 - s23
        rate = (chg / s23) if s23 > 0 else 0
        for j, v in enumerate([firm, s23, s24, s25, chg, rate, ""], 1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            if isinstance(v, (int, float)) and j <= 5:
                c.number_format = "#,##0"; c.alignment = Alignment(horizontal="right")
            elif j == 6:
                c.number_format = "0.0%"; c.alignment = Alignment(horizontal="right")
        r += 1
    # 합계
    t23, t24, t25 = agg["2023"].sum(), agg["2024"].sum(), agg["2025"].sum()
    for j, v in enumerate(["합계", t23, t24, t25, t25 - t23, (t25 - t23)/t23 if t23>0 else 0, ""], 1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = Font(bold=True); c.border = BORDER
        if isinstance(v, (int, float)) and j <= 5:
            c.number_format = "#,##0"
        elif j == 6:
            c.number_format = "0.0%"

    # 지입 제외 버전
    r += 2
    ws.cell(row=r, column=1, value="■ 3사 연간 매출 (지입 제외 — 제작공임 기준)").font = Font(bold=True, size=11)
    r += 1
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=r, column=j, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
    r += 1
    for firm in ["갑우", "비피", "더원"]:
        sub = agg[(agg["회사"] == firm) & (agg["지입여부"] == "일반")]
        s23, s24, s25 = sub["2023"].sum(), sub["2024"].sum(), sub["2025"].sum()
        chg = s25 - s23
        rate = (chg / s23) if s23 > 0 else 0
        for j, v in enumerate([firm, s23, s24, s25, chg, rate, "지입 사업자번호 제외"], 1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            if isinstance(v, (int, float)) and j <= 5:
                c.number_format = "#,##0"; c.alignment = Alignment(horizontal="right")
            elif j == 6:
                c.number_format = "0.0%"
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="■ 지입(아방) 거래처 — 매출 따로 보는 이유").font = Font(bold=True, size=11)
    r += 1
    notes = [
        "• 지입 = 거래처가 용지를 직접 구매·지급, 갑우는 제작공임만 받음",
        "• 세금계산서 금액은 제작공임만 반영되어 '일반 거래처 평균보다 낮아 보임'",
        "• 단가·수익성을 일반 거래처와 동일 기준으로 비교하면 안 됨 → 별도 시트(6_지입거래처_분리) 참고",
        f"• 지입 사업자번호(확정): {', '.join(JIIP_CUST_BIZ.values())}",
    ]
    for n in notes:
        ws.cell(row=r, column=1, value=n).font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
    autosize(ws)


def _write_trend_sheet(ws, df, title, cols_display):
    ws.cell(row=1, column=1, value=title).font = SECTION_FONT
    ws.cell(row=1, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols_display))
    for j, h in enumerate(cols_display, 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df[cols_display].itertuples(index=False), start=3):
        is_jiip_row = False
        try:
            idx = cols_display.index("지입여부")
            is_jiip_row = (row[idx] == "지입")
        except ValueError:
            pass
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = BORDER
            col_name = cols_display[j-1]
            if is_jiip_row: c.fill = JIIP_FILL
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")
                if "%" in col_name or "YoY" in col_name or "변화율" in col_name:
                    c.number_format = "0.0%"
                elif "마진율" in col_name or "원가율" in col_name:
                    c.number_format = "0.0%"
                else:
                    c.number_format = "#,##0"
    autosize(ws)


def write_decline_top(wb, agg):
    ws = wb.create_sheet("2_매출감소_TOP")
    df = agg[(agg["패턴"].isin(["📉 축소", "🚨 이탈(23→25)"])) &
             (agg["변화(25-23)"] < 0)].copy()
    df = df.sort_values("변화(25-23)").reset_index(drop=True).head(50)
    cols = ["회사", "거래처명", "2023", "2024", "2025",
            "변화(25-23)", "변화율_23→25", "YoY_24", "YoY_25", "패턴", "지입여부"]
    _write_trend_sheet(ws, df, "매출 감소 TOP 50 (23→25)", cols)


def write_growth_new(wb, agg):
    ws = wb.create_sheet("3_매출증가·신규")
    df = agg[(agg["패턴"].isin(["📈 성장", "🆕 신규(25~)", "🆕 신규"])) &
             (agg["2025"] > 0)].copy()
    df = df.sort_values("2025", ascending=False).reset_index(drop=True).head(50)
    cols = ["회사", "거래처명", "2023", "2024", "2025",
            "변화(25-23)", "변화율_23→25", "YoY_24", "YoY_25", "패턴", "지입여부"]
    _write_trend_sheet(ws, df, "매출 증가·신규 TOP 50 (25년)", cols)


def write_all(wb, agg):
    ws = wb.create_sheet("4_전체거래처")
    cols = ["회사", "거래처명", "사업자번호", "2023", "2024", "2025",
            "변화(25-23)", "변화율_23→25", "YoY_24", "YoY_25", "패턴", "지입여부", "제외여부"]
    _write_trend_sheet(ws, agg, f"전체 거래처 ({len(agg)}개사, 3년 중 1년 이상 1억+)", cols)


def write_pattern(wb, agg):
    ws = wb.create_sheet("5_패턴분류_요약")
    ws["A1"] = "회사별 × 패턴별 거래처 수 및 매출 변화 (지입 여부 포함)"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:H1")
    r = 3
    for firm in ["갑우", "비피", "더원"]:
        sub = agg[agg["회사"] == firm]
        ws.cell(row=r, column=1, value=f"■ {firm}").font = Font(bold=True, size=12)
        r += 1
        hdr = ["패턴", "거래처 수", "2023 합계", "2025 합계", "변화", "변화율", "지입 포함 수"]
        for j, h in enumerate(hdr, 1):
            c = ws.cell(row=r, column=j, value=h)
            c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
        r += 1
        for p, g in sub.groupby("패턴"):
            n = len(g)
            s23 = g["2023"].sum(); s25 = g["2025"].sum()
            chg = s25 - s23
            rate = (chg / s23) if s23 > 0 else None
            jiip = (g["지입여부"] == "지입").sum()
            for j, v in enumerate([p, n, s23, s25, chg, rate, jiip], 1):
                c = ws.cell(row=r, column=j, value=v); c.border = BORDER
                if isinstance(v, (int, float)):
                    if j in (3, 4, 5): c.number_format = "#,##0"
                    if j == 6 and v is not None: c.number_format = "0.0%"
            r += 1
        r += 1
    autosize(ws)


def write_jiip(wb, agg):
    ws = wb.create_sheet("6_지입거래처_분리")
    ws["A1"] = "지입(아방) 거래처 분리 분석 — 일반 거래처와 수익구조가 다름"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:L1")

    notes = [
        "• 지입(=아방): 거래처가 용지를 직접 구매·지급 → 세금계산서 금액은 갑우 제작공임만",
        "• 일반 거래처 단가와 비교 금지 — 매출 규모는 '제작공임 비용'이지 '총거래금액' 아님",
        "• 사업자번호 기반 매칭 (2026-04-20 기준 5개사). 이름 키워드는 백업 매칭.",
    ]
    r = 2
    for n in notes:
        c = ws.cell(row=r, column=1, value=n); c.font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        r += 1
    r += 1

    # 일반 vs 지입 요약
    summary_hdr = ["구분", "거래처 수", "2023 매출", "2024 매출", "2025 매출", "23→25 변화율"]
    for j, h in enumerate(summary_hdr, 1):
        c = ws.cell(row=r, column=j, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
    r += 1
    for firm in ["갑우", "비피", "더원"]:
        for kind in ["일반", "지입"]:
            sub = agg[(agg["회사"] == firm) & (agg["지입여부"] == kind)]
            if sub.empty: continue
            s23, s24, s25 = sub["2023"].sum(), sub["2024"].sum(), sub["2025"].sum()
            rate = (s25 - s23) / s23 if s23 > 0 else None
            vals = [f"{firm}·{kind}", len(sub), s23, s24, s25, rate]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=j, value=v); c.border = BORDER
                if kind == "지입": c.fill = JIIP_FILL
                if isinstance(v, (int, float)):
                    if j in (3, 4, 5): c.number_format = "#,##0"
                    if j == 6 and v is not None: c.number_format = "0.0%"
            r += 1
    r += 2

    # 지입 거래처 상세 리스트
    ws.cell(row=r, column=1, value="■ 지입 거래처 상세 (회사·연도별 매출)").font = Font(bold=True, size=11)
    r += 1
    cols = ["회사", "거래처명", "사업자번호", "2023", "2024", "2025",
            "변화(25-23)", "변화율_23→25", "YoY_24", "YoY_25", "패턴"]
    for j, h in enumerate(cols, 1):
        c = ws.cell(row=r, column=j, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
    r += 1
    jiip_df = agg[agg["지입여부"] == "지입"].sort_values(["회사", "2025"], ascending=[True, False])
    for row in jiip_df[cols].itertuples(index=False):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = BORDER; c.fill = JIIP_FILL
            col_name = cols[j-1]
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")
                if "%" in col_name or "YoY" in col_name or "변화율" in col_name:
                    c.number_format = "0.0%"
                else:
                    c.number_format = "#,##0"
        r += 1
    autosize(ws)


def write_gabwoo_profitability(wb, prof):
    ws = wb.create_sheet("7_갑우_수익성_3개년")
    ws["A1"] = "갑우 거래처별 3개년 수익성 (용지원가 실측 + 간접비 52.9% 배분)"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:R1")
    notes = [
        "• 용지원가: viewGabwoo_마감 → PUR_POH → PRT_WO → PRT_SOH (CD_PARTNER) 체인 실측 (갑우 10000만)",
        "• 간접비: 매출의 52.9% 일괄 (v1 감사보고서 기준, 연도별 변동 미반영 — 한계)",
        "• 지입 거래처(노란색): 용지원가를 0으로 처리. '용지원가_실측' 컬럼은 ERP가 잡은 매입액(참고용)",
        "• 추정마진 = 매출 - 용지원가 - 간접비. 초안(ERP 역산), 확정치 아님",
    ]
    r = 2
    for n in notes:
        c = ws.cell(row=r, column=1, value=n); c.font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=18)
        r += 1
    r += 1

    cols = ["거래처명", "사업자번호", "지입여부",
            "매출_2023", "매출_2024", "매출_2025",
            "용지원가_2023", "용지원가_2024", "용지원가_2025",
            "용지원가율_2025", "간접비_2025",
            "추정마진_2023", "추정마진_2024", "추정마진_2025",
            "추정마진율_2023", "추정마진율_2024", "추정마진율_2025",
            "용지원가_실측_2025"]
    for j, h in enumerate(cols, 1):
        c = ws.cell(row=r, column=j, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
    r += 1

    # 매출 1억+ 대상만
    prof_f = prof[(prof["매출_2023"] >= CUST_THRESHOLD) |
                  (prof["매출_2024"] >= CUST_THRESHOLD) |
                  (prof["매출_2025"] >= CUST_THRESHOLD)].copy()

    for _, row in prof_f.iterrows():
        is_jiip_row = (row["지입여부"] == "지입")
        for j, col in enumerate(cols, 1):
            val = row.get(col, None)
            c = ws.cell(row=r, column=j, value=val)
            c.border = BORDER
            if is_jiip_row: c.fill = JIIP_FILL
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")
                if "율_" in col: c.number_format = "0.0%"
                else: c.number_format = "#,##0"
        r += 1
    autosize(ws, max_width=22)


def write_limits(wb, agg):
    ws = wb.create_sheet("8_한계·출처")
    ws["A1"] = "데이터 한계 및 출처"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:F1")
    rows = [
        ("데이터 소스 (매출)", "제품매출현황 xls = 회계 프로그램의 세금계산서 발행 내역 (3사 시트 분리)"),
        ("데이터 소스 (용지원가)", "SNOTES ERP dbo.viewGabwoo_마감 → PUR_POH → PRT_WO → PRT_SOH (CD_PARTNER) 체인"),
        ("기간", "2023.01 ~ 2025.12 (36개월)"),
        ("거래처 임계", f"3개년 중 1년이라도 매출 1억+ = {len(agg)}개사"),
        ("거래처명 정규화", "(주)/주식회사/공백 차이 통합 + 사업자번호 기반 지입 매칭"),
        ("지입 식별 기준", "사업자번호(JIIP_CUST_BIZ 5건) + 이름 키워드(교원/이투스/에듀윌/동행복권) 백업"),
        ("", ""),
        ("한계 1", "간접비 52.9% 비율은 2025 감사보고서 기준. 23·24년 별도 감사 수치 미반영 → 연도별 비율 변동 감안 필요"),
        ("한계 2", "갑우 용지원가 체인 커버리지: 2025 99.8%, 2023·2024 추가 검증 필요"),
        ("한계 3", "비피(20000)·더원(30000) 용지원가 데이터는 SNOTES에 없음 → 수익성 분석 갑우만 가능"),
        ("한계 4", "지입 거래처의 용지원가 실측값이 0 아닌 경우 = ERP가 갑우 명의로 매입한 흔적. 실제 용지대 부담자 추가 확인 필요"),
        ("한계 5", "수정세금계산서(음수 대변)는 그대로 합산. 거래처별 분해는 별도 분석 필요"),
        ("한계 6", "세금계산서 거래처명 ↔ ERP CD_PARTNER 정규화 매칭 미스로 일부 거래처의 용지원가가 0으로 표시될 수 있음. 예: (주)에듀윌, (주)한성피앤아이 — v1(2025단독) 대비 수치 불일치 7/26건"),
        ("", ""),
        ("분석 제외 거래처", "(주)지에스리테일 V00749 — 프린트뱅크 이관 특수 케이스 (본부장 확인, 2026-04-16). 4·6·7 시트에서 배제, 전체 거래처(시트4) 매출 추이에는 '제외여부' 컬럼으로 표시"),
        ("주의", "수익성 숫자는 '초안(ERP 역산+추정 간접비)'. 영업자·재경 확인 전 확정치 아님"),
    ]
    for i, (a, b) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True) if a else Font()
        ws.cell(row=i, column=2, value=b)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    autosize(ws)


# ─────────────────────────────────────────
# 메인
# ─────────────────────────────────────────
def main():
    print("🔵 v2 경영증거패키지 빌드 시작")
    print("1) 세금계산서 xls 3개년 파싱")
    raw = parse_tax_invoice_xls()
    print(f"   raw rows: {len(raw):,}")
    agg = aggregate_by_customer(raw)
    print(f"   집계 거래처(1억+): {len(agg)}개")

    print("2) SNOTES ERP 접속 → 갑우 용지원가 3개년 체인")
    env = load_env()
    conn = pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=15,
    )
    cost = fetch_gabwoo_cost_3y(conn)
    conn.close()
    print(f"   갑우 용지원가 거래처: {len(cost)}개")

    print("3) 갑우 수익성 3개년 계산")
    prof = build_gabwoo_profitability_3y(agg, cost)
    print(f"   수익성 대상: {len(prof)}개")

    print("4) 엑셀 빌드")
    wb = Workbook(); wb.remove(wb.active)
    write_cover(wb)
    write_executive(wb, agg)
    write_decline_top(wb, agg)
    write_growth_new(wb, agg)
    write_all(wb, agg)
    write_pattern(wb, agg)
    write_jiip(wb, agg)
    write_gabwoo_profitability(wb, prof)
    write_limits(wb, agg)
    wb.save(OUT_XLSX)
    print(f"✅ 완료: {OUT_XLSX}")


if __name__ == "__main__":
    main()
