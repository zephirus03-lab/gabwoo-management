"""
세금계산서 3개년 월별 분석 — 히트맵 + 이탈/신규/발행패턴

입력:
  - /Users/jack/dev/gabwoo/23,24년 제품매출현황(갑우,비피,더원).xls
  - /Users/jack/dev/gabwoo/25년 제품매출현황(갑우,비피,더원).xls

출력 (모두 AX 전환 계획/YYYY-MM-DD/ 하위):
  1) 거래처_월별매출_3개년_YYYYMMDD.xlsx — 히트맵·이탈·신규·발행패턴 종합
  2) customer_monthly_matrix.json — 거래처 × 36개월 매트릭스
  3) customer_lifecycle.json — 이탈/신규/정착 지표
  4) billing_pattern.json — 월합산/건별/이월 분류

원칙:
  - 원본 xls 수정 금지 (read-only)
  - 음수(수정세금계산서)는 별도 집계, 합산 시 순매출로 반영
  - 월계/누계 요약행 제외 (사업자번호 비어있음)
  - "녹음"·"녹취" 용어 금지 (출처는 세금계산서 원본 기준)
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    import numpy as np
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌ pip3 install pandas numpy openpyxl xlrd: {e}")
    sys.exit(1)

# ───────── 경로 ─────────
XLS_23_24 = Path("/Users/jack/dev/gabwoo/23,24년 제품매출현황(갑우,비피,더원).xls")
XLS_25 = Path("/Users/jack/dev/gabwoo/25년 제품매출현황(갑우,비피,더원).xls")

TODAY = datetime.now().strftime("%Y-%m-%d")
OUT_DATE_DIR = Path(f"/Users/jack/dev/gabwoo/AX 전환 계획/{TODAY}")
OUT_DATE_DIR.mkdir(parents=True, exist_ok=True)
OUT_JSON_DIR = Path(__file__).parent / "output"
OUT_JSON_DIR.mkdir(exist_ok=True)

OUT_XLSX = OUT_DATE_DIR / f"거래처_월별매출_3개년_{datetime.now():%Y%m%d}.xlsx"
OUT_MATRIX_JSON = OUT_JSON_DIR / "customer_monthly_matrix.json"
OUT_LIFECYCLE_JSON = OUT_JSON_DIR / "customer_lifecycle.json"
OUT_BILLING_JSON = OUT_JSON_DIR / "billing_pattern.json"

FIRM_NAME = {"10000": "갑우", "20000": "비피", "30000": "더원"}
# 시트명 후보 (25년: 갑우/비피/더원, 23·24년: 갑우23,24 등 다른 이름 사용)
FIRM_SHEET_PATTERNS = {
    "10000": ("갑우", "갑우23,24", "갑우23,24년"),
    "20000": ("비피", "비피23,24", "비피23,24년"),
    "30000": ("더원", "더원23,24", "더원23,24년"),
}

# 이름 정규화 (중복 거래처 병합용) — "(주) 필통북스" = "(주)필통북스" 같은 변형 제거
_SUFFIX_RE = re.compile(r"(\(주\)|\(재\)|\(사\)|\(합\)|주식회사|유한회사|\s+)")

# 패턴 임계치
THRESHOLD_SIZE = 100_000_000   # 1억 이상만 패턴 분류 대상
GROWTH_UP = 20                 # 성장 기준 (%)
DECLINE_DOWN = -20             # 축소 기준 (%)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ───────── 1. 세금계산서 xls 파싱 ─────────
def norm_name(s) -> str:
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return ""
    return _SUFFIX_RE.sub("", str(s).strip()).upper()


def parse_one_xls(xls_path: Path) -> pd.DataFrame:
    """3사(갑우/비피/더원) 시트를 파싱해 통합. 시트명 변형 자동 탐지."""
    frames = []
    xl = pd.ExcelFile(xls_path)
    available = xl.sheet_names
    for firm_code, patterns in FIRM_SHEET_PATTERNS.items():
        sheet = None
        for p in patterns:
            if p in available:
                sheet = p
                break
        if sheet is None:
            # 부분 매칭 (갑우로 시작하는 시트 등)
            prefix = patterns[0]
            for s in available:
                if s.startswith(prefix):
                    sheet = s
                    break
        if sheet is None:
            continue
        try:
            df = pd.read_excel(xls_path, sheet_name=sheet, header=0)
        except ValueError:
            continue
        df.columns = [str(c).strip() for c in df.columns]
        col_map = {}
        for c in df.columns:
            if "날짜" in c: col_map[c] = "날짜"
            elif "적요" in c: col_map[c] = "적요"
            elif "거래처" in c: col_map[c] = "거래처명"
            elif "사업자" in c: col_map[c] = "사업자번호"
            elif "대변" in c: col_map[c] = "대변"
            elif "프로젝트" in c: col_map[c] = "프로젝트"
        df = df.rename(columns=col_map)
        need = {"날짜", "거래처명", "사업자번호", "대변"}
        if not need.issubset(df.columns):
            continue
        if "프로젝트" not in df.columns:
            df["프로젝트"] = None
        if "적요" not in df.columns:
            df["적요"] = None
        df["소속사코드"] = firm_code
        df["소속사"] = FIRM_NAME[firm_code]
        df = df[["소속사코드", "소속사", "날짜", "거래처명", "사업자번호", "대변", "프로젝트", "적요"]]
        # 요약행(사업자번호 비어있음) 제외
        df = df[df["사업자번호"].notna() & (df["사업자번호"].astype(str).str.strip() != "")]
        df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce")
        df["대변"] = pd.to_numeric(df["대변"], errors="coerce").fillna(0)
        df = df[df["날짜"].notna()]
        df["사업자번호"] = (
            df["사업자번호"].astype(str).str.strip()
            .str.replace("-", "", regex=False).str.replace(" ", "", regex=False)
        )
        df = df[df["사업자번호"].str.match(r"^\d{10}$")]
        df["거래처키"] = df["거래처명"].apply(norm_name)
        df = df[df["거래처키"] != ""]
        frames.append(df)
    return pd.concat(frames, ignore_index=True)


def load_all() -> pd.DataFrame:
    print(f"📥 23·24년 파싱: {XLS_23_24.name}")
    a = parse_one_xls(XLS_23_24)
    print(f"   {len(a):,}행")
    print(f"📥 25년 파싱: {XLS_25.name}")
    b = parse_one_xls(XLS_25)
    print(f"   {len(b):,}행")
    all_rows = pd.concat([a, b], ignore_index=True)
    all_rows["연도"] = all_rows["날짜"].dt.year
    all_rows["연월"] = all_rows["날짜"].dt.strftime("%Y-%m")
    # 범위 23/01 ~ 25/12만 사용
    all_rows = all_rows[(all_rows["연도"] >= 2023) & (all_rows["연도"] <= 2025)]
    print(f"   합계 23-25: {len(all_rows):,}행")
    return all_rows


# ───────── 2. 거래처 대표명/대표키 고정 ─────────
def canonical_name_map(rows: pd.DataFrame) -> pd.DataFrame:
    """거래처키 → (대표거래처명, 사업자번호 모음, 소속사)
    같은 거래처키는 가장 많이 쓰인 거래처명을 대표로 사용.
    """
    agg = (
        rows.groupby(["소속사", "거래처키", "거래처명"])
            .size().reset_index(name="cnt")
            .sort_values(["소속사", "거래처키", "cnt"], ascending=[True, True, False])
    )
    canonical = agg.drop_duplicates(["소속사", "거래처키"])[["소속사", "거래처키", "거래처명"]]
    canonical = canonical.rename(columns={"거래처명": "대표거래처명"})
    return canonical


# ───────── 3. 월별 매트릭스 ─────────
MONTHS_36 = [f"{y}-{m:02d}" for y in (2023, 2024, 2025) for m in range(1, 13)]


def build_monthly_matrix(rows: pd.DataFrame, canonical: pd.DataFrame) -> pd.DataFrame:
    """(소속사, 거래처키) × 연월 순매출 합산."""
    agg = rows.groupby(["소속사", "거래처키", "연월"], as_index=False)["대변"].sum()
    pivot = agg.pivot_table(
        index=["소속사", "거래처키"], columns="연월", values="대변", aggfunc="sum"
    ).fillna(0)
    for m in MONTHS_36:
        if m not in pivot.columns:
            pivot[m] = 0
    pivot = pivot[MONTHS_36].reset_index()
    pivot = pivot.merge(canonical, on=["소속사", "거래처키"], how="left")
    # 연도별 합계
    for y in (2023, 2024, 2025):
        cols = [f"{y}-{m:02d}" for m in range(1, 13)]
        pivot[f"{y}_합계"] = pivot[cols].sum(axis=1)
    pivot["3년_합계"] = pivot[[f"{y}_합계" for y in (2023, 2024, 2025)]].sum(axis=1)
    return pivot


# ───────── 4. 패턴 분류 (이탈/신규/성장/축소/유지) ─────────
def classify_pattern(r) -> str:
    s23, s24, s25 = r["2023_합계"], r["2024_합계"], r["2025_합계"]
    # 이탈: 23·24 중 1억+ 있었는데 25는 0원
    prev_max = max(s23, s24)
    if s25 == 0 and prev_max >= THRESHOLD_SIZE:
        return "C. 이탈"
    # 신규: 23 0원, 25에 1억+
    if s23 == 0 and s25 >= THRESHOLD_SIZE:
        return "N. 신규"
    # 규모 미달
    if max(s23, s24, s25) < THRESHOLD_SIZE:
        return "-. 소규모"
    if s23 == 0:
        return "기타"
    change = (s25 - s23) / s23 * 100
    if change <= DECLINE_DOWN:
        return "A. 축소"
    if change >= GROWTH_UP:
        return "G. 성장"
    return "S. 유지"


# ───────── 5. 이탈 시점 분석 ─────────
def analyze_churn(matrix: pd.DataFrame) -> pd.DataFrame:
    """이탈(C) 거래처의 마지막 거래월·이탈 전 월평균·경고신호."""
    churn = matrix[matrix["패턴"] == "C. 이탈"].copy()
    results = []
    for _, r in churn.iterrows():
        monthly = {m: r[m] for m in MONTHS_36}
        positive_months = [(m, v) for m, v in monthly.items() if v > 0]
        if not positive_months:
            continue
        last_month, last_amt = positive_months[-1]
        first_month, _ = positive_months[0]
        # 마지막 거래 이전 6개월 평균
        idx = MONTHS_36.index(last_month)
        pre_window = MONTHS_36[max(0, idx - 5): idx + 1]
        pre_vals = [monthly[m] for m in pre_window if monthly[m] > 0]
        pre_avg = np.mean(pre_vals) if pre_vals else 0
        # 이탈 후 공백 개월수 (마지막 거래 이후 0원 이어진 개월수)
        tail = MONTHS_36[idx + 1:]
        gap_months = len(tail)
        # 경고 신호: 마지막 6개월 중 월 감소 추세 있었는가
        decline_ratio = (last_amt / pre_avg) if pre_avg > 0 else None
        # 이탈 구간: 2024 / 2025 / 조기
        if last_month >= "2025-01":
            timing = "2025년 이탈"
        elif last_month >= "2024-07":
            timing = "2024 하반기 이탈"
        elif last_month >= "2024-01":
            timing = "2024 상반기 이탈"
        else:
            timing = "2023년 이탈"
        results.append({
            "회사": r["소속사"],
            "거래처": r["대표거래처명"],
            "거래처키": r["거래처키"],
            "첫거래월": first_month,
            "마지막거래월": last_month,
            "공백개월": gap_months,
            "이탈시점": timing,
            "이탈전_6개월_월평균(백만)": round(pre_avg / 1e6, 1),
            "마지막거래_금액(백만)": round(last_amt / 1e6, 1),
            "마지막/전6평균": round(decline_ratio, 2) if decline_ratio is not None else None,
            "2023_합계(억)": round(r["2023_합계"] / 1e8, 2),
            "2024_합계(억)": round(r["2024_합계"] / 1e8, 2),
            "2025_합계(억)": 0.0,
            "3년_합계(억)": round(r["3년_합계"] / 1e8, 2),
        })
    df = pd.DataFrame(results)
    if not df.empty:
        df = df.sort_values("3년_합계(억)", ascending=False)
    return df


# ───────── 6. 신규 정착도 ─────────
def analyze_new(matrix: pd.DataFrame) -> pd.DataFrame:
    """신규(N) 거래처의 첫 거래·활동월수·정착점수."""
    new = matrix[matrix["패턴"] == "N. 신규"].copy()
    results = []
    for _, r in new.iterrows():
        monthly = {m: r[m] for m in MONTHS_36}
        positive = [(m, v) for m, v in monthly.items() if v > 0]
        if not positive:
            continue
        first_m, _ = positive[0]
        last_m, last_amt = positive[-1]
        active_months = len(positive)
        total = r["3년_합계"]
        avg_per_active = total / active_months if active_months > 0 else 0
        # 최근 6개월 활동 여부
        last6 = MONTHS_36[-6:]
        recent_active = sum(1 for m in last6 if monthly[m] > 0)
        recent_sum = sum(monthly[m] for m in last6)
        # 시작 시점
        if first_m >= "2025-07":
            onset = "2025 하반기 진입"
        elif first_m >= "2025-01":
            onset = "2025 상반기 진입"
        elif first_m >= "2024-07":
            onset = "2024 하반기 진입"
        else:
            onset = "2024 상반기 진입"
        # 정착도: (최근6개월 활동월수/6) + (최근6개월 매출비중)
        settle_score = (recent_active / 6) * 0.5 + (recent_sum / total if total > 0 else 0) * 0.5
        if settle_score >= 0.55:
            settle_grade = "🟢 정착"
        elif settle_score >= 0.3:
            settle_grade = "🟡 관찰"
        else:
            settle_grade = "🟠 확인"
        results.append({
            "회사": r["소속사"],
            "거래처": r["대표거래처명"],
            "거래처키": r["거래처키"],
            "첫거래월": first_m,
            "마지막거래월": last_m,
            "활동월수": active_months,
            "진입시점": onset,
            "2024(억)": round(r["2024_합계"] / 1e8, 2),
            "2025(억)": round(r["2025_합계"] / 1e8, 2),
            "3년(억)": round(r["3년_합계"] / 1e8, 2),
            "최근6개월_활동월수": recent_active,
            "최근6개월매출(억)": round(recent_sum / 1e8, 2),
            "월평균매출(백만)": round(avg_per_active / 1e6, 1),
            "정착점수": round(settle_score, 2),
            "정착등급": settle_grade,
        })
    df = pd.DataFrame(results)
    if not df.empty:
        df = df.sort_values("3년(억)", ascending=False)
    return df


# ───────── 7. 발행 패턴 (월합산 / 건별 / 이월) ─────────
def analyze_billing_pattern(rows: pd.DataFrame, matrix: pd.DataFrame) -> pd.DataFrame:
    """거래처별 연간 발행건수 / 월당 건수 / 건당 평균 금액으로 패턴 분류."""
    # 양수 행만 발행건수로 집계 (수정세금계산서 음수 제외)
    pos = rows[rows["대변"] > 0].copy()
    pos["연도"] = pos["날짜"].dt.year
    ann = (
        pos.groupby(["소속사", "거래처키", "연도"], as_index=False)
           .agg(연발행건수=("대변", "size"), 연매출=("대변", "sum"),
                활동월수=("연월", "nunique"))
    )
    # 연도 평균 (활동 있는 연도만)
    agg = ann.groupby(["소속사", "거래처키"], as_index=False).agg(
        평균연발행건수=("연발행건수", "mean"),
        평균활동월수=("활동월수", "mean"),
        평균연매출=("연매출", "mean"),
        활동연도수=("연도", "nunique"),
    )
    # 매트릭스 정보 붙이기
    agg = agg.merge(matrix[["소속사", "거래처키", "대표거래처명", "패턴", "3년_합계"]], on=["소속사", "거래처키"], how="left")
    # 1억+ 만 분류 대상
    agg_big = agg[agg["3년_합계"] >= THRESHOLD_SIZE].copy()

    def classify(r):
        n = r["평균연발행건수"]
        m = r["평균활동월수"]
        avg_amt = r["평균연매출"] / n if n > 0 else 0
        # 활동 5개월 이하면 "간헐적/이월"
        if m <= 5:
            return "이월·간헐"
        # 연 발행 20건 미만 + 활동 8개월+ → 월 1~2건 = 월합산
        if n <= 24 and m >= 8:
            return "월합산"
        # 연 발행 100건+ → 건별 집중
        if n >= 100:
            return "건별 집중"
        return "건별 혼합"

    agg_big["발행패턴"] = agg_big.apply(classify, axis=1)
    agg_big["건당평균(백만)"] = agg_big.apply(
        lambda r: round((r["평균연매출"] / r["평균연발행건수"]) / 1e6, 2) if r["평균연발행건수"] > 0 else 0,
        axis=1,
    )
    agg_big["평균연발행건수"] = agg_big["평균연발행건수"].round(1)
    agg_big["평균활동월수"] = agg_big["평균활동월수"].round(1)
    agg_big["3년_합계(억)"] = (agg_big["3년_합계"] / 1e8).round(2)
    agg_big = agg_big[[
        "소속사", "대표거래처명", "거래처키", "패턴", "발행패턴",
        "평균연발행건수", "평균활동월수", "건당평균(백만)",
        "활동연도수", "3년_합계(억)",
    ]].rename(columns={"소속사": "회사", "대표거래처명": "거래처"})
    agg_big = agg_big.sort_values("3년_합계(억)", ascending=False)
    return agg_big


# ───────── 8. xlsx 작성 ─────────
def autosize(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            v = "" if cell.value is None else str(cell.value)
            ln = sum(2 if ord(c) > 127 else 1 for c in v)
            if ln > max_len:
                max_len = ln
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def write_df(ws, df: pd.DataFrame, title: str, notes=None, money_cols_billion=None):
    ws["A1"] = title
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    ncols = max(len(df.columns), 5)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    r = 2
    if notes:
        for n in (notes if isinstance(notes, list) else [notes]):
            ws.cell(row=r, column=1, value=n).font = Font(italic=True, size=10, color="555555")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            r += 1
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=r, column=j, value=str(col))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(df.itertuples(index=False), start=r + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = BORDER
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                c.alignment = Alignment(horizontal="right")
                if isinstance(val, float):
                    c.number_format = "#,##0.00"
                else:
                    c.number_format = "#,##0"
    autosize(ws)


def write_heatmap(ws, matrix: pd.DataFrame):
    """거래처 × 36개월 히트맵 (3년_합계 큰 순, 1억+ 만)."""
    big = matrix[matrix["3년_합계"] >= THRESHOLD_SIZE].sort_values("3년_합계", ascending=False).copy()
    ws["A1"] = "거래처 × 36개월 월별 매출 히트맵 (단위: 백만원, 3년 누적 1억 이상)"
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=40)
    ws["A2"] = "회색=매출없음 / 진할수록 높음 (거래처 내부 상대값, 거래처별로 컬러 스케일 독립)"
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws.merge_cells("A2:AR2")

    header = ["회사", "거래처", "패턴", "3년합(억)"] + MONTHS_36
    for j, h in enumerate(header, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    # 내용
    for i, row in enumerate(big.itertuples(index=False), start=4):
        r_dict = dict(zip(big.columns, row))
        ws.cell(row=i, column=1, value=r_dict["소속사"]).border = BORDER
        ws.cell(row=i, column=2, value=r_dict["대표거래처명"]).border = BORDER
        ws.cell(row=i, column=3, value=r_dict.get("패턴", "")).border = BORDER
        total_cell = ws.cell(row=i, column=4, value=round(r_dict["3년_합계"] / 1e8, 2))
        total_cell.border = BORDER
        total_cell.number_format = "#,##0.00"
        # 거래처 내부 최대값 (컬러 스케일)
        cust_max = max(r_dict[m] for m in MONTHS_36) if any(r_dict[m] for m in MONTHS_36) else 0
        for k, m in enumerate(MONTHS_36):
            v = r_dict[m]
            display_v = round(v / 1e6) if v > 0 else (round(v / 1e6, 1) if v < 0 else "")
            c = ws.cell(row=i, column=5 + k, value=display_v)
            c.border = BORDER
            c.alignment = Alignment(horizontal="right")
            if isinstance(display_v, (int, float)) and display_v != "":
                c.number_format = "#,##0"
            if v == 0:
                c.fill = PatternFill("solid", fgColor="F2F2F2")
            elif v < 0:
                c.fill = PatternFill("solid", fgColor="F4CCCC")
            else:
                # 0~1 상대값 → 파랑 농도
                ratio = v / cust_max if cust_max > 0 else 0
                grade = max(0, min(255, int(255 - ratio * 170)))
                color = f"{grade:02X}{grade:02X}FF"
                c.fill = PatternFill("solid", fgColor=color)
    # 열 너비: 거래처명 넓게, 월은 좁게
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    for k in range(len(MONTHS_36)):
        ws.column_dimensions[get_column_letter(5 + k)].width = 8
    ws.freeze_panes = "E4"


def write_cover(ws, counts: dict, sums: dict):
    ws["A1"] = "세금계산서 3개년 월별 거동 분석 (갑우·비피·더원)"
    ws["A1"].font = Font(bold=True, size=16, color="1F3864")
    ws.merge_cells("A1:F1")
    ws["A2"] = f"작성일: {datetime.now():%Y-%m-%d}  /  출처: 제품매출현황 xls 2개 파일(23·24년 + 25년)"
    ws["A2"].font = Font(size=10, color="555555")
    ws.merge_cells("A2:F2")

    ws["A4"] = "[이 보고서가 답하는 5가지 질문]"
    ws["A4"].font = SECTION_FONT
    ws["A4"].fill = SECTION_FILL
    ws.merge_cells("A4:F4")
    q = [
        "① 거래처 매출이 언제 빠지기 시작했는가? (36개월 히트맵)",
        "② 이탈한 거래처는 언제·어떻게 떠났는가? (마지막 거래월 + 경고신호)",
        "③ 신규로 잡힌 거래처는 정착하고 있는가? (진입 시점 + 정착점수)",
        "④ 거래처는 어떻게 세금계산서를 받아가는가? (월합산·건별·이월 분류)",
        "⑤ 매출 -49% 붕괴는 몇 개 거래처에 얼마나 몰려 있는가?",
    ]
    for i, line in enumerate(q, start=5):
        ws[f"A{i}"] = line
        ws.merge_cells(f"A{i}:F{i}")

    ws["A12"] = "[패턴 요약 — 1억 이상 거래처 기준]"
    ws["A12"].font = SECTION_FONT
    ws["A12"].fill = SECTION_FILL
    ws.merge_cells("A12:F12")
    headers = ["패턴", "거래처수", "2023 합(억)", "2024 합(억)", "2025 합(억)", "23→25 변화(억)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=13, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    order = ["A. 축소", "C. 이탈", "G. 성장", "N. 신규", "S. 유지"]
    for i, p in enumerate(order, start=14):
        n = counts.get(p, 0)
        s23 = sums.get((p, 2023), 0) / 1e8
        s24 = sums.get((p, 2024), 0) / 1e8
        s25 = sums.get((p, 2025), 0) / 1e8
        ws.cell(row=i, column=1, value=p).border = BORDER
        ws.cell(row=i, column=2, value=n).border = BORDER
        ws.cell(row=i, column=3, value=round(s23, 1)).border = BORDER
        ws.cell(row=i, column=4, value=round(s24, 1)).border = BORDER
        ws.cell(row=i, column=5, value=round(s25, 1)).border = BORDER
        ws.cell(row=i, column=6, value=round(s25 - s23, 1)).border = BORDER
        for j in (2, 3, 4, 5, 6):
            ws.cell(row=i, column=j).alignment = Alignment(horizontal="right")
            ws.cell(row=i, column=j).number_format = "#,##0.00"

    ws["A21"] = "[시트 구성]"
    ws["A21"].font = SECTION_FONT
    ws["A21"].fill = SECTION_FILL
    ws.merge_cells("A21:F21")
    sheets = [
        ("1_월별_히트맵", "1억+ 거래처 × 36개월 매출 (상대값 색조)"),
        ("2_이탈_시점분석", "25년에 사라진 거래처의 마지막 거래월·경고신호"),
        ("3_신규_정착도", "24·25년에 처음 잡힌 거래처의 정착 여부"),
        ("4_발행패턴", "월합산 / 건별 / 이월 분류 (교원 사급 구조 후보 탐지)"),
        ("5_패턴별_상세", "축소·이탈·성장·신규·유지 거래처 전체 3개년 합계"),
    ]
    for i, (s, d) in enumerate(sheets, start=22):
        ws[f"A{i}"] = s
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = d
        ws.merge_cells(f"B{i}:F{i}")

    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 18


def write_pattern_detail(ws, matrix: pd.DataFrame):
    df = matrix.copy()
    df["2023(억)"] = (df["2023_합계"] / 1e8).round(2)
    df["2024(억)"] = (df["2024_합계"] / 1e8).round(2)
    df["2025(억)"] = (df["2025_합계"] / 1e8).round(2)
    df["3년(억)"] = (df["3년_합계"] / 1e8).round(2)
    df["23→25변화(억)"] = (df["2025(억)"] - df["2023(억)"]).round(2)
    out = df[["소속사", "대표거래처명", "패턴", "2023(억)", "2024(억)", "2025(억)", "3년(억)", "23→25변화(억)"]]
    out = out.rename(columns={"소속사": "회사", "대표거래처명": "거래처"})
    out = out.sort_values(["패턴", "3년(억)"], ascending=[True, False])
    write_df(
        ws,
        out,
        "패턴별 상세 — 전체 거래처 3년 합계",
        "패턴: A.축소 / C.이탈 / G.성장 / N.신규 / S.유지 / -.소규모 / 기타 (1억 미만은 '-.소규모')",
    )


def main():
    rows = load_all()
    canonical = canonical_name_map(rows)

    matrix = build_monthly_matrix(rows, canonical)
    matrix["패턴"] = matrix.apply(classify_pattern, axis=1)
    print(f"   거래처(정규화 후): {len(matrix):,}개")

    # 패턴 카운트
    counts = matrix["패턴"].value_counts().to_dict()
    sums = {}
    for p in matrix["패턴"].unique():
        sub = matrix[matrix["패턴"] == p]
        for y in (2023, 2024, 2025):
            sums[(p, y)] = sub[f"{y}_합계"].sum()
    print("\n[패턴 집계]")
    for p in ["A. 축소", "C. 이탈", "G. 성장", "N. 신규", "S. 유지", "-. 소규모"]:
        n = counts.get(p, 0)
        if n == 0:
            continue
        s25 = sums.get((p, 2025), 0) / 1e8
        s23 = sums.get((p, 2023), 0) / 1e8
        print(f"  {p}: {n}개, 23={s23:.1f}억 → 25={s25:.1f}억 (Δ{s25-s23:+.1f})")

    # 이탈 분석
    churn = analyze_churn(matrix)
    print(f"\n[이탈] {len(churn)}개사")
    if not churn.empty:
        print(churn.groupby("이탈시점").size().to_dict())

    # 신규 정착
    new = analyze_new(matrix)
    print(f"\n[신규] {len(new)}개사")
    if not new.empty:
        print(new.groupby("정착등급").size().to_dict())

    # 발행 패턴
    billing = analyze_billing_pattern(rows, matrix)
    print(f"\n[발행 패턴] {len(billing)}개사")
    if not billing.empty:
        print(billing.groupby("발행패턴").size().to_dict())

    # xlsx 작성
    wb = Workbook()
    wb.active.title = "0_표지"
    write_cover(wb["0_표지"], counts, sums)

    ws1 = wb.create_sheet("1_월별_히트맵")
    write_heatmap(ws1, matrix)

    ws2 = wb.create_sheet("2_이탈_시점분석")
    if churn.empty:
        ws2["A1"] = "이탈 거래처 없음"
    else:
        write_df(ws2, churn, "이탈 거래처 시점 분석 (25년 0원 + 23·24 합이 1억 이상)")

    ws3 = wb.create_sheet("3_신규_정착도")
    if new.empty:
        ws3["A1"] = "신규 거래처 없음"
    else:
        write_df(ws3, new, "신규 거래처 정착도 (23년 0원 + 25년 1억 이상)")

    ws4 = wb.create_sheet("4_발행패턴")
    if billing.empty:
        ws4["A1"] = "데이터 없음"
    else:
        write_df(
            ws4,
            billing,
            "거래처별 세금계산서 발행 패턴 (1억+ 거래처)",
            [
                "월합산: 연발행건수 ≤ 24 + 활동월수 ≥ 8 (한 달에 1~2건 몰아 발행)",
                "건별 집중: 연발행건수 ≥ 100 (건건이 끊어 발행)",
                "이월·간헐: 활동월수 ≤ 5 (특정 기간만 거래)",
                "건별 혼합: 위 3개에 속하지 않는 일반 패턴",
            ],
        )

    ws5 = wb.create_sheet("5_패턴별_상세")
    write_pattern_detail(ws5, matrix)

    wb.save(OUT_XLSX)
    print(f"\n✅ 저장: {OUT_XLSX}")

    # JSON 저장
    matrix_out = []
    for _, r in matrix.iterrows():
        if r["3년_합계"] < THRESHOLD_SIZE:
            continue
        monthly = {m: int(r[m]) for m in MONTHS_36}
        matrix_out.append({
            "회사": r["소속사"],
            "거래처": r["대표거래처명"],
            "거래처키": r["거래처키"],
            "패턴": r["패턴"],
            "2023_합계": int(r["2023_합계"]),
            "2024_합계": int(r["2024_합계"]),
            "2025_합계": int(r["2025_합계"]),
            "3년_합계": int(r["3년_합계"]),
            "월별": monthly,
        })
    OUT_MATRIX_JSON.write_text(json.dumps(matrix_out, ensure_ascii=False, indent=2))
    print(f"✅ JSON: {OUT_MATRIX_JSON}")

    lifecycle = {
        "churn": churn.to_dict(orient="records") if not churn.empty else [],
        "new": new.to_dict(orient="records") if not new.empty else [],
    }
    OUT_LIFECYCLE_JSON.write_text(json.dumps(lifecycle, ensure_ascii=False, indent=2, default=str))
    print(f"✅ JSON: {OUT_LIFECYCLE_JSON}")

    OUT_BILLING_JSON.write_text(
        json.dumps(billing.to_dict(orient="records") if not billing.empty else [], ensure_ascii=False, indent=2, default=str)
    )
    print(f"✅ JSON: {OUT_BILLING_JSON}")


if __name__ == "__main__":
    main()
