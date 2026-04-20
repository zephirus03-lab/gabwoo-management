"""
경영진단 잠정 보고서 v2 — 3사 분리 + 지입거래처 구분

[3사 분리] CD_CUST_OWN 코드로 구분
  - 10000 = 갑우문화사 (출판 본체)
  - 20000 = 비피앤피 (패키지)
  - 30000 = 더원프린팅 (제본)

[지입거래처] 거래처가 용지를 직접 구매·지급하고 제작만 맡김
  → 매출에 용지비 미포함 → 다른 거래처와 단가/원가율 직접 비교 불가
  - V00661 (주)교원구몬
  - V00712 (주)교원
  - V01222 이투스에듀 서초지점
  - V1526  (주)에듀윌
  - V01119 (주)교원프라퍼티 인천공장
  - 그 외 '교원/이투스/에듀윌' 키워드 포함 거래처

출력: 잠정_경영진단보고서_v2_YYYYMMDD.xlsx
"""
import sys
from pathlib import Path
from datetime import datetime

try:
    import pymssql
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"pip3 install pymssql pandas openpyxl: {e}"); sys.exit(1)

ENV_FILE = Path("/Users/jack/dev/gabwoo/견적계산기/.env.local")
OUT_DIR = Path(__file__).parent / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_XLSX = OUT_DIR / f"잠정_경영진단보고서_v2_{datetime.now():%Y%m%d}.xlsx"

FIRM_MAP = {"10000": "갑우", "20000": "비피", "30000": "더원"}
FIRM_NAMES = {"갑우": "갑우문화사", "비피": "비피앤피", "더원": "더원프린팅"}

# 지입 거래처 키워드 (거래처명 부분일치)
지입_KEYWORDS = ["교원", "이투스", "에듀윌"]

# 용지 원가 상승률 (2023→2025, 가중평균)
COST_INFLATION_23_25 = 22.6

# 스타일
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
SUBSECTION_FILL = PatternFill("solid", fgColor="BDD7EE")
BORDER = Border(left=Side(style="thin", color="BFBFBF"),
                right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="thin", color="BFBFBF"))


def load_env():
    env = {}
    for line in ENV_FILE.read_text().splitlines():
        if "=" in line and not line.strip().startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def get_conn(env):
    return pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=15,
    )


def is_지입(name):
    if not name:
        return False
    return any(kw in name for kw in 지입_KEYWORDS)


# ─────────────────────────────────────────────────────────────────────
# 데이터 수집
# ─────────────────────────────────────────────────────────────────────

def fetch_all(conn):
    print("📥 SAL_SALESH 헤더 조회...")
    header = pd.read_sql("""
        SELECT
            LEFT(h.DT_SALES, 4) AS 연도,
            h.CD_CUST_OWN AS 소속코드,
            h.CD_CUST AS 거래처코드,
            c.NM_CUST AS 거래처명,
            COUNT(DISTINCT h.NO_SALES) AS 매출건수,
            SUM(h.AM) AS 매출
        FROM SAL_SALESH h
        LEFT JOIN MAS_CUST c ON h.CD_CUST=c.CD_CUST AND h.CD_FIRM=c.CD_FIRM
        WHERE h.CD_FIRM='7000'
          AND h.DT_SALES >= '20210101' AND h.DT_SALES <= '20261231'
          AND h.AM > 0 AND (h.ST_SALES='Y' OR h.ST_SALES IS NULL)
          AND h.CD_CUST_OWN IN ('10000','20000','30000')
        GROUP BY LEFT(h.DT_SALES, 4), h.CD_CUST_OWN, h.CD_CUST, c.NM_CUST
    """, conn)

    print("📥 SAL_SALESL 라인 집계...")
    line = pd.read_sql("""
        SELECT
            LEFT(h.DT_SALES, 4) AS 연도,
            h.CD_CUST_OWN AS 소속코드,
            h.CD_CUST AS 거래처코드,
            SUM(l.QT) AS 수량,
            SUM(l.AM) AS 라인매출
        FROM SAL_SALESH h
        JOIN SAL_SALESL l ON h.CD_FIRM=l.CD_FIRM AND h.NO_SALES=l.NO_SALES
        WHERE h.CD_FIRM='7000'
          AND h.DT_SALES >= '20210101' AND h.DT_SALES <= '20261231'
          AND (h.ST_SALES='Y' OR h.ST_SALES IS NULL)
          AND l.QT > 0 AND l.AM > 0
          AND h.CD_CUST_OWN IN ('10000','20000','30000')
        GROUP BY LEFT(h.DT_SALES, 4), h.CD_CUST_OWN, h.CD_CUST
    """, conn)

    print("📥 용지 매입 월별...")
    paper_monthly = pd.read_sql("""
        SELECT FORMAT(일자, 'yyyy') AS 연도,
               SUM(공급가액) AS 용지매입,
               SUM(수량) AS 용지수량
        FROM dbo.viewGabwoo_마감
        WHERE 일자 >= '2021-01-01' AND 일자 <= '2026-12-31'
        GROUP BY FORMAT(일자, 'yyyy')
    """, conn)

    print("📥 외주 매입 연도별...")
    outsource_yearly = pd.read_sql("""
        SELECT LEFT(DT_PUR, 4) AS 연도,
               SUM(AM_SUPPLY) AS 외주매입,
               CD_CUST_OWN AS 소속코드
        FROM PUR_ETCCLSH
        WHERE CD_FIRM='7000'
          AND DT_PUR >= '20210101' AND DT_PUR <= '20261231'
          AND AM_SUPPLY > 0
        GROUP BY LEFT(DT_PUR, 4), CD_CUST_OWN
    """, conn)

    print("📥 용지 단가 연도별 가중평균...")
    paper_yearly = pd.read_sql("""
        SELECT YEAR(일자) AS 연도,
               SUM(수량) AS 수량,
               SUM(공급가액) AS 매입액,
               CASE WHEN SUM(수량)>0 THEN SUM(공급가액)*1.0/SUM(수량) ELSE 0 END AS 단가
        FROM dbo.viewGabwoo_마감
        WHERE 일자 >= '2021-01-01' AND 일자 <= '2026-12-31' AND 수량 > 0
        GROUP BY YEAR(일자)
        ORDER BY YEAR(일자)
    """, conn)

    return header, line, paper_monthly, outsource_yearly, paper_yearly


# ─────────────────────────────────────────────────────────────────────
# 가공
# ─────────────────────────────────────────────────────────────────────

def enrich(header, line):
    """헤더·라인 merge + 지입·소속 라벨 추가."""
    merged = header.merge(line, on=["연도", "소속코드", "거래처코드"], how="left")
    merged["수량"] = merged["수량"].fillna(0)
    merged["라인매출"] = merged["라인매출"].fillna(0)
    merged["단가"] = merged.apply(
        lambda r: r["라인매출"]/r["수량"] if r["수량"]>0 else 0, axis=1
    )
    merged["소속사"] = merged["소속코드"].map(FIRM_MAP)
    merged["지입여부"] = merged["거래처명"].apply(lambda n: "지입" if is_지입(n) else "일반")
    return merged


def firm_yearly_summary(merged, paper_yearly, outsource_yearly):
    """3사 × 연도별 매출·원가율 요약."""
    rows = []
    years = sorted(merged["연도"].unique())
    paper_by_year = paper_yearly.set_index("연도")["매입액"].to_dict() if "연도" in paper_yearly.columns else {}
    # 외주는 소속별 없음(갑우 본체만으로 간주) → 전체로 취급
    out_by_year = outsource_yearly.groupby("연도")["외주매입"].sum().to_dict()

    for year in years:
        year_int = int(year)
        for firm in ["갑우", "비피", "더원"]:
            sub = merged[(merged["연도"]==year) & (merged["소속사"]==firm)]
            매출 = sub["매출"].sum()
            if 매출 == 0:
                continue
            rows.append({
                "연도": year,
                "소속사": firm,
                "매출(억)": round(매출/1e8, 1),
                "거래처수": sub["거래처코드"].nunique(),
                "매출건수": sub["매출건수"].sum(),
            })
    summary = pd.DataFrame(rows)
    # 연도별 합계 row 추가
    totals = summary.groupby("연도").agg(
        매출억=("매출(억)", "sum"),
        거래처수=("거래처수", "sum"),
        매출건수=("매출건수", "sum"),
    ).reset_index()
    # paper/외주 붙여서 원가율(그룹 전체) 행 생성
    group_rows = []
    for _, r in totals.iterrows():
        y = r["연도"]
        paper = paper_by_year.get(int(y), 0)
        out = out_by_year.get(y, 0)
        매출액 = r["매출억"] * 1e8
        총매입 = paper + out
        원가율 = (총매입/매출액*100) if 매출액>0 else 0
        group_rows.append({
            "연도": y,
            "그룹매출(억)": round(매출액/1e8, 1),
            "용지매입(억)": round(paper/1e8, 1),
            "외주매입(억)": round(out/1e8, 1),
            "총매입(억)": round(총매입/1e8, 1),
            "원가율(%)": round(원가율, 1),
        })
    cost_ratio = pd.DataFrame(group_rows)
    return summary, cost_ratio


def customer_yoy(merged, firm, 지입필터="일반"):
    """특정 회사 × 지입여부 거래처별 2023→2025 YoY."""
    sub = merged[(merged["소속사"]==firm) & (merged["지입여부"]==지입필터)].copy()
    if sub.empty:
        return pd.DataFrame()
    amt = sub.pivot_table(index=["거래처코드","거래처명"], columns="연도", values="매출", aggfunc="sum").fillna(0)
    qty = sub.pivot_table(index=["거래처코드","거래처명"], columns="연도", values="수량", aggfunc="sum").fillna(0)
    um  = sub.pivot_table(index=["거래처코드","거래처명"], columns="연도", values="단가", aggfunc="mean").fillna(0)
    r = pd.DataFrame(index=amt.index)
    r["매출_2023"] = amt.get("2023", 0)
    r["매출_2025"] = amt.get("2025", 0)
    r["매출Δ(%)"] = (r["매출_2025"]-r["매출_2023"])/r["매출_2023"].replace(0, pd.NA)*100
    r["수량_2023"] = qty.get("2023", 0)
    r["수량_2025"] = qty.get("2025", 0)
    r["수량Δ(%)"] = (r["수량_2025"]-r["수량_2023"])/r["수량_2023"].replace(0, pd.NA)*100
    r["단가_2023"] = um.get("2023", 0)
    r["단가_2025"] = um.get("2025", 0)
    r["단가Δ(%)"] = (r["단가_2025"]-r["단가_2023"])/r["단가_2023"].replace(0, pd.NA)*100
    r = r.reset_index()
    # 의미있는 규모 (2023 매출 5천만 이상)
    return r[r["매출_2023"] >= 50_000_000].copy()


def classify_pattern(row):
    if row["매출_2023"] == 0:
        return "N/A"
    if row["매출_2025"] == 0:
        return "C. 완전이탈"
    try:
        d_amt = float(row["매출Δ(%)"])
        d_um = float(row["단가Δ(%)"])
    except (ValueError, TypeError):
        return "기타"
    if pd.isna(d_amt) or pd.isna(d_um):
        return "기타"
    if d_amt <= -20 and d_um <= -10:
        return "A. 단가↓물량↓"
    if d_amt <= -20 and d_um >= 5:
        return "B. 단가↑이탈"
    if d_amt <= -20 and -10 < d_um < 5:
        return "D. 물량감소·단가유지"
    if d_amt > -20 and d_amt < 20:
        return "E. 안정"
    if d_amt >= 20:
        return "F. 성장"
    return "기타"


def classify_cost_pass(row):
    """원가 +22.6% 대비 단가 인상률 판정."""
    if row["매출_2023"] == 0 or row["매출_2025"] == 0:
        return "N/A"
    try:
        d = float(row["단가Δ(%)"])
    except (ValueError, TypeError):
        return "N/A"
    if pd.isna(d):
        return "N/A"
    if d >= COST_INFLATION_23_25:
        return "✅ 반영충분"
    if d >= 0:
        return "🟡 일부반영"
    return "🔴 역행(단가↓)"


def fmt_display(df):
    d = df.copy()
    if len(d) == 0:
        return d
    d["매출23(억)"] = (d["매출_2023"]/1e8).round(2)
    d["매출25(억)"] = (d["매출_2025"]/1e8).round(2)
    for c in ["매출Δ(%)","수량Δ(%)","단가Δ(%)"]:
        d[c] = pd.to_numeric(d[c], errors="coerce").round(1)
    cols = ["거래처코드","거래처명","매출23(억)","매출25(억)","매출Δ(%)","수량Δ(%)","단가Δ(%)"]
    if "패턴" in d.columns: cols.append("패턴")
    if "원가반영" in d.columns: cols.append("원가반영")
    return d[cols]


# ─────────────────────────────────────────────────────────────────────
# Excel 작성
# ─────────────────────────────────────────────────────────────────────

def autosize(ws, max_width=55):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            try:
                v = str(cell.value) if cell.value is not None else ""
                ln = sum(2 if ord(c)>127 else 1 for c in v)
                if ln > max_len: max_len = ln
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def write_df_sheet(ws, df, title, notes=None):
    ws["A1"] = title
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    ncols = max(len(df.columns), 5) if len(df) > 0 else 5
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    r = 2
    if notes:
        for note in (notes if isinstance(notes, list) else [notes]):
            ws.cell(row=r, column=1, value=note).font = Font(italic=True, size=10, color="555555")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            r += 1
    # header
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=r, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    # body
    for i, row in enumerate(df.itertuples(index=False), start=r+1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = BORDER
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")
                if abs(val) >= 1000:
                    c.number_format = "#,##0"
                elif isinstance(val, float):
                    c.number_format = "#,##0.00"
    autosize(ws)


def write_cover(ws):
    ws["A1"] = "갑우문화사 경영진단 — 잠정 보고서 v2 (3사 분리 + 지입거래처 구분)"
    ws["A1"].font = Font(bold=True, size=16, color="1F3864")
    ws.merge_cells("A1:F1")
    ws["A2"] = f"작성일: {datetime.now():%Y-%m-%d}  /  분석범위: 2021~2025 + 2026Q1  /  출처: SNOTES ERP 직접 조회 (SELECT only)"
    ws["A2"].font = Font(size=10, color="555555")
    ws.merge_cells("A2:F2")

    ws["A4"] = "[v2 핵심 변경점]"
    ws["A4"].font = SECTION_FONT
    ws["A4"].fill = SECTION_FILL
    ws.merge_cells("A4:F4")
    changes = [
        ("1", "CD_CUST_OWN 필드로 3사 분리: 10000=갑우 / 20000=비피앤피 / 30000=더원. 기존 보고서는 3사 혼재였음"),
        ("2", "지입거래처 별도 분류: 교원·이투스·에듀윌 계열 (거래처가 용지 직접 구매, 우리는 제작만). 단가 비교에서 제외 또는 별도 집계"),
        ("3", "원가율 계산 주의: 매입(용지+외주)은 CD_CUST_OWN 필드가 없어 그룹 전체 기준. 매출은 회사별로 분리하되 원가율은 그룹 단위"),
    ]
    for i, (n, t) in enumerate(changes, start=5):
        ws[f"A{i}"] = n
        ws[f"B{i}"] = t
        ws.merge_cells(f"B{i}:F{i}")

    ws["A9"] = "[시트 구성]"
    ws["A9"].font = SECTION_FONT
    ws["A9"].fill = SECTION_FILL
    ws.merge_cells("A9:F9")
    sheets = [
        ("1_3사_연도별매출", "소속사별 매출 추이 + 그룹 원가율"),
        ("2_갑우_일반거래처", "갑우 본체 일반 거래처 YoY + 패턴 + 원가반영"),
        ("3_갑우_지입거래처", "갑우 지입(교원/이투스/에듀윌) — 단가 비교 분리 해석"),
        ("4_비피_일반거래처", "비피앤피 일반 거래처 YoY + 패턴 + 원가반영"),
        ("5_비피_지입거래처", "비피앤피 지입 거래처 (교원구몬 등)"),
        ("6_더원_전체", "더원프린팅 (소규모, 2023부터 데이터 있음)"),
        ("7_용지단가_추이", "용지 매입 단가 YoY (전체 공통)"),
        ("8_후속제안", "다음 액션"),
    ]
    for i, (s, d) in enumerate(sheets, start=10):
        ws[f"A{i}"] = s
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = d
        ws.merge_cells(f"B{i}:F{i}")

    ws["A19"] = "[지입거래처 — 단가 비교에서 왜 분리하는가]"
    ws["A19"].font = SECTION_FONT
    ws["A19"].fill = SECTION_FILL
    ws.merge_cells("A19:F19")
    explain = [
        "지입(紙入)거래처: 거래처가 용지를 직접 구매·지급하고 갑우에는 제작만 맡김",
        "→ 매출에 용지비가 포함되지 않음. 일반 거래처와 같은 '단가' 기준으로 비교하면 왜곡",
        "→ 용지 원가 상승(+70%)의 영향도 거의 받지 않음 (용지비 부담이 거래처 쪽)",
        "→ 단가 인하·인상 판정은 '일반 거래처' 기준으로만 해석해야 함",
        "해당: 교원구몬·교원·교원프라퍼티·이투스에듀·에듀윌 등",
    ]
    for i, t in enumerate(explain, start=20):
        ws[f"A{i}"] = "•"
        ws[f"B{i}"] = t
        ws.merge_cells(f"B{i}:F{i}")

    ws.column_dimensions["A"].width = 4
    for col in "BCDEF":
        ws.column_dimensions[col].width = 22


def write_firm_sheet(wb, sheet_name, firm, 지입필터, merged, title, note):
    ws = wb.create_sheet(sheet_name)
    df = customer_yoy(merged, firm, 지입필터)
    if df.empty:
        ws["A1"] = f"{title} — 데이터 없음"
        return
    df["패턴"] = df.apply(classify_pattern, axis=1)
    df["원가반영"] = df.apply(classify_cost_pass, axis=1)
    df_sorted = df.sort_values("매출_2023", ascending=False)
    display = fmt_display(df_sorted)

    # 요약
    summary = df.groupby("패턴").agg(
        거래처수=("거래처코드","count"),
        매출23억=("매출_2023", lambda x: round(x.sum()/1e8, 1)),
        매출25억=("매출_2025", lambda x: round(x.sum()/1e8, 1)),
    ).reset_index()
    summary["손실(억)"] = summary["매출23억"] - summary["매출25억"]

    cost_summary = df.groupby("원가반영").agg(
        거래처수=("거래처코드","count"),
        매출23억=("매출_2023", lambda x: round(x.sum()/1e8, 1)),
        매출25억=("매출_2025", lambda x: round(x.sum()/1e8, 1)),
    ).reset_index()
    cost_summary["손익(억)"] = cost_summary["매출25억"] - cost_summary["매출23억"]

    # 쓰기: 상단 요약 2개 + 아래 상세 테이블
    ws["A1"] = title
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:I1")
    ws["A2"] = note
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws.merge_cells("A2:I2")

    # 요약1 — 패턴
    r = 4
    ws.cell(row=r, column=1, value="[거래처 패턴 요약]").font = Font(bold=True, size=11)
    r += 1
    for j, col in enumerate(summary.columns, start=1):
        c = ws.cell(row=r, column=j, value=col)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    for _, row in summary.iterrows():
        r += 1
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = BORDER
            if isinstance(val, (int,float)):
                c.alignment = Alignment(horizontal="right")

    # 요약2 — 원가반영
    r += 2
    ws.cell(row=r, column=1, value="[원가 반영 판정 (벤치마크: 용지단가 +22.6%)]").font = Font(bold=True, size=11)
    r += 1
    for j, col in enumerate(cost_summary.columns, start=1):
        c = ws.cell(row=r, column=j, value=col)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    for _, row in cost_summary.iterrows():
        r += 1
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = BORDER
            if isinstance(val, (int,float)):
                c.alignment = Alignment(horizontal="right")

    # 상세 테이블
    r += 2
    ws.cell(row=r, column=1, value="[거래처별 상세 — 2023 매출 5천만 이상, 2023 매출 큰 순]").font = Font(bold=True, size=11)
    r += 1
    for j, col in enumerate(display.columns, start=1):
        c = ws.cell(row=r, column=j, value=col)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    for _, row in display.iterrows():
        r += 1
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = BORDER
            if isinstance(val, (int,float)):
                c.alignment = Alignment(horizontal="right")
                if abs(val) >= 1000: c.number_format = "#,##0"
                elif isinstance(val, float): c.number_format = "#,##0.00"

    autosize(ws)


def main():
    env = load_env()
    conn = get_conn(env)
    print(f"✅ ERP 연결 ({env['ERP_HOST']})")

    header, line, paper_monthly, outsource_yearly, paper_yearly = fetch_all(conn)
    conn.close()

    merged = enrich(header, line)
    print(f"   → 총 {len(merged):,}행, 소속사 분포:")
    for firm, n in merged["소속사"].value_counts().items():
        print(f"     {firm}: {n:,}행")

    # Excel
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "0_표지"
    write_cover(ws0)

    # 1. 3사 연도별
    print("📝 시트 1: 3사 연도별 매출·원가율")
    ws1 = wb.create_sheet("1_3사_연도별매출")
    summary, cost_ratio = firm_yearly_summary(merged, paper_yearly, outsource_yearly)
    # pivot
    pivot = summary.pivot_table(index="연도", columns="소속사", values="매출(억)", aggfunc="sum").fillna(0)
    pivot["합계(억)"] = pivot.sum(axis=1)
    pivot = pivot.reset_index()

    write_df_sheet(ws1, pivot,
        "1-A. 소속사별 연도별 매출 (억)",
        "CD_CUST_OWN: 10000=갑우 / 20000=비피 / 30000=더원. 합계 = 그룹 전체 매출")
    # 이어서 원가율
    ws1b = wb.create_sheet("1-B_그룹_원가율")
    write_df_sheet(ws1b, cost_ratio,
        "1-B. 그룹 전체 원가율 추이 (용지 + 외주)",
        ["매입 데이터(용지·외주)는 CD_CUST_OWN 컬럼이 없어 회사별 분리 불가 → 그룹 합산 기준",
         "2021 14% → 2023 10%(호황) → 2025 18.5% → 2026Q1 25%로 급등"])

    # 2~6. 회사별 × 지입/일반 상세
    print("📝 시트 2-6: 회사별 거래처 분석")
    write_firm_sheet(wb, "2_갑우_일반거래처", "갑우", "일반", merged,
        "2. 갑우문화사 — 일반 거래처",
        "코스알엑스(B패턴 1위, -34.6억), 지에스리테일, 에듀윌 제외한 일반 거래처 기준")
    write_firm_sheet(wb, "3_갑우_지입거래처", "갑우", "지입", merged,
        "3. 갑우문화사 — 지입 거래처 (교원·이투스·에듀윌 계열)",
        "용지는 거래처가 직접 구매 → 단가/원가율 판정을 일반 거래처와 같은 기준으로 보면 왜곡됨")
    write_firm_sheet(wb, "4_비피_일반거래처", "비피", "일반", merged,
        "4. 비피앤피 — 일반 거래처",
        "한국조폐공사(A패턴 1위, -36.4억)·케이랩·지담디앤피 등 패키지 본진")
    write_firm_sheet(wb, "5_비피_지입거래처", "비피", "지입", merged,
        "5. 비피앤피 — 지입 거래처",
        "교원구몬(172억, 3사 걸쳐 최대 지입거래처)이 대부분")
    write_firm_sheet(wb, "6_더원_전체", "더원", "일반", merged,
        "6. 더원프린팅 — 전체 거래처",
        "2023년부터 데이터 있음, 연간 4~6억대 소규모")

    # 7. 용지단가
    print("📝 시트 7: 용지단가")
    ws7 = wb.create_sheet("7_용지단가_추이")
    pp = paper_yearly.copy()
    pp["매입액(억)"] = (pp["매입액"]/1e8).round(1)
    pp["단가(원)"] = pp["단가"].round(0)
    pp = pp[["연도","수량","매입액(억)","단가(원)"]]
    write_df_sheet(ws7, pp,
        "7. 용지 매입 단가 추이 (viewGabwoo_마감, 그룹 공통)",
        "2021 27,553원 → 2025 46,991원 = +70.6%. 수량은 198k→137k(-30%)로 감소 중인데 단가가 급등")

    # 8. 후속제안
    ws8 = wb.create_sheet("8_후속제안")
    ws8["A1"] = "8. 후속 제안 (v2 기준)"
    ws8["A1"].font = SECTION_FONT
    ws8["A1"].fill = SECTION_FILL
    ws8.merge_cells("A1:C1")
    props = [
        ("우선순위","행동","기대효과"),
        ("1순위","갑우 본체(10000) 매출 2021→2025 -50% 붕괴 원인 심층 분석", "감사보고서 매출 -50%와 일치 — 가장 시급한 문제"),
        ("1순위","(주)코스알엑스 단가 +72% 인상 시점·사유 시계열 (갑우 B패턴 1위)", "단독 -34.6억 손실의 재현·복구 가능성 점검"),
        ("1순위","한국조폐공사 -36억 손실(비피앤피)의 단가 -28% 인하 의사결정 추적", "비피 본진 거래처 이탈 방지, 재협상 기준점"),
        ("2순위","비피앤피(+35% 성장) 구체적 성장 동력 분석 → 갑우에 이식 가능한지 검토", "그룹 리밸런싱 — 패키지 중심 재편의 근거"),
        ("2순위","지입거래처(교원/이투스/에듀윌)의 실제 수익구조 재점검", "매출은 커 보이지만 제작공임만 받으므로 이익률 다름 — 진짜 기여 재평가"),
        ("3순위","매입 테이블(SCT/PUR)에 CD_CUST_OWN 도입 필요성 제기", "회사별 원가율 분리 분석이 현재 불가능 → 근본 데이터 개선"),
        ("3순위","용지 단가 +70% 급등의 제조사별 기여도 분석 → 재협상 타겟", "매입 원가 절감 (원가율 18%→15% 복귀 시 매출총이익 +10억대)"),
    ]
    for i, row in enumerate(props, start=3):
        for j, val in enumerate(row, start=1):
            c = ws8.cell(row=i, column=j, value=val)
            c.border = BORDER
            if i == 3:
                c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if j == 1 and i != 3:
                c.font = Font(bold=True, color="C00000" if "1" in str(val) else "1F3864")
    ws8.column_dimensions["A"].width = 10
    ws8.column_dimensions["B"].width = 55
    ws8.column_dimensions["C"].width = 60

    wb.save(OUT_XLSX)
    print(f"\n✅ 완료: {OUT_XLSX}")
    print(f"   시트: {', '.join(wb.sheetnames)}")

    # 콘솔 요약
    print("\n" + "="*60)
    print("소속사 × 연도 매출 (억)")
    print("="*60)
    print(pivot.to_string(index=False))
    print("\n그룹 원가율 (%)")
    print(cost_ratio.to_string(index=False))


if __name__ == "__main__":
    main()
