"""
인쇄기별 매출 시뮬레이션 V0.

데이터원:
  1. Supabase Storage gabwoo-data/data.json 의 all_completed (재단지시서 변환 결과, 매일 갱신)
  2. /Users/jack/dev/gabwoo/출판_생산 진행 현황/재단지시서_2025년.xlsx (과거 1년치)
  3. SNOTES SAL_SALESH/L (ERP 매출)

흐름:
  A. 두 데이터원에서 작업 행 추출 (equipment, client, product, 통수, 일자)
  B. 거래처+품명 fuzzy 매칭 → ERP 매출 라인
  C. 한 견적이 여러 호기에 걸치면 통수 비율로 매출 분할
  D. 호기 × 월별 매출 합산
"""
from pathlib import Path
import json
import re
import os
import sys
import urllib.request
import openpyxl
import pymssql

ENV = Path("/Users/jack/dev/gabwoo/견적계산기/.env.local")
OUT = Path(__file__).parent / "output" / "machine_revenue_v0.txt"
JAEDAN_2025 = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/재단지시서_2025년.xlsx")

SUPABASE_URL = "https://btbqzbrtsmwoolurpqgx.supabase.co"
SUPABASE_SR = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJ0YnF6YnJ0c213b29sdXJwcWd4Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTQ0MzgwNSwiZXhwIjoyMDkxMDE5ODA1fQ.M6H48_EfhK_GbNB8LC557VukouFaYdXXTjSD8S5azqw"
DATA_LOCAL = Path("/tmp/gw_check/data.json")


def load_env():
    env = {}
    for line in ENV.read_text().splitlines():
        line = line.strip()
        if line and "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def connect():
    env = load_env()
    return pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=10,
    )


class Tee:
    def __init__(self, path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.f = open(path, "w", encoding="utf-8")

    def __call__(self, *args):
        msg = " ".join(str(a) for a in args)
        print(msg)
        self.f.write(msg + "\n")

    def close(self):
        self.f.close()


def section(log, title):
    log("\n" + "=" * 80)
    log(title)
    log("=" * 80)


def normalize_product(s):
    """제품명 정규화 — 공백/슬래시/구두점 제거, 소문자."""
    if not s:
        return ""
    s = re.sub(r"[\s/_,()\[\]·\-]+", "", str(s)).lower()
    return s


def normalize_client(s):
    """거래처명 정규화 — 공백/괄호/주식회사/(주) 제거."""
    if not s:
        return ""
    s = str(s)
    s = re.sub(r"\(주\)|주식회사|㈜|\(유\)|유한회사", "", s)
    s = re.sub(r"\s+", "", s)
    return s.lower()


def fetch_supabase_data():
    """Storage 인증 다운로드. /tmp 캐시."""
    DATA_LOCAL.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(
        f"{SUPABASE_URL}/storage/v1/object/gabwoo-data/data.json",
        headers={"apikey": SUPABASE_SR, "Authorization": f"Bearer {SUPABASE_SR}"},
    )
    with urllib.request.urlopen(req) as r:
        DATA_LOCAL.write_bytes(r.read())
    return json.loads(DATA_LOCAL.read_text())


def parse_jaedan_2025_xlsx():
    """2025년 재단지시서 엑셀 파싱.
    동일 구조: 헤더 row3 [설비, 거래처, 제품명, ..., 통수, 판수, ..., 후가공, 납기]
    설비 forward-fill 필요 (같은 호기 묶음)."""
    if not JAEDAN_2025.exists():
        return []
    wb = openpyxl.load_workbook(str(JAEDAN_2025), read_only=True, data_only=True)
    if "재단지시서" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["재단지시서"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows or len(rows) < 4:
        return []

    # 헤더 자동 탐지
    header_idx = None
    for i, r in enumerate(rows[:5]):
        if r and any(c == "설비" for c in r if c):
            header_idx = i
            break
    if header_idx is None:
        return []

    header = rows[header_idx]
    col = {name: idx for idx, name in enumerate(header) if name}
    needed = {"설비", "거래처", "제품명", "통수", "판수"}
    if not needed.issubset(col):
        return []

    out = []
    current_equip = None
    for r in rows[header_idx + 1:]:
        if not r:
            continue
        eq = r[col["설비"]]
        client = r[col["거래처"]]
        product = r[col["제품명"]]
        cnt = r[col["통수"]]
        # forward-fill 설비
        if eq and str(eq).strip() not in ("-", ""):
            current_equip = str(eq).strip()
        if not (client and product) or str(client).strip() in ("-", ""):
            continue
        try:
            cnt_int = int(float(cnt)) if cnt and str(cnt).replace(".", "").isdigit() else None
        except (ValueError, TypeError):
            cnt_int = None
        out.append({
            "equipment": current_equip,
            "client": str(client).strip(),
            "product": str(product).strip(),
            "cuts": cnt_int,
            "source": "2025_xlsx",
        })
    return out


def main():
    log = Tee(OUT)

    # ─── 1. 데이터원 로드 ─────────────────────────────────────────
    section(log, "1. 데이터원 로드")

    # 1a. Supabase data.json
    log("\n[a] Supabase Storage data.json 다운로드 중…")
    try:
        sdata = fetch_supabase_data()
        log(f"   updated_at: {sdata.get('updated_at')}")
        log(f"   uploaded_by: {sdata.get('uploaded_by')}")
        log(f"   all_completed: {len(sdata.get('all_completed', []))}건")
        log(f"   today_work: {len(sdata.get('today_work', []))}건")
        log(f"   waiting: {len(sdata.get('waiting', []))}건")
    except Exception as e:
        log(f"   ❌ 실패: {e}")
        sdata = {"all_completed": []}

    # 1b. 2025 엑셀
    log("\n[b] 2025년 재단지시서 엑셀 파싱…")
    jobs_2025 = parse_jaedan_2025_xlsx()
    log(f"   파싱 결과: {len(jobs_2025)}건")

    # 1c. 통합 작업 리스트 (equipment, client, product, cuts, source)
    jobs_all = []
    for j in sdata.get("all_completed", []):
        jobs_all.append({
            "equipment": j.get("equipment"),
            "client": j.get("client"),
            "product": j.get("product"),
            "cuts": j.get("cuts"),  # 통수
            "source": "2026_supabase",
        })
    jobs_all.extend(jobs_2025)
    log(f"\n   → 통합 작업 행: {len(jobs_all)}건")

    # 호기별 분포
    by_eq = {}
    for j in jobs_all:
        eq = j["equipment"] or "(unknown)"
        by_eq[eq] = by_eq.get(eq, 0) + 1
    log("\n   호기별 분포 (Top 10):")
    for eq, n in sorted(by_eq.items(), key=lambda x: -x[1])[:15]:
        log(f"      {eq:<20} {n:>6}건")

    # 거래처별 분포
    by_cust = {}
    for j in jobs_all:
        c = j["client"] or "(unknown)"
        by_cust[c] = by_cust.get(c, 0) + 1
    log("\n   거래처별 분포 (Top 10):")
    for c, n in sorted(by_cust.items(), key=lambda x: -x[1])[:10]:
        log(f"      {c:<30} {n:>6}건")

    # ─── 2. ERP 매출 라인 로드 (2025-2026, 갑우 ST=Y) ────────────
    section(log, "2. ERP 매출 라인 로드 (2025~2026 갑우)")
    conn = connect()
    cur = conn.cursor(as_dict=True)
    cur.execute("""
        SELECT h.NO_SALES, h.DT_SALES, h.CD_CUST, c.NM_CUST,
               i.CD_ITEM, i.NM_ITEM,
               CAST(l.AM AS BIGINT) am
        FROM SAL_SALESL l
        INNER JOIN SAL_SALESH h ON h.NO_SALES = l.NO_SALES
        LEFT JOIN MAS_CUST c ON c.CD_CUST = h.CD_CUST
        LEFT JOIN PRT_ITEM i ON i.CD_ITEM = l.CD_ITEM
        WHERE (h.DT_SALES LIKE '2025%' OR h.DT_SALES LIKE '2026%')
          AND h.CD_CUST_OWN='10000' AND h.ST_SALES='Y' AND h.AM > 0
    """)
    erp_lines = cur.fetchall()
    log(f"   ERP 매출 라인: {len(erp_lines)}건 / 합계 {sum(int(l['am'] or 0) for l in erp_lines):,}원")
    conn.close()

    # 거래처 정규화 인덱스
    erp_by_cust = {}
    for l in erp_lines:
        nc = normalize_client(l["NM_CUST"])
        if not nc:
            continue
        erp_by_cust.setdefault(nc, []).append(l)

    # ─── 3. 매칭 시뮬레이션 ─────────────────────────────────────
    section(log, "3. 매칭 시뮬레이션 (작업행 → ERP 라인)")
    matched_jobs = 0
    matched_am = 0
    sample_matches = []
    sample_misses = []

    for j in jobs_all:
        nc = normalize_client(j["client"])
        np = normalize_product(j["product"])
        if not nc or not np:
            continue
        # 1차: 거래처 정규화 동일 + 제품명 부분일치
        candidates = erp_by_cust.get(nc, [])
        if not candidates:
            # 2차: 거래처 부분 매칭
            for ec in erp_by_cust.keys():
                if nc and (nc in ec or ec in nc):
                    candidates = erp_by_cust[ec]
                    break
        if not candidates:
            sample_misses.append((j["client"], j["product"], "거래처 매칭실패"))
            continue
        # 제품명 부분 매칭
        hit = None
        for c in candidates:
            ep = normalize_product(c["NM_ITEM"])
            if not ep:
                continue
            # 양방향 부분 매칭 + 최소 길이 5
            if len(np) >= 5 and len(ep) >= 5:
                if np in ep or ep in np:
                    hit = c
                    break
                # 토큰 단위 — 슬래시/공백 분리한 토큰 절반 이상 일치
                jt = set(re.findall(r"[가-힣A-Za-z0-9]{2,}", j["product"]))
                et = set(re.findall(r"[가-힣A-Za-z0-9]{2,}", c["NM_ITEM"] or ""))
                if jt and et and len(jt & et) / len(jt) >= 0.5:
                    hit = c
                    break
        if hit:
            matched_jobs += 1
            matched_am += int(hit["am"] or 0)
            if len(sample_matches) < 10:
                sample_matches.append((j["equipment"], j["client"], j["product"][:30], hit["NM_ITEM"][:30] if hit["NM_ITEM"] else "?", int(hit["am"] or 0)))
        else:
            if len(sample_misses) < 10:
                sample_misses.append((j["client"], j["product"][:40], "제품명 매칭실패"))

    log(f"\n   매칭 성공: {matched_jobs:>6,} / 전체 {len(jobs_all):>6,} ({matched_jobs / max(len(jobs_all), 1) * 100:.1f}%)")
    log(f"   매칭 매출 합계: {matched_am:>17,}원")
    log(f"   ERP 전체 매출: {sum(int(l['am'] or 0) for l in erp_lines):>17,}원")
    log(f"   커버율: {matched_am / max(sum(int(l['am'] or 0) for l in erp_lines), 1) * 100:.1f}%")

    log("\n   매칭 샘플 (Top 10):")
    for eq, cl, p, ep, am in sample_matches:
        log(f"      [{eq:<8}] {cl[:15]:<15} | {p:<30} ↔ {ep:<30} ₩{am:>12,}")

    log("\n   매칭 실패 샘플:")
    for c, p, reason in sample_misses[:10]:
        log(f"      [{reason}] {str(c)[:20]:<20} | {p}")

    # ─── 4. 호기별 매출 분할 (단순: 매칭 1:1로 매출 부착) ──────────
    section(log, "4. 호기별 매출 합계 (V0: 매칭 라인 매출을 호기에 단순 합산)")
    by_eq_revenue = {}
    by_eq_jobs = {}
    for j in jobs_all:
        nc = normalize_client(j["client"])
        np = normalize_product(j["product"])
        if not nc or not np:
            continue
        candidates = erp_by_cust.get(nc, [])
        if not candidates:
            for ec in erp_by_cust.keys():
                if nc and (nc in ec or ec in nc):
                    candidates = erp_by_cust[ec]
                    break
        hit = None
        for c in candidates:
            ep = normalize_product(c["NM_ITEM"])
            if len(np) >= 5 and len(ep) >= 5:
                if np in ep or ep in np:
                    hit = c
                    break
                jt = set(re.findall(r"[가-힣A-Za-z0-9]{2,}", j["product"]))
                et = set(re.findall(r"[가-힣A-Za-z0-9]{2,}", c["NM_ITEM"] or ""))
                if jt and et and len(jt & et) / len(jt) >= 0.5:
                    hit = c
                    break
        if hit:
            eq = j["equipment"] or "(unknown)"
            by_eq_revenue[eq] = by_eq_revenue.get(eq, 0) + int(hit["am"] or 0)
            by_eq_jobs[eq] = by_eq_jobs.get(eq, 0) + 1

    log(f"\n   {'호기':<20}{'매칭 작업':>12}{'추정 매출':>20}")
    for eq in sorted(by_eq_revenue.keys(), key=lambda x: -by_eq_revenue[x]):
        log(f"   {eq:<20}{by_eq_jobs.get(eq, 0):>12,}{by_eq_revenue[eq]:>20,}")
    log(f"   {'-' * 60}")
    log(f"   {'합계':<20}{sum(by_eq_jobs.values()):>12,}{sum(by_eq_revenue.values()):>20,}")

    # ⚠️ 주의: 같은 매출 라인이 여러 호기에 매칭되면 중복 합산됨 → V1에서 통수 비율 분할 필요
    log("\n   ⚠️ V0 한계: 같은 매출 라인이 여러 호기 작업과 매칭되면 중복 합산. V1에서 통수 비율 분할 예정.")

    log.close()


if __name__ == "__main__":
    main()
