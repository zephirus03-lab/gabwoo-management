"""
거래처별 '뭐가 이상한지' + 물어볼 질문 리스트 (2025 매출 기준)

목적:
  라포·영향력이 아직 쌓이기 전 단계에서 판정·결론 언어를 피하고,
  거래처별로 '숫자가 이상한 지점'과 '담당 영업자에게 물어볼 질문'을 생성.

구조:
  - 2025 매출(세금계산서 우선, 없으면 ERP) 기준으로 거래처 3단 분리:
    5억+ / 1억~5억 / 1억 미만 (참고용)
  - 각 거래처에 대해 이상점 탐지 (4개 카테고리: 데이터·추세·운영·손익)
  - 각 이상점에 대해 구체 수치 + 물어볼 질문 생성
  - 영업담당자 이름 포함 (MAS_EMP JOIN)

입력:
  - 내 건전도 프로파일 xlsx (이미 생성)
  - Evidence v0 xlsx (5억+ 손익 + 3년 추세)
  - ERP 직접 쿼리 (영업담당자)

출력:
  output/거래처_이상_질문리스트_2025_YYYYMMDD.xlsx
"""

from __future__ import annotations

import sys
import re
from datetime import datetime
from pathlib import Path

try:
    import numpy as np
    import pandas as pd
    import pymssql
except ImportError as e:
    print(f"❌ pip3 install pymssql pandas numpy openpyxl xlrd: {e}")
    sys.exit(1)


# ───────── 경로 ─────────
ENV_FILE = Path("/Users/jack/dev/gabwoo/견적계산기/.env.local")
MY_XLSX = Path("/Users/jack/dev/gabwoo/관리 대시보드/scripts/output/거래처_건전도_2025_20260416.xlsx")
EVID_XLSX = Path("/Users/jack/Downloads/갑우그룹_경영증거패키지_v0_통합.xlsx")
OUT_DIR = Path(__file__).parent / "output"
OUT_XLSX = OUT_DIR / f"거래처_이상_질문리스트_2025_{datetime.now():%Y%m%d}.xlsx"


# ───────── 거래처명 정규화 ─────────
_SUFFIX_RE = re.compile(r"(\(주\)|\(재\)|\(사\)|주식회사|유한회사|\s+)")

def norm_name(s) -> str:
    if pd.isna(s):
        return ""
    return _SUFFIX_RE.sub("", str(s).strip()).upper()


def load_env(p: Path) -> dict:
    env = {}
    for line in p.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip()
    return env


# ───────── ERP: 거래처별 영업담당자 ─────────
def fetch_salespeople() -> pd.DataFrame:
    """각 거래처의 주담당자(건수 최다) + 전체 담당자 목록."""
    env = load_env(ENV_FILE)
    conn = pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=15,
    )
    q = """
    SELECT
        h.CD_CUST,
        h.CD_EMP, e.NM_EMP,
        COUNT(*) AS 건수,
        SUM(CAST(h.AM AS BIGINT)) AS 담당매출
    FROM SAL_SALESH h
    LEFT JOIN MAS_EMP e ON h.CD_EMP = e.CD_EMP
    WHERE h.CD_FIRM='7000'
      AND h.DT_SALES BETWEEN '20250101' AND '20251231'
      AND (h.ST_SALES='Y' OR h.ST_SALES IS NULL) AND h.AM > 0
    GROUP BY h.CD_CUST, h.CD_EMP, e.NM_EMP
    """
    df = pd.read_sql(q, conn)
    conn.close()
    df["NM_EMP"] = df["NM_EMP"].fillna("(미지정)")
    df = df.sort_values(["CD_CUST", "담당매출"], ascending=[True, False]).reset_index(drop=True)

    # 주담당자 (매출 최대)
    top = df.drop_duplicates(subset=["CD_CUST"], keep="first")[
        ["CD_CUST", "NM_EMP", "담당매출", "건수"]
    ].rename(columns={"NM_EMP": "주담당자", "담당매출": "주담당매출원",
                       "건수": "주담당건수"})
    top["주담당매출억"] = (top["주담당매출원"] / 1e8).round(2)

    # 전체 담당자 목록
    def join_all(g):
        parts = [f"{r.NM_EMP}({int(r.건수)}건, {round(r.담당매출/1e8, 2)}억)"
                 for r in g.itertuples()]
        return pd.Series({"담당자수": len(g), "전체담당자": ", ".join(parts)})

    all_df = df.groupby("CD_CUST").apply(join_all).reset_index()

    return top[["CD_CUST", "주담당자", "주담당건수", "주담당매출억"]].merge(all_df, on="CD_CUST", how="left")


# ───────── 건전도 프로파일 로드 ─────────
def load_my_profile() -> pd.DataFrame:
    frames = []
    for sh, cat in [("2_신뢰도부실_판단보류", "부실"), ("3_전체_평가가능", "평가가능"),
                    ("4_지입거래처_별도", "지입"), ("5_소규모_1억미만", "소규모")]:
        df = pd.read_excel(MY_XLSX, sheet_name=sh)
        df["원천시트"] = cat
        frames.append(df)
    all_df = pd.concat(frames, ignore_index=True)
    pri = {"평가가능": 0, "부실": 1, "지입": 2, "소규모": 3}
    all_df["_pri"] = all_df["원천시트"].map(pri)
    all_df = all_df.sort_values("_pri").drop_duplicates(subset=["NO_BIZ", "소속사"], keep="first")
    return all_df.drop(columns=["_pri"]).reset_index(drop=True)


# ───────── Evidence 파싱 ─────────
def parse_evidence_margin() -> pd.DataFrame:
    df = pd.read_excel(EVID_XLSX, sheet_name="3_손익랭킹_5억이상15개사", header=3)
    df = df[df["순위"].apply(lambda x: str(x).isdigit() if pd.notna(x) else False)]
    keep = ["거래처", "매출", "용지원가", "용지원가율", "간접비", "간접비율",
            "총원가(추정)", "추정마진", "추정마진율", "용지원가 신뢰도", "분류"]
    df = df[[c for c in keep if c in df.columns]]
    df["매칭키"] = df["거래처"].apply(norm_name)
    df = df.rename(columns={
        "매출": "EvidM_매출", "용지원가": "EvidM_용지원가", "용지원가율": "EvidM_용지원가율",
        "간접비": "EvidM_간접비", "간접비율": "EvidM_간접비율",
        "총원가(추정)": "EvidM_총원가", "추정마진": "EvidM_마진", "추정마진율": "EvidM_마진율",
        "용지원가 신뢰도": "EvidM_용지신뢰도", "분류": "EvidM_손익분류",
    })
    return df.drop(columns=["거래처"])


def parse_evidence_trend() -> pd.DataFrame:
    frames = []
    df5 = pd.read_excel(EVID_XLSX, sheet_name="5_계속하락_9개사", header=3)
    df5["원가반영"] = "계속하락"
    frames.append(df5)
    df6 = pd.read_excel(EVID_XLSX, sheet_name="6_역행_단가인하_23개사", header=3)
    df6["원가반영"] = "역행(단가인하)"
    frames.append(df6)
    df7a = pd.read_excel(EVID_XLSX, sheet_name="7_일부+반영충분_28개사", header=4, nrows=8)
    df7a["원가반영"] = "일부반영"
    frames.append(df7a)
    df7b = pd.read_excel(EVID_XLSX, sheet_name="7_일부+반영충분_28개사", header=15, nrows=20)
    df7b["원가반영"] = "반영충분"
    frames.append(df7b)
    combined = pd.concat(frames, ignore_index=True)
    combined = combined[combined["거래처코드"].notna()]
    combined = combined[combined["거래처코드"].astype(str).str.strip() != ""]
    pri = {"계속하락": 0, "역행(단가인하)": 1, "일부반영": 2, "반영충분": 3}
    combined["_tp"] = combined["원가반영"].map(pri)
    combined = combined.sort_values("_tp").drop_duplicates(subset=["거래처코드"], keep="first")
    keep = ["거래처명", "매출23(억)", "매출24(억)", "매출25(억)",
            "단가23", "단가24", "단가25", "단가Δ 23→25(%)",
            "단가추세", "원가반영", "원인주석"]
    combined = combined[[c for c in keep if c in combined.columns]]
    combined.columns = ["EvidT_거래처명", "매출23_억", "매출24_억", "매출25_억",
                         "단가23", "단가24", "단가25", "단가변화_23_25",
                         "EvidT_단가추세", "EvidT_원가반영", "EvidT_원인주석"]
    combined["매칭키"] = combined["EvidT_거래처명"].apply(norm_name)
    return combined.drop(columns=["EvidT_거래처명"]).reset_index(drop=True)


# ───────── 이상점 탐지 함수 ─────────
# 각 함수는 (이상점_구체, 물어볼_질문) 튜플 리스트 반환

def anomalies_data(row) -> list[tuple[str, str]]:
    """데이터 정합성 이상 — ERP vs 세금계산서."""
    out = []
    erp = row.get("ERP매출(억)")
    tax = row.get("세금매출(억)")
    erp = 0 if pd.isna(erp) else float(erp)
    tax = 0 if pd.isna(tax) else float(tax)
    gap = round(erp - tax, 2)

    if erp == 0 and tax > 0.3:  # 세금 0.3억 이상인데 ERP 없음
        out.append((
            f"세금계산서에는 {tax:.2f}억이 발행됐는데 ERP 매출 기록이 없음",
            f"세금계산서상 2025년 {tax:.2f}억이 발행됐는데 ERP(SAL_SALESH)에 해당 거래처 매출이 등록돼 있지 않습니다. "
            f"혹시 다른 소속사 계정에 입력되었거나, ERP 등록이 누락된 건지 확인 부탁드립니다.",
        ))
    elif tax == 0 and erp > 0.3:
        out.append((
            f"ERP에는 {erp:.2f}억이 있는데 세금계산서 발행 기록이 없음",
            f"ERP에 2025년 {erp:.2f}억의 매출이 등록돼 있는데 세금계산서는 해당 거래처로 발행되지 않았습니다. "
            f"다른 법인(관계사) 명의로 발행됐는지, 발행 전 단계인지, 또는 취소 처리 누락인지 확인 부탁드립니다.",
        ))
    elif erp > 0.3 and tax > 0.3:
        ratio = erp / tax
        if ratio > 1.4:
            out.append((
                f"ERP 매출 {erp:.2f}억 > 세금계산서 {tax:.2f}억 (ERP가 {ratio:.2f}배, 갭 +{gap:.2f}억)",
                f"ERP에 {erp:.2f}억이 잡혀 있는데 실제 세금계산서는 {tax:.2f}억만 발행됐습니다 (갭 +{gap:.2f}억). "
                f"라인 이중등록, 월별 이월 발행, 취소된 매출의 ST_SALES 미수정, 또는 다른 법인으로 세금계산서 발행 중 어느 경우인지 "
                f"영업자에게 확인 부탁드립니다.",
            ))
        elif ratio < 0.75:
            out.append((
                f"ERP 매출 {erp:.2f}억 < 세금계산서 {tax:.2f}억 (ERP가 {ratio:.2f}배, 갭 {gap:.2f}억)",
                f"세금계산서는 {tax:.2f}억이 발행됐는데 ERP에는 {erp:.2f}억만 등록돼 있습니다 (갭 {gap:.2f}억). "
                f"ST_SALES='NULL' 미처리 상태로 남은 건인지, 다른 소속사에 입력됐는지 확인 부탁드립니다.",
            ))
    return out


def anomalies_trend(row) -> list[tuple[str, str]]:
    """단가·매출 추세 이상 — Evidence 시트 5/6/7 기반."""
    out = []
    trend_label = str(row.get("EvidT_단가추세", "") or "")
    reflect = str(row.get("EvidT_원가반영", "") or "")
    delta = row.get("단가변화_23_25")
    u23 = row.get("단가23")
    u25 = row.get("단가25")
    y23 = row.get("매출23_억")
    y24 = row.get("매출24_억")
    y25 = row.get("매출25_억")

    # 3년 연속 단가 하락
    if "계속하락" in trend_label and pd.notna(delta):
        u23s = f"{int(u23)}원" if pd.notna(u23) else "—"
        u25s = f"{int(u25)}원" if pd.notna(u25) else "—"
        delta_pct = float(delta) * 100 if pd.notna(delta) and abs(float(delta)) < 5 else float(delta)
        out.append((
            f"2023→24→25 평균 판매단가 3년 연속 하락 ({u23s}→{u25s}, 누적 {delta_pct:+.1f}%)",
            f"이 거래처는 2023년부터 평균 판매단가가 매년 내려가고 있습니다 ({u23s} → {u25s}, 누적 {delta_pct:+.1f}%). "
            f"같은 기간 갑우 용지 매입 평균 단가는 +22.6% 올랐습니다. "
            f"거래처 측 단가 인하 요청이 있었는지, 경쟁사 견적이 들어온 적 있는지, 계약 갱신 시 어떤 논의가 있었는지 확인 부탁드립니다.",
        ))
    elif reflect == "역행(단가인하)" and pd.notna(delta):
        delta_pct = float(delta) * 100 if pd.notna(delta) and abs(float(delta)) < 5 else float(delta)
        out.append((
            f"2023→2025 평균 판매단가 누적 {delta_pct:+.1f}% (용지원가 +22.6% 동안 역행)",
            f"2023→2025 평균 판매단가가 {delta_pct:+.1f}% 움직여서 용지 원가 +22.6% 상승을 흡수하지 못했습니다. "
            f"단가 재협상 제안을 해 본 적이 있는지, 거래처 쪽 반응은 어땠는지 확인 부탁드립니다.",
        ))

    # 매출 -50% 이상 급감 (24→25)
    if pd.notna(y24) and pd.notna(y25) and y24 > 0.5:
        pct = (y25 - y24) / y24
        if pct <= -0.4:
            out.append((
                f"2024 {y24:.2f}억 → 2025 {y25:.2f}억 매출 {pct*100:.0f}% 급감",
                f"2024년 {y24:.2f}억에서 2025년 {y25:.2f}억으로 매출이 {abs(pct)*100:.0f}% 줄었습니다. "
                f"거래 중단인지, 담당자·발주 책임자 변경인지, 일부 품목만 축소됐는지 확인 부탁드립니다. "
                f"또한 경쟁사로 이관된 물량이 있는지, 복귀 가능성은 있는지도 함께 여쭤봐 주세요.",
            ))

    # 매출 상승→하락 전환 (23→24 상승, 24→25 하락 10%+)
    if pd.notna(y23) and pd.notna(y24) and pd.notna(y25) and y23 > 0.5 and y24 > 0.5:
        g1 = (y24 - y23) / y23
        g2 = (y25 - y24) / y24
        if g1 > 0.1 and g2 < -0.1:
            out.append((
                f"매출 상승→하락 전환 ({y23:.2f}→{y24:.2f}→{y25:.2f}억, 2024 대비 {g2*100:.0f}%)",
                f"2023→2024는 매출이 {g1*100:.0f}% 늘었는데 2024→2025에는 {g2*100:.0f}% 줄었습니다. "
                f"2025년 하반기부터 달라진 상황이 있는지, 일회성 프로젝트 종료인지 확인 부탁드립니다.",
            ))
    return out


def anomalies_ops(row) -> list[tuple[str, str]]:
    """운영 이상 — 취소·미처리·할인·변동·공백."""
    out = []
    # 취소율
    cancel = row.get("취소율_pct")
    cancel_amt = row.get("매출_취소N", 0)
    if pd.notna(cancel) and cancel >= 3:
        ca = float(cancel_amt) / 1e8 if pd.notna(cancel_amt) else 0
        out.append((
            f"취소된 매출 비중 {cancel:.1f}% (ST_SALES='N' 약 {ca:.2f}억)",
            f"2025년 이 거래처에서 취소(ST_SALES='N') 처리된 매출이 전체의 {cancel:.1f}% ({ca:.2f}억)입니다. "
            f"어떤 건이 취소됐는지, 이후 재발행된 건이 있는지, 취소 사유는 무엇인지 구체 건 확인 부탁드립니다.",
        ))

    # 미처리율
    pend = row.get("미처리율_pct")
    pend_amt = row.get("매출_미처리NULL", 0)
    if pd.notna(pend) and pend >= 15:
        pa = float(pend_amt) / 1e8 if pd.notna(pend_amt) else 0
        out.append((
            f"ST_SALES='NULL' 미처리 매출 {pend:.1f}% (약 {pa:.2f}억)",
            f"ERP에 등록된 이 거래처 매출 중 {pend:.1f}% ({pa:.2f}억)가 아직 ST_SALES='NULL' 상태입니다. "
            f"세금계산서가 발행됐는데 ERP 상태만 업데이트가 안 된 건지, 확정 전 단계로 대기 중인 건지 확인 부탁드립니다.",
        ))

    # 승인 없이 확정
    approve = row.get("승인없이확정_pct")
    if pd.notna(approve) and approve >= 10:
        out.append((
            f"승인 없이(YN_APP='N') 확정된 매출 비중 {approve:.1f}%",
            f"YN_APP='N' 인데도 ST_SALES='Y'로 확정 처리된 매출이 {approve:.1f}%입니다. "
            f"승인 프로세스를 건너뛴 사유가 있는지 (긴급 건, 일괄 확정 등), 정상 절차로 처리된 건지 확인 부탁드립니다.",
        ))

    # 평균 할인율
    disc = row.get("평균할인율")
    if pd.notna(disc) and disc >= 10:
        out.append((
            f"라인 평균 할인율 {disc:.1f}%",
            f"라인 기준 평균 할인율이 {disc:.1f}%로 잡혀 있습니다. "
            f"연간 계약 단가 자체가 할인 반영된 건지, 건별로 네고해서 할인을 적용한 건지 확인 부탁드립니다.",
        ))

    # 월 변동계수
    vol = row.get("ERP_월변동계수")
    if pd.notna(vol) and vol >= 0.8:
        out.append((
            f"월매출 변동계수 {vol:.2f} (월별 편차가 평균의 80%+)",
            f"월별 매출 변동이 큽니다 (변동계수 {vol:.2f}). "
            f"프로젝트 단위 발주라서 자연스러운 건지, 특정 달에 몰리는 패턴인지, 아니면 불규칙 발주인지 확인 부탁드립니다.",
        ))

    # 거래 공백 90일+
    gap = row.get("거래공백일")
    last = row.get("최근거래_dt")
    if pd.notna(gap) and gap >= 90:
        ld = ""
        if pd.notna(last):
            try:
                ld = f" (마지막 거래 {pd.to_datetime(last).strftime('%Y-%m-%d')})"
            except Exception:
                ld = ""
        out.append((
            f"최근 거래 이후 {int(gap)}일 공백{ld}",
            f"최근 거래일 이후 {int(gap)}일간 새 거래가 없습니다{ld}. "
            f"거래가 중단된 상태인지, 2026년 발주 예정이 있는지, 담당자·연락 채널에 변화가 있었는지 확인 부탁드립니다.",
        ))
    return out


def anomalies_pnl(row) -> list[tuple[str, str]]:
    """손익 이상 — Evidence v0 5억+ 15개사 중에서만 가능."""
    out = []
    margin = row.get("EvidM_마진율")
    paper_ratio = row.get("EvidM_용지원가율")
    rev = row.get("EvidM_매출")
    paper = row.get("EvidM_용지원가")
    indirect = row.get("EvidM_간접비")
    cred = row.get("EvidM_용지신뢰도", "")

    if pd.notna(margin) and margin < 0 and pd.notna(rev):
        rev_eok = float(rev) / 1e8
        paper_eok = float(paper) / 1e8 if pd.notna(paper) else 0
        ind_eok = float(indirect) / 1e8 if pd.notna(indirect) else 0
        out.append((
            f"추정 마진율 {margin*100:.1f}% (매출 {rev_eok:.2f}억 vs 총원가 추정 {paper_eok+ind_eok:.2f}억, 용지 {paper_eok:.2f}+간접 {ind_eok:.2f})",
            f"매출 대비 원가(용지비 {paper_eok:.2f}억 + 간접비 {ind_eok:.2f}억)가 매출 {rev_eok:.2f}억을 넘어서 추정 마진이 -{abs(margin)*100:.1f}%입니다. "
            f"용지비는 {cred} 기준이라 실측에 가깝습니다. 실제로 적자 거래인지, 용지비 배분 산식에 문제가 있는 건지 확인 부탁드립니다. "
            f"계속 거래해야 할 전략적 사유가 있는지도 함께 여쭤봐 주세요.",
        ))
    elif pd.notna(paper_ratio) and 0.45 <= paper_ratio < 0.99:
        rev_eok = float(rev) / 1e8 if pd.notna(rev) else 0
        paper_eok = float(paper) / 1e8 if pd.notna(paper) else 0
        out.append((
            f"용지원가율 {paper_ratio*100:.1f}% (매출 {rev_eok:.2f}억 중 용지비 {paper_eok:.2f}억)",
            f"매출 {rev_eok:.2f}억 중 용지비가 {paper_eok:.2f}억({paper_ratio*100:.1f}%)을 차지합니다. "
            f"특수지·고급지 사용인지, 일반 인쇄 대비 원가 구조가 왜 이렇게 높은지, 단가 재협상 여지가 있는지 확인 부탁드립니다. "
            f"간접비까지 포함하면 실마진이 매우 얇아 보입니다.",
        ))
    elif pd.notna(paper_ratio) and paper_ratio < 0.01 and pd.notna(rev) and float(rev) > 5e8:
        rev_eok = float(rev) / 1e8
        out.append((
            f"용지원가 거의 0원 (매출 {rev_eok:.2f}억 규모 대비 비정상)",
            f"매출 {rev_eok:.2f}억 규모인데 용지 매입이 거의 잡히지 않았습니다. "
            f"거래처가 용지를 직접 공급하는 방식(용지사급)인지, 외부 윤전기 공장 외주인지 확인 부탁드립니다. "
            f"만약 용지사급이라면 원가 구조 해석이 일반 거래처와 다르다는 점을 반영해야 합니다.",
        ))
    return out


# ───────── 통합 ─────────
def build():
    print("📥 ERP 영업담당자 조회...")
    sales = fetch_salespeople()
    print(f"   → {len(sales):,} (거래처×소속사) 담당자 매핑")

    print("📥 건전도 프로파일...")
    my = load_my_profile()
    print(f"   → {len(my):,}건")

    print("📥 Evidence 손익(5억+) + 추세(51개사)...")
    em = parse_evidence_margin()
    et = parse_evidence_trend()
    print(f"   → 손익 {len(em)} / 추세 {len(et)}")

    my["매칭키"] = my["거래처명"].apply(norm_name)

    # 영업담당자 JOIN — 내 프로파일에 CD_CUST 없으므로 NO_BIZ로 CD_CUST를 먼저 확보해야 함
    # 대신 소속사 + 거래처명 기준 매칭은 어렵고, CD_CUST가 필요해서 ERP에서 추가 매핑
    env = load_env(ENV_FILE)
    conn = pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=15,
    )
    cust_map = pd.read_sql("""
        SELECT c.CD_CUST, c.NO_BIZ, c.NM_CUST
        FROM MAS_CUST c
        WHERE c.CD_FIRM='7000'
    """, conn)
    conn.close()
    def norm_biz(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        if s.endswith(".0"): s = s[:-2]
        return s.replace("-", "").replace(" ", "")
    cust_map["NO_BIZ"] = cust_map["NO_BIZ"].apply(norm_biz)
    my["NO_BIZ"] = my["NO_BIZ"].apply(norm_biz)
    my2 = my.merge(cust_map[["CD_CUST", "NO_BIZ"]], on="NO_BIZ", how="left")

    # CD_CUST_OWN은 내 프로파일의 '소속사' 컬럼에 한글로 있음 → 코드 변환
    firm_code = {"갑우": "10000", "비피": "20000", "더원": "30000"}
    my2["CD_CUST_OWN"] = my2["소속사"].map(firm_code)

    # 영업담당자 JOIN — CD_CUST 기준
    my2 = my2.merge(sales, on="CD_CUST", how="left")
    print(f"   주담당자 매칭: {my2['주담당자'].notna().sum()}/{len(my2)}")

    # Evidence JOIN
    my2 = my2.merge(em, on="매칭키", how="left")
    my2 = my2.merge(et, on="매칭키", how="left")

    # 이상점 탐지
    print("🔎 이상점 탐지 + 질문 생성...")
    rows_detail = []
    summary_rows = []
    for _, r in my2.iterrows():
        anom = (
            [("데이터", a, q) for a, q in anomalies_data(r)] +
            [("추세", a, q) for a, q in anomalies_trend(r)] +
            [("운영", a, q) for a, q in anomalies_ops(r)] +
            [("손익", a, q) for a, q in anomalies_pnl(r)]
        )
        tax = r.get("세금매출(억)") or 0
        erp = r.get("ERP매출(억)") or 0
        rev = max(float(tax) if pd.notna(tax) else 0, float(erp) if pd.notna(erp) else 0)

        for cat, desc, q in anom:
            rows_detail.append({
                "소속사": r.get("소속사"),
                "거래처명": r.get("거래처명"),
                "주담당자": r.get("주담당자", "(미지정)"),
                "전체담당자": r.get("전체담당자", ""),
                "2025매출_기준억": round(rev, 2),
                "카테고리": cat,
                "이상한_점": desc,
                "물어볼_질문": q,
            })

        if anom:
            summary_rows.append({
                "소속사": r.get("소속사"),
                "거래처명": r.get("거래처명"),
                "주담당자": r.get("주담당자", "(미지정)"),
                "전체담당자": r.get("전체담당자", ""),
                "2025매출_기준억": round(rev, 2),
                "ERP매출억": r.get("ERP매출(억)"),
                "세금매출억": r.get("세금매출(억)"),
                "이상점_수": len(anom),
                "이상점_요약": " · ".join(f"[{c}] {d}" for c, d, _ in anom),
                "물어볼_질문_전체": "\n\n".join(
                    f"Q{i+1}. [{c}] {q}"
                    for i, (c, _, q) in enumerate(anom)
                ),
            })
        else:
            summary_rows.append({
                "소속사": r.get("소속사"),
                "거래처명": r.get("거래처명"),
                "주담당자": r.get("주담당자", "(미지정)"),
                "전체담당자": r.get("전체담당자", ""),
                "2025매출_기준억": round(rev, 2),
                "ERP매출억": r.get("ERP매출(억)"),
                "세금매출억": r.get("세금매출(억)"),
                "이상점_수": 0,
                "이상점_요약": "특이사항 없음",
                "물어볼_질문_전체": "",
            })

    detail = pd.DataFrame(rows_detail)
    summary = pd.DataFrame(summary_rows)

    # 규모 분리
    summary = summary.sort_values("2025매출_기준억", ascending=False).reset_index(drop=True)
    # 중복 제거 (같은 거래처가 다른 소속사로 두 번 들어올 수 있음)
    summary = summary.drop_duplicates(subset=["소속사", "거래처명"], keep="first")

    tier_A = summary[summary["2025매출_기준억"] >= 5].copy()
    tier_B = summary[(summary["2025매출_기준억"] >= 1) & (summary["2025매출_기준억"] < 5)].copy()
    tier_C = summary[summary["2025매출_기준억"] < 1].copy()

    detail_A = detail[detail["2025매출_기준억"] >= 5].copy()
    detail_B = detail[(detail["2025매출_기준억"] >= 1) & (detail["2025매출_기준억"] < 5)].copy()

    # 메타
    meta = pd.DataFrame({
        "항목": [
            "목적", "기준 기간", "매출 기준", "5억+ 거래처 수", "1억~5억 거래처 수", "1억 미만",
            "이상점 탐지 카테고리", "담당자 출처", "외부 결론·판정 금지",
        ],
        "값": [
            "거래처별 '뭐가 이상한지' + 영업담당자에게 물어볼 질문 정리",
            "2025-01-01 ~ 2025-12-31",
            "세금계산서 발행 금액 우선, 없으면 ERP 매출(AM, ST=Y/NULL)",
            f"{len(tier_A):,}",
            f"{len(tier_B):,}",
            f"{len(tier_C):,}",
            "데이터 정합성 / 단가·매출 추세 / 운영 플래그 / 손익(5억+만)",
            "SAL_SALESH.CD_EMP JOIN MAS_EMP (2025 매출 건수 기준 주담당 + 전체 담당자)",
            "이 문서는 질문 생성용. 거래축소·재협상 같은 결론은 담당자 답변 후 본부장·대표가 결정",
        ],
    })

    return meta, tier_A, tier_B, tier_C, detail_A, detail_B


def main():
    meta, tA, tB, tC, dA, dB = build()

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        meta.to_excel(w, sheet_name="0_메타", index=False)
        tA.to_excel(w, sheet_name=f"A_5억이상_{len(tA)}개사_요약", index=False)
        tB.to_excel(w, sheet_name=f"B_1억~5억_{len(tB)}개사_요약", index=False)
        dA.to_excel(w, sheet_name="C_5억이상_이상점상세", index=False)
        dB.to_excel(w, sheet_name="D_1억~5억_이상점상세", index=False)
        tC.to_excel(w, sheet_name=f"E_1억미만_{len(tC)}개사(참고)", index=False)

    # 엑셀 열 너비·줄바꿈
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    wb = load_workbook(OUT_XLSX)
    for sh in wb.sheetnames:
        ws = wb[sh]
        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = 0
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if cell.value is not None:
                    l = len(str(cell.value))
                    if l > max_len:
                        max_len = min(l, 80)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
    wb.save(OUT_XLSX)

    print(f"\n📤 {OUT_XLSX}\n")
    print("=" * 100)
    print(f"▼ A. 5억+ 거래처 {len(tA)}개사 — 이상점 있는 거 먼저")
    print("=" * 100)
    cols = ["소속사", "거래처명", "주담당자", "2025매출_기준억", "이상점_수", "이상점_요약"]
    tA_sorted = tA.sort_values(["이상점_수", "2025매출_기준억"], ascending=[False, False])
    pd.set_option("display.max_colwidth", 100)
    pd.set_option("display.width", 250)
    print(tA_sorted[cols].head(20).to_string(index=False))

    print("\n" + "=" * 100)
    print(f"▼ B. 1억~5억 거래처 {len(tB)}개사 — 이상점 상위 10")
    print("=" * 100)
    tB_sorted = tB.sort_values(["이상점_수", "2025매출_기준억"], ascending=[False, False])
    print(tB_sorted[cols].head(10).to_string(index=False))


if __name__ == "__main__":
    main()
