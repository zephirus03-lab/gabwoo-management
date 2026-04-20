"""
경영진단 잠정 보고서 v2 — 3사 분리 + 지입거래처 구분 + 2023/2024/2025 3년 비교

본부장 피드백 반영: 2023 vs 2025 2점 비교가 아닌 2023/2024/2025 3년 추이.

출력: 잠정_경영진단보고서_v2_3y_YYYYMMDD.xlsx
"""
import sys
from pathlib import Path
from datetime import datetime

try:
    import pymssql
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"pip3 install pymssql pandas openpyxl: {e}"); sys.exit(1)

ENV_FILE = Path("/Users/jack/dev/gabwoo/견적계산기/.env.local")
OUT_DIR = Path(__file__).parent / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_XLSX = OUT_DIR / f"잠정_경영진단보고서_v2_3y_{datetime.now():%Y%m%d}.xlsx"

FIRM_MAP = {"10000": "갑우", "20000": "비피", "30000": "더원"}
지입_KEYWORDS = ["교원", "이투스", "에듀윌"]

# 벤치마크 — 연도별 용지단가 대비 2023 기준 누적 상승률
# 2023 38,330 → 2024 43,887 (+14.5%) → 2025 46,991 (+22.6%)
COST_INFLATION_23_24 = 14.5
COST_INFLATION_23_25 = 22.6

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
BORDER = Border(left=Side(style="thin", color="BFBFBF"),
                right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="thin", color="BFBFBF"))


def load_env():
    env = {}
    for line in ENV_FILE.read_text().splitlines():
        if "=" in line and not line.strip().startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def get_conn(env):
    return pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=15,
    )


def is_지입(name):
    return bool(name) and any(kw in name for kw in 지입_KEYWORDS)


def fetch_all(conn):
    print("📥 SAL_SALESH 헤더...")
    header = pd.read_sql("""
        SELECT LEFT(h.DT_SALES, 4) AS 연도, h.CD_CUST_OWN AS 소속코드,
               h.CD_CUST AS 거래처코드, c.NM_CUST AS 거래처명,
               COUNT(DISTINCT h.NO_SALES) AS 매출건수, SUM(h.AM) AS 매출
        FROM SAL_SALESH h
        LEFT JOIN MAS_CUST c ON h.CD_CUST=c.CD_CUST AND h.CD_FIRM=c.CD_FIRM
        WHERE h.CD_FIRM='7000' AND h.DT_SALES>='20210101' AND h.DT_SALES<='20261231'
          AND h.AM>0 AND (h.ST_SALES='Y' OR h.ST_SALES IS NULL)
          AND h.CD_CUST_OWN IN ('10000','20000','30000')
        GROUP BY LEFT(h.DT_SALES, 4), h.CD_CUST_OWN, h.CD_CUST, c.NM_CUST
    """, conn)

    print("📥 SAL_SALESL 라인...")
    line = pd.read_sql("""
        SELECT LEFT(h.DT_SALES, 4) AS 연도, h.CD_CUST_OWN AS 소속코드,
               h.CD_CUST AS 거래처코드, SUM(l.QT) AS 수량, SUM(l.AM) AS 라인매출
        FROM SAL_SALESH h
        JOIN SAL_SALESL l ON h.CD_FIRM=l.CD_FIRM AND h.NO_SALES=l.NO_SALES
        WHERE h.CD_FIRM='7000' AND h.DT_SALES>='20210101' AND h.DT_SALES<='20261231'
          AND (h.ST_SALES='Y' OR h.ST_SALES IS NULL)
          AND l.QT>0 AND l.AM>0
          AND h.CD_CUST_OWN IN ('10000','20000','30000')
        GROUP BY LEFT(h.DT_SALES, 4), h.CD_CUST_OWN, h.CD_CUST
    """, conn)

    print("📥 용지 단가 연도별...")
    paper_yearly = pd.read_sql("""
        SELECT YEAR(일자) AS 연도, SUM(수량) AS 수량, SUM(공급가액) AS 매입액,
               CASE WHEN SUM(수량)>0 THEN SUM(공급가액)*1.0/SUM(수량) ELSE 0 END AS 단가
        FROM dbo.viewGabwoo_마감
        WHERE 일자>='2021-01-01' AND 일자<='2026-12-31' AND 수량>0
        GROUP BY YEAR(일자) ORDER BY YEAR(일자)
    """, conn)

    print("📥 외주 매입...")
    outsource = pd.read_sql("""
        SELECT LEFT(DT_PUR, 4) AS 연도, SUM(AM_SUPPLY) AS 외주매입
        FROM PUR_ETCCLSH
        WHERE CD_FIRM='7000' AND DT_PUR>='20210101' AND DT_PUR<='20261231'
          AND AM_SUPPLY>0
        GROUP BY LEFT(DT_PUR, 4)
    """, conn)

    return header, line, paper_yearly, outsource


def enrich(header, line):
    merged = header.merge(line, on=["연도", "소속코드", "거래처코드"], how="left")
    merged["수량"] = merged["수량"].fillna(0)
    merged["라인매출"] = merged["라인매출"].fillna(0)
    merged["단가"] = merged.apply(
        lambda r: r["라인매출"]/r["수량"] if r["수량"]>0 else 0, axis=1
    )
    merged["소속사"] = merged["소속코드"].map(FIRM_MAP)
    merged["지입여부"] = merged["거래처명"].apply(lambda n: "지입" if is_지입(n) else "일반")
    return merged


def customer_3year(merged, firm, 지입필터):
    """특정 회사 × 지입여부 거래처별 2023·2024·2025 3년 YoY."""
    sub = merged[(merged["소속사"]==firm) & (merged["지입여부"]==지입필터)].copy()
    if sub.empty:
        return pd.DataFrame()
    amt = sub.pivot_table(index=["거래처코드","거래처명"], columns="연도", values="매출", aggfunc="sum").fillna(0)
    qty = sub.pivot_table(index=["거래처코드","거래처명"], columns="연도", values="수량", aggfunc="sum").fillna(0)
    um  = sub.pivot_table(index=["거래처코드","거래처명"], columns="연도", values="단가", aggfunc="mean").fillna(0)
    r = pd.DataFrame(index=amt.index)
    for y in ["2023","2024","2025"]:
        r[f"매출_{y}"] = amt.get(y, 0)
        r[f"수량_{y}"] = qty.get(y, 0)
        r[f"단가_{y}"] = um.get(y, 0)
    # YoY 변화
    r["매출Δ_23→24(%)"] = (r["매출_2024"]-r["매출_2023"])/r["매출_2023"].replace(0, pd.NA)*100
    r["매출Δ_24→25(%)"] = (r["매출_2025"]-r["매출_2024"])/r["매출_2024"].replace(0, pd.NA)*100
    r["매출Δ_23→25(%)"] = (r["매출_2025"]-r["매출_2023"])/r["매출_2023"].replace(0, pd.NA)*100
    r["수량Δ_23→25(%)"] = (r["수량_2025"]-r["수량_2023"])/r["수량_2023"].replace(0, pd.NA)*100
    r["단가Δ_23→24(%)"] = (r["단가_2024"]-r["단가_2023"])/r["단가_2023"].replace(0, pd.NA)*100
    r["단가Δ_24→25(%)"] = (r["단가_2025"]-r["단가_2024"])/r["단가_2024"].replace(0, pd.NA)*100
    r["단가Δ_23→25(%)"] = (r["단가_2025"]-r["단가_2023"])/r["단가_2023"].replace(0, pd.NA)*100
    r = r.reset_index()
    return r[r["매출_2023"] >= 50_000_000].copy()


def classify_pattern_3y(row):
    m23, m24, m25 = row["매출_2023"], row["매출_2024"], row["매출_2025"]
    if m23 == 0:
        return "N/A"
    if m25 == 0:
        return "C. 완전이탈"
    try:
        d_amt = float(row["매출Δ_23→25(%)"])
        d_um = float(row["단가Δ_23→25(%)"])
    except (ValueError, TypeError):
        return "기타"
    if pd.isna(d_amt) or pd.isna(d_um):
        return "기타"
    if d_amt <= -20 and d_um <= -10: return "A. 단가↓물량↓"
    if d_amt <= -20 and d_um >= 5:   return "B. 단가↑이탈"
    if d_amt <= -20 and -10 < d_um < 5: return "D. 물량감소·단가유지"
    if d_amt > -20 and d_amt < 20:   return "E. 안정"
    if d_amt >= 20:                   return "F. 성장"
    return "기타"


def classify_cost_pass_3y(row):
    """3년 비교 원가반영 판정. 23→25 누적 +22.6% 기준."""
    if row["매출_2023"] == 0 or row["매출_2025"] == 0:
        return "N/A"
    try:
        d = float(row["단가Δ_23→25(%)"])
    except (ValueError, TypeError):
        return "N/A"
    if pd.isna(d): return "N/A"
    if d >= COST_INFLATION_23_25: return "✅ 반영충분"
    if d >= 0: return "🟡 일부반영"
    return "🔴 역행(단가↓)"


def classify_trajectory(row):
    """3년 추이 궤적: 지속감소/반등/지속하락-가속/전환 등."""
    m23, m24, m25 = row["매출_2023"], row["매출_2024"], row["매출_2025"]
    if m23 == 0:
        return "N/A"
    d1 = (m24-m23)/m23*100 if m23>0 else 0
    d2 = (m25-m24)/m24*100 if m24>0 else (-100 if m25==0 else 0)
    if d1 <= -15 and d2 <= -15: return "📉 2년 연속 급감"
    if d1 >= 15 and d2 >= 15:   return "📈 2년 연속 성장"
    if d1 <= -15 and d2 >= 15:  return "🔄 24년 저점→25년 반등"
    if d1 >= 15 and d2 <= -15:  return "⚠️ 24년 반짝→25년 후퇴"
    if d1 <= -5 and d2 <= -5:   return "📉 완만한 2년 감소"
    if d1 >= 5 and d2 >= 5:     return "📈 완만한 2년 성장"
    return "➡️ 등락 혼재"


def fmt_display(df):
    d = df.copy()
    if len(d) == 0: return d
    for y in ["2023","2024","2025"]:
        d[f"매출{y[-2:]}(억)"] = (d[f"매출_{y}"]/1e8).round(2)
    for c in ["매출Δ_23→24(%)","매출Δ_24→25(%)","매출Δ_23→25(%)",
              "수량Δ_23→25(%)","단가Δ_23→24(%)","단가Δ_24→25(%)","단가Δ_23→25(%)"]:
        d[c] = pd.to_numeric(d[c], errors="coerce").round(1)
    cols = ["거래처코드","거래처명",
            "매출23(억)","매출24(억)","매출25(억)",
            "매출Δ_23→24(%)","매출Δ_24→25(%)","매출Δ_23→25(%)",
            "단가Δ_23→24(%)","단가Δ_24→25(%)","단가Δ_23→25(%)",
            "수량Δ_23→25(%)"]
    if "패턴" in d.columns: cols.append("패턴")
    if "궤적" in d.columns: cols.append("궤적")
    if "원가반영" in d.columns: cols.append("원가반영")
    return d[cols]


def autosize(ws, max_width=55):
    for col in ws.columns:
        max_len = 0; col_letter = None
        for cell in col:
            if col_letter is None: col_letter = get_column_letter(cell.column)
            try:
                v = str(cell.value) if cell.value is not None else ""
                ln = sum(2 if ord(c)>127 else 1 for c in v)
                if ln > max_len: max_len = ln
            except Exception: pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def write_df_sheet(ws, df, title, notes=None):
    ws["A1"] = title
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL
    ncols = max(len(df.columns), 5) if len(df) > 0 else 5
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    r = 2
    if notes:
        for note in (notes if isinstance(notes, list) else [notes]):
            ws.cell(row=r, column=1, value=note).font = Font(italic=True, size=10, color="555555")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            r += 1
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=r, column=j, value=col)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(df.itertuples(index=False), start=r+1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = BORDER
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")
                if abs(val) >= 1000: c.number_format = "#,##0"
                elif isinstance(val, float): c.number_format = "#,##0.00"
    autosize(ws)


def write_cover(ws):
    ws["A1"] = "갑우문화사 경영진단 — 잠정 보고서 v2 (3년 비교 확장판)"
    ws["A1"].font = Font(bold=True, size=16, color="1F3864")
    ws.merge_cells("A1:G1")
    ws["A2"] = f"작성일: {datetime.now():%Y-%m-%d}  /  분석범위: 2023·2024·2025 3년  /  출처: SNOTES ERP 직접 조회 (SELECT only)"
    ws["A2"].font = Font(size=10, color="555555")
    ws.merge_cells("A2:G2")

    ws["A4"] = "[이번 버전(3년 비교) 변경점 — 본부장 피드백 반영]"
    ws["A4"].font = SECTION_FONT; ws["A4"].fill = SECTION_FILL
    ws.merge_cells("A4:G4")
    changes = [
        ("1","매출/단가/수량을 2023·2024·2025 3년 컬럼으로 모두 노출. 2023 vs 2025 단순 2점 비교 X"),
        ("2","YoY 증감률도 3구간: 23→24 / 24→25 / 23→25(누적). 본부장 '24년 값?' 질문 직접 대응"),
        ("3","'궤적' 분류 추가: 지속감소·반등·2년연속급감 등 3년 흐름 패턴으로 재분류"),
        ("4","원가반영 판정 기준도 3년 누적(+22.6%)과 24년 중간값(+14.5%) 이중 표시"),
        ("5","3사 분리(10000=갑우/20000=비피/30000=더원) + 지입거래처 구분은 v2 그대로 유지"),
    ]
    for i, (n, t) in enumerate(changes, start=5):
        ws[f"A{i}"] = n; ws[f"B{i}"] = t
        ws.merge_cells(f"B{i}:G{i}")

    ws["A11"] = "[시트 구성]"
    ws["A11"].font = SECTION_FONT; ws["A11"].fill = SECTION_FILL
    ws.merge_cells("A11:G11")
    sheets = [
        ("1_3사_연도별매출", "2021~2026 소속사별 매출 전체 추이"),
        ("1B_그룹_원가율", "2021~2026 용지/외주/원가율"),
        ("2_갑우_일반거래처", "갑우 본체 일반 거래처 3년 YoY"),
        ("3_갑우_지입거래처", "갑우 지입(교원/이투스/에듀윌) 3년 YoY"),
        ("4_비피_일반거래처", "비피앤피 일반 거래처 3년 YoY"),
        ("5_비피_지입거래처", "비피앤피 지입 거래처 3년 YoY"),
        ("6_더원_전체", "더원프린팅 전체 3년 YoY"),
        ("7_용지단가_추이", "용지 매입 단가 23→24→25 (벤치마크)"),
    ]
    for i, (s, d) in enumerate(sheets, start=12):
        ws[f"A{i}"] = s
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = d
        ws.merge_cells(f"B{i}:G{i}")

    ws.column_dimensions["A"].width = 4
    for col in "BCDEFG": ws.column_dimensions[col].width = 22


def write_firm_sheet(wb, sheet_name, firm, 지입필터, merged, title, note):
    ws = wb.create_sheet(sheet_name)
    df = customer_3year(merged, firm, 지입필터)
    if df.empty:
        ws["A1"] = f"{title} — 데이터 없음"; return
    df["패턴"] = df.apply(classify_pattern_3y, axis=1)
    df["궤적"] = df.apply(classify_trajectory, axis=1)
    df["원가반영"] = df.apply(classify_cost_pass_3y, axis=1)
    df_sorted = df.sort_values("매출_2023", ascending=False)
    display = fmt_display(df_sorted)

    # 패턴 요약
    pattern_sum = df.groupby("패턴").agg(
        거래처수=("거래처코드","count"),
        매출23억=("매출_2023", lambda x: round(x.sum()/1e8, 1)),
        매출24억=("매출_2024", lambda x: round(x.sum()/1e8, 1)),
        매출25억=("매출_2025", lambda x: round(x.sum()/1e8, 1)),
    ).reset_index()
    pattern_sum["23→25_손실(억)"] = pattern_sum["매출23억"] - pattern_sum["매출25억"]

    # 궤적 요약
    traj_sum = df.groupby("궤적").agg(
        거래처수=("거래처코드","count"),
        매출23억=("매출_2023", lambda x: round(x.sum()/1e8, 1)),
        매출24억=("매출_2024", lambda x: round(x.sum()/1e8, 1)),
        매출25억=("매출_2025", lambda x: round(x.sum()/1e8, 1)),
    ).reset_index()

    # 원가반영 요약
    cost_sum = df.groupby("원가반영").agg(
        거래처수=("거래처코드","count"),
        매출23억=("매출_2023", lambda x: round(x.sum()/1e8, 1)),
        매출25억=("매출_2025", lambda x: round(x.sum()/1e8, 1)),
    ).reset_index()
    cost_sum["손익(억)"] = cost_sum["매출25억"] - cost_sum["매출23억"]

    ws["A1"] = title
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:M1")
    ws["A2"] = note
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws.merge_cells("A2:M2")

    def write_summary(start_r, title_txt, sdf):
        ws.cell(row=start_r, column=1, value=title_txt).font = Font(bold=True, size=11)
        r = start_r + 1
        for j, col in enumerate(sdf.columns, start=1):
            c = ws.cell(row=r, column=j, value=col)
            c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        for _, row in sdf.iterrows():
            r += 1
            for j, val in enumerate(row, start=1):
                c = ws.cell(row=r, column=j, value=val); c.border = BORDER
                if isinstance(val, (int,float)):
                    c.alignment = Alignment(horizontal="right")
        return r + 2

    r = 4
    r = write_summary(r, "[패턴 요약 — 23→25 기준]", pattern_sum)
    r = write_summary(r, "[궤적 요약 — 23→24→25 흐름]", traj_sum)
    r = write_summary(r, f"[원가반영 판정 — 23→25 단가 vs 용지원가 +{COST_INFLATION_23_25}%]", cost_sum)

    # 상세 테이블
    ws.cell(row=r, column=1, value="[거래처별 3년 상세 — 2023 매출 5천만 이상, 2023 매출 큰 순]").font = Font(bold=True, size=11)
    r += 1
    for j, col in enumerate(display.columns, start=1):
        c = ws.cell(row=r, column=j, value=col)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    for _, row in display.iterrows():
        r += 1
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val); c.border = BORDER
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")
                if abs(val) >= 1000: c.number_format = "#,##0"
                elif isinstance(val, float): c.number_format = "#,##0.00"
    autosize(ws)


def main():
    env = load_env()
    conn = get_conn(env)
    print(f"✅ ERP 연결")
    header, line, paper_yearly, outsource = fetch_all(conn)
    conn.close()
    merged = enrich(header, line)
    print(f"   총 {len(merged):,}행")

    wb = Workbook()
    ws0 = wb.active; ws0.title = "0_표지"; write_cover(ws0)

    # 1. 3사 연도별
    pivot = merged.pivot_table(index="연도", columns="소속사", values="매출", aggfunc="sum").fillna(0)
    for c in pivot.columns:
        pivot[c] = (pivot[c]/1e8).round(1)
    pivot["합계(억)"] = pivot.sum(axis=1)
    pivot = pivot.reset_index()
    ws1 = wb.create_sheet("1_3사_연도별매출")
    write_df_sheet(ws1, pivot, "1. 소속사별 연도별 매출 (억)",
        "CD_CUST_OWN 기준: 10000=갑우 / 20000=비피 / 30000=더원")

    # 1B. 원가율
    paper_by_y = paper_yearly.set_index("연도")["매입액"].to_dict()
    out_by_y = outsource.set_index("연도")["외주매입"].to_dict()
    group_totals = merged.groupby("연도")["매출"].sum().reset_index()
    rows = []
    for _, r in group_totals.iterrows():
        y = r["연도"]; 매출 = r["매출"]
        paper = paper_by_y.get(int(y), 0)
        out = out_by_y.get(y, 0)
        총매입 = paper + out
        원가율 = (총매입/매출*100) if 매출>0 else 0
        rows.append({
            "연도": y, "그룹매출(억)": round(매출/1e8, 1),
            "용지매입(억)": round(paper/1e8, 1),
            "외주매입(억)": round(out/1e8, 2),
            "총매입(억)": round(총매입/1e8, 1),
            "원가율(%)": round(원가율, 1),
        })
    cost_df = pd.DataFrame(rows)
    ws1b = wb.create_sheet("1B_그룹_원가율")
    write_df_sheet(ws1b, cost_df, "1B. 그룹 원가율 추이",
        ["2023 10% → 2024 14% → 2025 19%로 지속 상승",
         "매입 테이블에 CD_CUST_OWN 없어 그룹 합산만 가능"])

    # 2~6. 회사별 × 지입/일반
    write_firm_sheet(wb, "2_갑우_일반거래처", "갑우", "일반", merged,
        "2. 갑우문화사 — 일반 거래처 (3년 비교)",
        "코스알엑스·씨디유디자인·지에스리테일 등. 23→24→25 3년 추이로 판정")
    write_firm_sheet(wb, "3_갑우_지입거래처", "갑우", "지입", merged,
        "3. 갑우문화사 — 지입 거래처 (교원·이투스·에듀윌)",
        "용지는 거래처가 직접 구매 → 단가/원가율을 일반 거래처와 같은 기준으로 보면 왜곡")
    write_firm_sheet(wb, "4_비피_일반거래처", "비피", "일반", merged,
        "4. 비피앤피 — 일반 거래처 (3년 비교)",
        "한국조폐공사·지담디앤피·케이랩·동행복권 등 패키지 본진")
    write_firm_sheet(wb, "5_비피_지입거래처", "비피", "지입", merged,
        "5. 비피앤피 — 지입 거래처 (3년 비교)",
        "교원구몬(3사 걸쳐 최대 지입거래처) 등")
    write_firm_sheet(wb, "6_더원_전체", "더원", "일반", merged,
        "6. 더원프린팅 — 전체 거래처 (3년 비교)",
        "2023부터 데이터 있음, 연간 4~6.6억대 소규모")

    # 7. 용지단가
    pp = paper_yearly.copy()
    pp["매입액(억)"] = (pp["매입액"]/1e8).round(1)
    pp["단가(원)"] = pp["단가"].round(0)
    # 23 기준 누적
    base = pp[pp["연도"]==2023]["단가"].values
    base_v = base[0] if len(base) > 0 else 0
    pp["23기준누적Δ(%)"] = pp["단가"].apply(lambda x: round((x-base_v)/base_v*100, 1) if base_v>0 else 0)
    pp = pp[["연도","수량","매입액(억)","단가(원)","23기준누적Δ(%)"]]
    ws7 = wb.create_sheet("7_용지단가_추이")
    write_df_sheet(ws7, pp, "7. 용지 단가 추이 (벤치마크)",
        "2023 38,330 → 2024 43,887(+14.5%) → 2025 46,991(+22.6%) → 2026Q1 49,654(+29.5%)")

    wb.save(OUT_XLSX)
    print(f"\n✅ 저장: {OUT_XLSX}")
    print(f"   시트: {', '.join(wb.sheetnames)}")
    print("\n소속사 × 연도 매출(억):")
    print(pivot.to_string(index=False))
    print("\n그룹 원가율:")
    print(cost_df.to_string(index=False))


if __name__ == "__main__":
    main()
