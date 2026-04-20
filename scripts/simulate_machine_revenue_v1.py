"""
인쇄기별 매출 시뮬레이션 V1.

V0 → V1 개선 사항:
  1. ERP 매출 라인을 3사 모두 로드 (CD_CUST_OWN 필터 제거)
     이유: 교원 학습교재가 비피앤피(OWN=20000)에 등록되어 있어 갑우만 필터 시 누락
  2. 거래처 별칭 매핑 (교원 → (주)교원구몬, 교원턴키 → ?, 이투스에듀 → 이투스 등)
  3. 매칭 단계 분리:
     a. 강한 매칭: 거래처+제품명(슬래시 분리 토큰 매칭)
     b. 통수 비율 분할: 한 매출 라인이 N개 작업과 매칭되면 통수 비율로 분배
  4. 월별 시계열: 매출 라인 DT_SALES(YYYYMM) 기준 호기 × 월 매트릭스
"""
from pathlib import Path
import json
import re
import urllib.request
import openpyxl
import pymssql

ENV = Path("/Users/jack/dev/gabwoo/견적계산기/.env.local")
OUT = Path(__file__).parent / "output" / "machine_revenue_v1.txt"
OUT_JSON = Path(__file__).parent / "output" / "machine_revenue_v1.json"
JAEDAN_2025 = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/재단지시서_2025년.xlsx")

SUPABASE_URL = "https://btbqzbrtsmwoolurpqgx.supabase.co"
SUPABASE_SR = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJ0YnF6YnJ0c213b29sdXJwcWd4Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTQ0MzgwNSwiZXhwIjoyMDkxMDE5ODA1fQ.M6H48_EfhK_GbNB8LC557VukouFaYdXXTjSD8S5azqw"
DATA_LOCAL = Path("/tmp/gw_check/data.json")

# ─── 거래처 별칭 매핑 (재단지시서 → ERP 정식명) ────────────────────
CLIENT_ALIAS = {
    "교원": ["(주)교원구몬", "(주)교원프라퍼티 인천공장", "(주)교원"],
    "교원구몬": ["(주)교원구몬"],
    "교원턴키": ["(주)교원구몬", "(주)교원"],  # 턴키도 교원 학습교재 라인으로 추정
    "이투스에듀": ["(주)이투스에듀"],
    "한국조폐공사": ["한국조폐공사"],
    "에듀윌": ["(주)에듀윌"],
}

# ─── 교원 정형 패턴 매칭 키워드 ────────────────────────────────────
KYOWON_CATEGORIES = ["정교재", "전집", "구몬", "성장노트", "한자", "수학", "일어", "중국어"]


def load_env():
    env = {}
    for line in ENV.read_text().splitlines():
        line = line.strip()
        if line and "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def connect():
    env = load_env()
    return pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=10,
    )


class Tee:
    def __init__(self, path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.f = open(path, "w", encoding="utf-8")

    def __call__(self, *args):
        msg = " ".join(str(a) for a in args)
        print(msg)
        self.f.write(msg + "\n")

    def close(self):
        self.f.close()


def section(log, title):
    log("\n" + "=" * 80)
    log(title)
    log("=" * 80)


# ─── 정규화 ────────────────────────────────────────────────────────
def normalize_client(s):
    if not s:
        return ""
    s = re.sub(r"\(주\)|주식회사|㈜|\(유\)|유한회사", "", str(s))
    s = re.sub(r"\s+", "", s)
    return s.lower()


def normalize_product(s):
    if not s:
        return ""
    s = re.sub(r"[\s/_,()\[\]·\-]+", "", str(s)).lower()
    return s


def split_tokens(s):
    if not s:
        return set()
    return set(t for t in re.split(r"[\s/_,()\[\]·\-]+", str(s)) if len(t) >= 2)


# ─── 데이터 로드 ───────────────────────────────────────────────────
def fetch_supabase_data():
    DATA_LOCAL.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(
        f"{SUPABASE_URL}/storage/v1/object/gabwoo-data/data.json",
        headers={"apikey": SUPABASE_SR, "Authorization": f"Bearer {SUPABASE_SR}"},
    )
    with urllib.request.urlopen(req) as r:
        DATA_LOCAL.write_bytes(r.read())
    return json.loads(DATA_LOCAL.read_text())


def parse_jaedan_xlsx(path):
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if "재단지시서" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["재단지시서"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header_idx = None
    for i, r in enumerate(rows[:5]):
        if r and any(c == "설비" for c in r if c):
            header_idx = i
            break
    if header_idx is None:
        return []
    header = rows[header_idx]
    col = {name: idx for idx, name in enumerate(header) if name}
    out = []
    current_equip = None
    for r in rows[header_idx + 1:]:
        if not r or len(r) <= max(col.values()):
            continue
        eq = r[col["설비"]]
        client = r[col.get("거래처", -1)] if "거래처" in col else None
        product = r[col.get("제품명", -1)] if "제품명" in col else None
        cnt = r[col.get("통수", -1)] if "통수" in col else None
        if eq and str(eq).strip() not in ("-", "", "설비"):
            current_equip = str(eq).strip()
        if not (client and product) or str(client).strip() in ("-", "", "거래처"):
            continue
        try:
            cnt_int = int(float(cnt)) if cnt and str(cnt).replace(".", "").replace("-", "").isdigit() else None
        except (ValueError, TypeError):
            cnt_int = None
        out.append({
            "equipment": current_equip,
            "client": str(client).strip(),
            "product": str(product).strip(),
            "cuts": cnt_int,
            "source": str(path.name),
        })
    return out


def load_all_jobs(log):
    log("\n[a] Supabase data.json 다운로드…")
    sdata = fetch_supabase_data()
    log(f"   updated_at={sdata.get('updated_at')}, all_completed={len(sdata.get('all_completed', []))}건")

    log("\n[b] 2025 재단지시서 엑셀…")
    jobs_2025 = parse_jaedan_xlsx(JAEDAN_2025)
    log(f"   파싱 {len(jobs_2025)}건")

    jobs = []
    for j in sdata.get("all_completed", []):
        eq = j.get("equipment")
        if eq and str(eq).strip() not in ("-", "", "설비"):
            jobs.append({
                "equipment": str(eq).strip(),
                "client": j.get("client"),
                "product": j.get("product"),
                "cuts": j.get("cuts"),
                "source": "2026_supabase",
            })
    jobs.extend(jobs_2025)
    return jobs


def load_erp_lines(log):
    """3사 모두 + 2024-2026 매출 라인."""
    conn = connect()
    cur = conn.cursor(as_dict=True)
    cur.execute("""
        SELECT h.NO_SALES, h.DT_SALES, h.CD_CUST, h.CD_CUST_OWN,
               c.NM_CUST,
               l.CD_ITEM, i.NM_ITEM,
               CAST(l.AM AS BIGINT) am
        FROM SAL_SALESL l
        INNER JOIN SAL_SALESH h ON h.NO_SALES = l.NO_SALES
        LEFT JOIN MAS_CUST c ON c.CD_CUST = h.CD_CUST
        LEFT JOIN PRT_ITEM i ON i.CD_ITEM = l.CD_ITEM
        WHERE (h.DT_SALES LIKE '2024%' OR h.DT_SALES LIKE '2025%' OR h.DT_SALES LIKE '2026%')
          AND h.ST_SALES='Y' AND h.AM > 0
    """)
    lines = cur.fetchall()
    conn.close()
    log(f"   ERP 매출 라인 (3사, 2024-2026): {len(lines):,}건 / 합계 {sum(int(l['am'] or 0) for l in lines):,}원")
    return lines


# ─── 매칭 ──────────────────────────────────────────────────────────
def build_erp_index(erp_lines):
    """ERP 매출 라인 → 거래처(정규화) 인덱스 + 거래처별 라인 리스트."""
    idx = {}
    for l in erp_lines:
        nc = normalize_client(l["NM_CUST"])
        if nc:
            idx.setdefault(nc, []).append(l)
    return idx


def find_candidate_lines(job, erp_index):
    """작업 행 → ERP 후보 라인 리스트 (별칭 + 부분매칭)."""
    raw_client = job["client"]
    nc = normalize_client(raw_client)
    candidates = []
    # 1) 별칭에 등록된 거래처
    if raw_client and raw_client.strip() in CLIENT_ALIAS:
        for alias_name in CLIENT_ALIAS[raw_client.strip()]:
            ac = normalize_client(alias_name)
            candidates.extend(erp_index.get(ac, []))
    # 2) 정규화 동일
    if nc and nc in erp_index:
        candidates.extend(erp_index[nc])
    # 3) 부분 매칭 (양방향)
    if not candidates and nc:
        for ec in erp_index:
            if (nc in ec or ec in nc) and len(nc) >= 2:
                candidates.extend(erp_index[ec])
                if len(candidates) > 200:
                    break
    return candidates


def product_matches(job_product, erp_item):
    """제품명 매칭 — 토큰 기반 + 부분 문자열."""
    if not job_product or not erp_item:
        return 0.0
    np = normalize_product(job_product)
    ep = normalize_product(erp_item)
    if len(np) < 4 or len(ep) < 4:
        return 0.0
    # 강한 매칭: 부분 문자열 양방향
    if np in ep or ep in np:
        return 1.0
    # 토큰 기반 (자카드 유사도)
    jt = split_tokens(job_product)
    et = split_tokens(erp_item)
    if not jt or not et:
        return 0.0
    inter = jt & et
    if not inter:
        return 0.0
    jaccard = len(inter) / len(jt | et)
    # 교원 카테고리 토큰 우선
    if any(k in inter for k in KYOWON_CATEGORIES):
        return max(jaccard, 0.5)
    return jaccard


def main():
    log = Tee(OUT)

    # ─── 1. 데이터 로드 ────────────────────────────────────────────
    section(log, "1. 데이터 로드")
    jobs = load_all_jobs(log)
    log(f"\n   통합 작업 행: {len(jobs):,}건")
    erp_lines = load_erp_lines(log)
    erp_index = build_erp_index(erp_lines)
    log(f"   ERP 거래처 (정규화): {len(erp_index):,}개")

    # ─── 2. 매칭 (작업 → 매출 라인 리스트) ──────────────────────────
    section(log, "2. 매칭 — 작업 행 → 후보 매출 라인 (강한 매칭만)")
    job_to_lines = {}  # job_idx → [(line, score)]
    line_to_jobs = {}  # NO_SALES → [job_idx]

    for ji, job in enumerate(jobs):
        candidates = find_candidate_lines(job, erp_index)
        best_score = 0.0
        best_line = None
        for c in candidates:
            score = product_matches(job["product"], c["NM_ITEM"])
            if score > best_score and score >= 0.5:
                best_score = score
                best_line = c
        if best_line:
            job_to_lines[ji] = (best_line, best_score)
            line_to_jobs.setdefault(best_line["NO_SALES"], []).append(ji)

    log(f"\n   매칭 성공: {len(job_to_lines):,} / {len(jobs):,} ({len(job_to_lines) / len(jobs) * 100:.1f}%)")
    log(f"   매칭에 사용된 매출 라인: {len(line_to_jobs):,}건")
    log(f"   매출 라인 1건당 평균 매칭 작업: {len(job_to_lines) / max(len(line_to_jobs), 1):.1f}개")

    # 거래처별 매칭율
    log("\n   거래처별 매칭율 (Top 15):")
    by_cust = {}
    for ji, job in enumerate(jobs):
        c = job["client"] or "(unknown)"
        by_cust.setdefault(c, [0, 0])[0] += 1
        if ji in job_to_lines:
            by_cust[c][1] += 1
    for c, (total, matched) in sorted(by_cust.items(), key=lambda x: -x[1][0])[:15]:
        log(f"      {c[:25]:<25} 작업={total:>6,} 매칭={matched:>6,} ({matched / total * 100:>5.1f}%)")

    # ─── 3. 통수 비율 분할 ─────────────────────────────────────────
    section(log, "3. 통수 비율로 매출 분할")
    split_revenue = {}  # ji → (line, allocated_am)

    def to_int_cuts(v):
        if v is None:
            return 0
        try:
            return int(float(str(v).replace(",", "")))
        except (ValueError, TypeError):
            return 0

    for no_sales, job_indexes in line_to_jobs.items():
        line = job_to_lines[job_indexes[0]][0]
        am = int(line["am"] or 0)
        cuts_list = [to_int_cuts(jobs[ji]["cuts"]) for ji in job_indexes]
        valid_cuts = [c for c in cuts_list if c > 0]
        avg_cut = sum(valid_cuts) / len(valid_cuts) if valid_cuts else 1
        normalized_cuts = [c if c > 0 else int(avg_cut) for c in cuts_list]
        total_cuts = sum(normalized_cuts) or len(job_indexes)

        for ji, c in zip(job_indexes, normalized_cuts):
            allocated = int(am * c / total_cuts)
            split_revenue[ji] = (line, allocated)

    total_allocated = sum(v[1] for v in split_revenue.values())
    log(f"\n   분할 후 총 매출: {total_allocated:>17,}원")
    log(f"   매칭 ERP 매출 합: {sum(int(line_to_jobs and job_to_lines[line_to_jobs[ns][0]][0]['am'] or 0) for ns in line_to_jobs):>17,}원 (참고)")
    matched_erp_sum = sum(int(job_to_lines[line_to_jobs[ns][0]][0]['am'] or 0) for ns in line_to_jobs)
    log(f"   매칭 ERP 매출 합 (정확): {matched_erp_sum:>17,}원")
    log(f"   분할/매칭 비율: {total_allocated / max(matched_erp_sum, 1) * 100:.1f}% (100%면 손실 없음)")

    # ─── 4. 호기별 합계 ────────────────────────────────────────────
    section(log, "4. 호기별 매출 합계 (V1 — 통수 비율 분할 후)")
    by_eq_revenue = {}
    by_eq_jobs = {}
    for ji, (line, am) in split_revenue.items():
        eq = jobs[ji]["equipment"] or "(unknown)"
        by_eq_revenue[eq] = by_eq_revenue.get(eq, 0) + am
        by_eq_jobs[eq] = by_eq_jobs.get(eq, 0) + 1

    log(f"\n   {'호기':<15}{'매칭 작업':>12}{'추정 매출':>20}{'평균/작업':>15}")
    for eq in sorted(by_eq_revenue.keys(), key=lambda x: -by_eq_revenue[x]):
        avg = by_eq_revenue[eq] / max(by_eq_jobs[eq], 1)
        log(f"   {eq:<15}{by_eq_jobs.get(eq, 0):>12,}{by_eq_revenue[eq]:>20,}{int(avg):>15,}")
    log(f"   {'-' * 62}")
    log(f"   {'합계':<15}{sum(by_eq_jobs.values()):>12,}{sum(by_eq_revenue.values()):>20,}")

    # ─── 5. 호기 × 월별 매출 (매출 라인 DT_SALES 기준) ─────────────
    section(log, "5. 호기 × 월별 매출 (매출 라인 DT_SALES 기준 YYYYMM)")
    matrix = {}  # equip → {ym: amount}
    months = set()
    for ji, (line, am) in split_revenue.items():
        eq = jobs[ji]["equipment"] or "(unknown)"
        dt = str(line.get("DT_SALES") or "")
        ym = dt[:6] if len(dt) >= 6 else "unknown"
        months.add(ym)
        matrix.setdefault(eq, {})[ym] = matrix.setdefault(eq, {}).get(ym, 0) + am

    sorted_months = sorted(m for m in months if m != "unknown")
    sorted_eqs = sorted(matrix.keys(), key=lambda e: -sum(matrix[e].values()))

    # 헤더
    header = f"   {'호기':<10}" + "".join(f"{m:>10}" for m in sorted_months) + f"{'합계':>15}"
    log(f"\n{header}")
    for eq in sorted_eqs:
        row = f"   {eq:<10}"
        for m in sorted_months:
            v = matrix[eq].get(m, 0)
            row += f"{v / 1_000_000:>10.0f}" if v else f"{'.':>10}"
        row += f"{sum(matrix[eq].values()) / 1_000_000:>15,.0f}"
        log(row)
    log(f"   (단위: 백만 원)")

    # JSON 저장 (대시보드 연동용)
    out_data = {
        "by_equipment_total": {eq: by_eq_revenue[eq] for eq in sorted(by_eq_revenue.keys(), key=lambda x: -by_eq_revenue[x])},
        "by_equipment_jobs": by_eq_jobs,
        "by_equipment_month": matrix,
        "match_stats": {
            "total_jobs": len(jobs),
            "matched_jobs": len(job_to_lines),
            "match_rate": len(job_to_lines) / len(jobs),
            "matched_lines": len(line_to_jobs),
            "total_allocated_revenue": total_allocated,
        },
    }
    OUT_JSON.write_text(json.dumps(out_data, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"\n→ JSON 저장: {OUT_JSON}")

    log.close()


if __name__ == "__main__":
    main()
