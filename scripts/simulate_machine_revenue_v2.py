"""
인쇄기별 매출 시뮬레이션 V2.

V1 → V2 개선:
  1. 진짜 통수 = `quantity` 사용 (V1 cuts는 대시 비어있음)
  2. 인쇄 완료일 = `completed_date` (Supabase). V1은 ERP DT_SALES로 우회했음
  3. 출판 거래처 화이트리스트 = 재단지시서에 1번이라도 등장한 모든 거래처
     → 출판 매출만 분모로 잡아 V1 커버율 재계산
  4. 호기 × 월별 매트릭스를 인쇄 완료일 기준으로
"""
from pathlib import Path
import json
import re
import urllib.request
import openpyxl
import pymssql
from collections import Counter

ENV = Path("/Users/jack/dev/gabwoo/견적계산기/.env.local")
OUT = Path(__file__).parent / "output" / "machine_revenue_v2.txt"
OUT_JSON = Path(__file__).parent / "output" / "machine_revenue_v2.json"
JAEDAN_2025 = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/재단지시서_2025년.xlsx")

SUPABASE_URL = "https://btbqzbrtsmwoolurpqgx.supabase.co"
SUPABASE_SR = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJ0YnF6YnJ0c213b29sdXJwcWd4Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTQ0MzgwNSwiZXhwIjoyMDkxMDE5ODA1fQ.M6H48_EfhK_GbNB8LC557VukouFaYdXXTjSD8S5azqw"
DATA_LOCAL = Path("/tmp/gw_check/data.json")

CLIENT_ALIAS = {
    "교원": ["(주)교원구몬", "(주)교원프라퍼티 인천공장", "(주)교원"],
    "교원구몬": ["(주)교원구몬"],
    "교원턴키": ["(주)교원구몬", "(주)교원"],
    "이투스에듀": ["(주)이투스에듀"],
    "한국조폐공사": ["한국조폐공사"],
    "에듀윌": ["(주)에듀윌"],
}
KYOWON_CATEGORIES = ["정교재", "전집", "구몬", "성장노트", "한자", "수학", "일어", "중국어"]


def load_env():
    env = {}
    for line in ENV.read_text().splitlines():
        line = line.strip()
        if line and "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def connect():
    env = load_env()
    return pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=10,
    )


class Tee:
    def __init__(self, path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.f = open(path, "w", encoding="utf-8")

    def __call__(self, *args):
        msg = " ".join(str(a) for a in args)
        print(msg)
        self.f.write(msg + "\n")

    def close(self):
        self.f.close()


def section(log, title):
    log("\n" + "=" * 80)
    log(title)
    log("=" * 80)


def normalize_client(s):
    if not s:
        return ""
    s = re.sub(r"\(주\)|주식회사|㈜|\(유\)|유한회사", "", str(s))
    s = re.sub(r"\s+", "", s)
    return s.lower()


def normalize_product(s):
    if not s:
        return ""
    return re.sub(r"[\s/_,()\[\]·\-]+", "", str(s)).lower()


def split_tokens(s):
    if not s:
        return set()
    return set(t for t in re.split(r"[\s/_,()\[\]·\-]+", str(s)) if len(t) >= 2)


def to_int(v):
    if v is None:
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def fetch_supabase_data():
    DATA_LOCAL.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(
        f"{SUPABASE_URL}/storage/v1/object/gabwoo-data/data.json",
        headers={"apikey": SUPABASE_SR, "Authorization": f"Bearer {SUPABASE_SR}"},
    )
    with urllib.request.urlopen(req) as r:
        DATA_LOCAL.write_bytes(r.read())
    return json.loads(DATA_LOCAL.read_text())


def parse_jaedan_xlsx(path):
    """2025 엑셀: header row3 [설비, 거래처, 제품명, ..., 통수, 판수, ..., 후가공, 납기]
    quantity가 없으니 통수 컬럼만 사용."""
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if "재단지시서" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["재단지시서"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header_idx = None
    for i, r in enumerate(rows[:5]):
        if r and any(c == "설비" for c in r if c):
            header_idx = i
            break
    if header_idx is None:
        return []
    header = rows[header_idx]
    col = {n: i for i, n in enumerate(header) if n}
    out = []
    current_eq = None
    for r in rows[header_idx + 1:]:
        if not r or len(r) <= max(col.values()):
            continue
        eq = r[col["설비"]]
        client = r[col.get("거래처", -1)] if "거래처" in col else None
        product = r[col.get("제품명", -1)] if "제품명" in col else None
        cnt = r[col.get("통수", -1)] if "통수" in col else None
        post = r[col.get("후가공", -1)] if "후가공" in col else None
        if eq and str(eq).strip() not in ("-", "", "설비"):
            current_eq = str(eq).strip()
        if not (client and product) or str(client).strip() in ("-", "", "거래처"):
            continue
        out.append({
            "equipment": current_eq,
            "client": str(client).strip(),
            "product": str(product).strip(),
            "quantity": to_int(cnt),
            "post_process": str(post).strip() if post else None,
            "completed_date": None,  # 2025 엑셀에는 행별 일자 없음
            "source": "2025_xlsx",
        })
    return out


def load_all_jobs(log):
    log("\n[a] Supabase data.json…")
    sdata = fetch_supabase_data()
    jobs = []
    for j in sdata.get("all_completed", []):
        eq = j.get("equipment")
        if eq and str(eq).strip() not in ("-", "", "설비"):
            jobs.append({
                "equipment": str(eq).strip(),
                "client": j.get("client"),
                "product": j.get("product"),
                "quantity": to_int(j.get("quantity")),  # ⭐ 진짜 통수
                "post_process": j.get("post_process"),
                "completed_date": j.get("completed_date"),  # ⭐ 인쇄 완료일
                "source": "2026_supabase",
            })
    log(f"   2026 supabase: {len(jobs):,}건 (quantity·completed_date 보유)")

    log("\n[b] 2025 엑셀…")
    jobs_2025 = parse_jaedan_xlsx(JAEDAN_2025)
    log(f"   2025 xlsx: {len(jobs_2025):,}건 (통수만, 일자 없음)")
    jobs.extend(jobs_2025)
    return jobs


def load_erp_lines(log):
    conn = connect()
    cur = conn.cursor(as_dict=True)
    cur.execute("""
        SELECT h.NO_SALES, h.DT_SALES, h.CD_CUST, h.CD_CUST_OWN,
               c.NM_CUST, l.CD_ITEM, i.NM_ITEM,
               CAST(l.AM AS BIGINT) am
        FROM SAL_SALESL l
        INNER JOIN SAL_SALESH h ON h.NO_SALES = l.NO_SALES
        LEFT JOIN MAS_CUST c ON c.CD_CUST = h.CD_CUST
        LEFT JOIN PRT_ITEM i ON i.CD_ITEM = l.CD_ITEM
        WHERE (h.DT_SALES LIKE '2024%' OR h.DT_SALES LIKE '2025%' OR h.DT_SALES LIKE '2026%')
          AND h.ST_SALES='Y' AND h.AM > 0
    """)
    lines = cur.fetchall()
    conn.close()
    log(f"   ERP 매출 라인 (3사, 2024-2026): {len(lines):,}건 / 합계 {sum(int(l['am'] or 0) for l in lines):,}원")
    return lines


def build_erp_index(erp_lines):
    idx = {}
    for l in erp_lines:
        nc = normalize_client(l["NM_CUST"])
        if nc:
            idx.setdefault(nc, []).append(l)
    return idx


def find_candidate_lines(job, erp_index):
    raw_client = job["client"]
    nc = normalize_client(raw_client)
    candidates = []
    if raw_client and raw_client.strip() in CLIENT_ALIAS:
        for alias_name in CLIENT_ALIAS[raw_client.strip()]:
            ac = normalize_client(alias_name)
            candidates.extend(erp_index.get(ac, []))
    if nc and nc in erp_index:
        candidates.extend(erp_index[nc])
    if not candidates and nc:
        for ec in erp_index:
            if (nc in ec or ec in nc) and len(nc) >= 2:
                candidates.extend(erp_index[ec])
                if len(candidates) > 200:
                    break
    return candidates


def product_matches(job_product, erp_item):
    if not job_product or not erp_item:
        return 0.0
    np = normalize_product(job_product)
    ep = normalize_product(erp_item)
    if len(np) < 4 or len(ep) < 4:
        return 0.0
    if np in ep or ep in np:
        return 1.0
    jt = split_tokens(job_product)
    et = split_tokens(erp_item)
    if not jt or not et:
        return 0.0
    inter = jt & et
    if not inter:
        return 0.0
    jaccard = len(inter) / len(jt | et)
    if any(k in inter for k in KYOWON_CATEGORIES):
        return max(jaccard, 0.5)
    return jaccard


def main():
    log = Tee(OUT)

    # ─── 1. 데이터 로드 ────────────────────────────────────────────
    section(log, "1. 데이터 로드")
    jobs = load_all_jobs(log)
    log(f"\n   통합 작업 행: {len(jobs):,}건")
    erp_lines = load_erp_lines(log)
    erp_index = build_erp_index(erp_lines)

    # ─── 2. 출판 거래처 화이트리스트 = 재단지시서 등장 거래처 ──────
    section(log, "2. 출판 거래처 화이트리스트 (재단지시서 등장 거래처 전부)")
    pub_clients_raw = Counter(j["client"] for j in jobs if j["client"])
    log(f"   재단지시서 등장 distinct 거래처: {len(pub_clients_raw):,}개")

    # 정규화 화이트리스트 + 별칭
    pub_clients_norm = set()
    for c in pub_clients_raw:
        pub_clients_norm.add(normalize_client(c))
        if c.strip() in CLIENT_ALIAS:
            for alias in CLIENT_ALIAS[c.strip()]:
                pub_clients_norm.add(normalize_client(alias))

    # ERP 거래처 중 매칭 (정확/부분)
    matched_erp_clients = set()
    for nc in pub_clients_norm:
        if nc in erp_index:
            matched_erp_clients.add(nc)
        # 부분 매칭
        for ec in erp_index:
            if (nc in ec or ec in nc) and len(nc) >= 2:
                matched_erp_clients.add(ec)

    log(f"   ERP 매출 라인 거래처 (정규화): {len(erp_index):,}개")
    log(f"   재단지시서 거래처와 매칭된 ERP 거래처: {len(matched_erp_clients):,}개")

    # 출판 매출 (= 출판 거래처 화이트리스트의 ERP 매출 합계)
    pub_lines = []
    for ec in matched_erp_clients:
        pub_lines.extend(erp_index[ec])
    pub_revenue_total = sum(int(l["am"] or 0) for l in pub_lines)
    erp_total = sum(int(l["am"] or 0) for l in erp_lines)
    log(f"\n   ERP 전체 매출 (3사, 2024-2026): {erp_total:>17,}원")
    log(f"   출판 매출 (화이트리스트 합): {pub_revenue_total:>17,}원 ({pub_revenue_total / erp_total * 100:.1f}%)")
    log(f"   출판 매출 라인: {len(pub_lines):,}건")

    # 연도별 출판 매출
    log(f"\n   연도별 출판 매출:")
    by_year = {}
    for l in pub_lines:
        y = str(l["DT_SALES"])[:4]
        by_year[y] = by_year.get(y, 0) + int(l["am"] or 0)
    for y in sorted(by_year):
        log(f"      {y}: {by_year[y]:>17,}원")

    # ─── 3. 매칭 (작업 → 매출 라인) — V1과 동일 로직 ────────────
    section(log, "3. 매칭 — 작업 행 → 매출 라인 (강한 매칭)")
    job_to_line = {}
    line_to_jobs = {}
    for ji, job in enumerate(jobs):
        candidates = find_candidate_lines(job, erp_index)
        best_score = 0.0
        best_line = None
        for c in candidates:
            score = product_matches(job["product"], c["NM_ITEM"])
            if score > best_score and score >= 0.5:
                best_score = score
                best_line = c
        if best_line:
            job_to_line[ji] = (best_line, best_score)
            line_to_jobs.setdefault(best_line["NO_SALES"], []).append(ji)

    matched_n = len(job_to_line)
    log(f"\n   매칭 성공: {matched_n:,} / {len(jobs):,} ({matched_n / len(jobs) * 100:.1f}%)")
    log(f"   매칭 사용 매출 라인: {len(line_to_jobs):,}건")

    # ─── 4. quantity(진짜 통수) 비율 분할 ───────────────────────────
    section(log, "4. quantity(진짜 통수) 비율로 매출 분할 — V1 cuts 대신")
    split_revenue = {}
    for no_sales, job_indexes in line_to_jobs.items():
        line = job_to_line[job_indexes[0]][0]
        am = int(line["am"] or 0)
        qts = [jobs[ji]["quantity"] for ji in job_indexes]
        valid_qts = [q for q in qts if q > 0]
        avg = sum(valid_qts) / len(valid_qts) if valid_qts else 1
        norm_qts = [q if q > 0 else int(avg) for q in qts]
        total = sum(norm_qts) or len(job_indexes)
        for ji, q in zip(job_indexes, norm_qts):
            split_revenue[ji] = (line, int(am * q / total))

    total_alloc = sum(v[1] for v in split_revenue.values())
    matched_erp_sum = sum(int(job_to_line[line_to_jobs[ns][0]][0]['am'] or 0) for ns in line_to_jobs)
    log(f"\n   분할 매출 합계: {total_alloc:>17,}원")
    log(f"   매칭 ERP 매출 합: {matched_erp_sum:>17,}원")
    log(f"   분할 무손실: {total_alloc / max(matched_erp_sum, 1) * 100:.1f}%")

    # ⭐ 진짜 커버율 = 분할 매출 / 출판 매출
    log(f"\n   ⭐ 출판 매출 대비 커버율: {total_alloc / max(pub_revenue_total, 1) * 100:.1f}%")
    log(f"      (V1은 전체 매출 대비 8.6%였음)")

    # ─── 5. 호기별 합계 ────────────────────────────────────────────
    section(log, "5. 호기별 추정 매출 (V2 — quantity 분할)")
    by_eq_rev = {}
    by_eq_jobs = {}
    by_eq_qty = {}
    for ji, (line, am) in split_revenue.items():
        eq = jobs[ji]["equipment"] or "(?)"
        by_eq_rev[eq] = by_eq_rev.get(eq, 0) + am
        by_eq_jobs[eq] = by_eq_jobs.get(eq, 0) + 1
        by_eq_qty[eq] = by_eq_qty.get(eq, 0) + jobs[ji]["quantity"]

    log(f"\n   {'호기':<10}{'매칭작업':>10}{'통수합':>14}{'추정매출':>18}{'평균/작업':>14}")
    for eq in sorted(by_eq_rev.keys(), key=lambda x: -by_eq_rev[x]):
        avg = by_eq_rev[eq] / max(by_eq_jobs[eq], 1)
        log(f"   {eq:<10}{by_eq_jobs[eq]:>10,}{by_eq_qty.get(eq, 0):>14,}{by_eq_rev[eq]:>18,}{int(avg):>14,}")
    log(f"   {'-' * 66}")
    log(f"   {'합계':<10}{sum(by_eq_jobs.values()):>10,}{sum(by_eq_qty.values()):>14,}{sum(by_eq_rev.values()):>18,}")

    # ─── 6. 호기 × 월별 (completed_date 기준) ──────────────────────
    section(log, "6. 호기 × 월별 매출 (completed_date 기준 — 2026 supabase 한정)")
    matrix = {}  # equip → {ym: amount}
    months = set()
    no_date_jobs = 0
    for ji, (line, am) in split_revenue.items():
        eq = jobs[ji]["equipment"] or "(?)"
        cd = jobs[ji].get("completed_date")
        if cd and len(str(cd)) >= 7:
            ym = str(cd)[:7].replace("-", "")  # YYYYMM
        else:
            no_date_jobs += 1
            ym = "no_date"
        months.add(ym)
        matrix.setdefault(eq, {})[ym] = matrix.setdefault(eq, {}).get(ym, 0) + am

    log(f"\n   completed_date 없는 작업(2025 엑셀): {no_date_jobs:,}건")

    sorted_months = sorted(m for m in months if m != "no_date")
    sorted_eqs = sorted(matrix.keys(), key=lambda e: -sum(v for k, v in matrix[e].items() if k != "no_date"))

    header = f"   {'호기':<10}" + "".join(f"{m:>10}" for m in sorted_months) + f"{'no_date':>12}{'합계(M원)':>14}"
    log(f"\n{header}")
    for eq in sorted_eqs:
        row = f"   {eq:<10}"
        for m in sorted_months:
            v = matrix[eq].get(m, 0)
            row += f"{v / 1_000_000:>10.0f}" if v else f"{'.':>10}"
        nv = matrix[eq].get("no_date", 0)
        row += f"{nv / 1_000_000:>12.0f}" if nv else f"{'.':>12}"
        total = sum(matrix[eq].values())
        row += f"{total / 1_000_000:>14,.0f}"
        log(row)
    log(f"   (단위: 백만 원)")

    # ─── 7. V1 vs V2 ─────────────────────────────────────────────
    section(log, "7. V1 vs V2 비교")
    v1_path = Path(__file__).parent / "output" / "machine_revenue_v1.json"
    if v1_path.exists():
        v1 = json.loads(v1_path.read_text())
        v1_total = sum(v1.get("by_equipment_total", {}).values())
        v1_rate = f"{v1['match_stats']['match_rate'] * 100:.1f}%"
        v2_rate = f"{matched_n / len(jobs) * 100:.1f}%"
        v1_cov = f"{v1_total / 52_802_522_891 * 100:.1f}%"
        v2_cov = f"{total_alloc / max(pub_revenue_total, 1) * 100:.1f}%"
        log(f"\n   {'지표':<25}{'V1':>18}{'V2':>18}")
        log(f"   {'매칭율':<25}{v1_rate:>18}{v2_rate:>18}")
        log(f"   {'매칭 작업':<25}{v1['match_stats']['matched_jobs']:>18,}{matched_n:>18,}")
        log(f"   {'분할 매출':<25}{v1_total:>18,}{total_alloc:>18,}")
        log(f"   {'분모':<25}{'ERP전체 528억':>18}{'출판 매출':>18}")
        log(f"   {'커버율':<25}{v1_cov:>18}{v2_cov:>18}")

    # JSON 저장
    out_data = {
        "by_equipment_total": {eq: by_eq_rev[eq] for eq in sorted(by_eq_rev.keys(), key=lambda x: -by_eq_rev[x])},
        "by_equipment_jobs": by_eq_jobs,
        "by_equipment_qty": by_eq_qty,
        "by_equipment_month": matrix,
        "publishing_revenue_total": pub_revenue_total,
        "match_stats": {
            "total_jobs": len(jobs),
            "matched_jobs": matched_n,
            "match_rate": matched_n / len(jobs),
            "matched_lines": len(line_to_jobs),
            "total_allocated_revenue": total_alloc,
            "publishing_coverage": total_alloc / max(pub_revenue_total, 1),
        },
    }
    OUT_JSON.write_text(json.dumps(out_data, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"\n→ JSON 저장: {OUT_JSON}")
    log.close()


if __name__ == "__main__":
    main()
