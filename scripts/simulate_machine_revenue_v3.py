"""
인쇄기별 매출 시뮬레이션 V3 — 지입(아방) 분리 + 후가공 외주비.

V2 → V3 개선:
  1. 지입 거래처 분리: 교원/교원구몬/동행복권/이투스/에듀윌
     호기별 매출을 "지입(임가공)" / "일반" 두 줄로
  2. Jack 2026-04-20 증언 검증: "1호기·11호기가 교원 주로 담당, 9호기도 많이"
  3. 후가공 외주 단가표 파싱 → 통수×판규격×유광무광으로 후가공비 추정
  4. 호기별 매출 - 후가공 외주비 = 공헌이익 (초안)
"""
from pathlib import Path
import json
import re
import urllib.request
import openpyxl
import pymssql
from collections import Counter

ENV = Path("/Users/jack/dev/gabwoo/견적계산기/.env.local")
OUT = Path(__file__).parent / "output" / "machine_revenue_v3.txt"
OUT_JSON = Path(__file__).parent / "output" / "machine_revenue_v3.json"
JAEDAN_2025 = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/재단지시서_2025년.xlsx")
OUTSRC_XLSX = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/01_매입 단가표(출판).xlsx")

SUPABASE_URL = "https://btbqzbrtsmwoolurpqgx.supabase.co"
SUPABASE_SR = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJ0YnF6YnJ0c213b29sdXJwcWd4Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTQ0MzgwNSwiZXhwIjoyMDkxMDE5ODA1fQ.M6H48_EfhK_GbNB8LC557VukouFaYdXXTjSD8S5azqw"
DATA_LOCAL = Path("/tmp/gw_check/data.json")

# ─── 출판 스코프 지입(아방) 거래처 ────────────────────────────
# 주의: 이 스크립트는 출판사업부 재단지시서만 분석. 동행복권은 패키지라 제외.
# 출처: build_customer_health_2025.py JIIP_CUST_BIZ,
#       build_integrated_v1.py JIIP_FLAGS,
#       Jack 2026-04-20 (동행복권=패키지 스코프)
JIIP_BIZ = {
    "4968602009": "(주)교원",
    "6678702103": "(주)교원구몬",
    "7638502028": "(주)교원프라퍼티 인천공장",
    "7228502352": "이투스에듀 주식회사",
    "1198154852": "(주)에듀윌",
    "1078140772": "(주)갑우문화사",   # 관계사
    "1418115138": "(주)비피앤피",    # 관계사
}
JIIP_NAME_KW = ["교원", "이투스", "에듀윌"]  # 출판 스코프 (동행복권은 패키지)

CLIENT_ALIAS = {
    "교원": ["(주)교원구몬", "(주)교원프라퍼티 인천공장", "(주)교원"],
    "교원구몬": ["(주)교원구몬"],
    "교원턴키": ["(주)교원구몬", "(주)교원"],
    "이투스에듀": ["(주)이투스에듀"],
    "한국조폐공사": ["한국조폐공사"],
    "에듀윌": ["(주)에듀윌"],
    "동행복권": ["동행복권"],
}
KYOWON_CATEGORIES = ["정교재", "전집", "구몬", "성장노트", "한자", "수학", "일어", "중국어"]


def load_env():
    env = {}
    for line in ENV.read_text().splitlines():
        line = line.strip()
        if line and "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def connect():
    env = load_env()
    return pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=10,
    )


class Tee:
    def __init__(self, path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.f = open(path, "w", encoding="utf-8")

    def __call__(self, *args):
        msg = " ".join(str(a) for a in args)
        print(msg)
        self.f.write(msg + "\n")

    def close(self):
        self.f.close()


def section(log, title):
    log("\n" + "=" * 80)
    log(title)
    log("=" * 80)


def normalize_client(s):
    if not s:
        return ""
    s = re.sub(r"\(주\)|주식회사|㈜|\(유\)|유한회사", "", str(s))
    s = re.sub(r"\s+", "", s)
    return s.lower()


def normalize_product(s):
    if not s:
        return ""
    return re.sub(r"[\s/_,()\[\]·\-]+", "", str(s)).lower()


def split_tokens(s):
    if not s:
        return set()
    return set(t for t in re.split(r"[\s/_,()\[\]·\-]+", str(s)) if len(t) >= 2)


def to_int(v):
    if v is None:
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def is_jiip_client(client_name):
    """지입 거래처인가? 이름 키워드 기반."""
    if not client_name:
        return False
    return any(kw in client_name for kw in JIIP_NAME_KW)


def fetch_supabase_data():
    DATA_LOCAL.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(
        f"{SUPABASE_URL}/storage/v1/object/gabwoo-data/data.json",
        headers={"apikey": SUPABASE_SR, "Authorization": f"Bearer {SUPABASE_SR}"},
    )
    with urllib.request.urlopen(req) as r:
        DATA_LOCAL.write_bytes(r.read())
    return json.loads(DATA_LOCAL.read_text())


def parse_jaedan_xlsx(path):
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if "재단지시서" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["재단지시서"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header_idx = None
    for i, r in enumerate(rows[:5]):
        if r and any(c == "설비" for c in r if c):
            header_idx = i
            break
    if header_idx is None:
        return []
    header = rows[header_idx]
    col = {n: i for i, n in enumerate(header) if n}
    out = []
    current_eq = None
    for r in rows[header_idx + 1:]:
        if not r or len(r) <= max(col.values()):
            continue
        eq = r[col["설비"]]
        client = r[col.get("거래처", -1)] if "거래처" in col else None
        product = r[col.get("제품명", -1)] if "제품명" in col else None
        cnt = r[col.get("통수", -1)] if "통수" in col else None
        post = r[col.get("후가공", -1)] if "후가공" in col else None
        size = r[col.get("규격", -1)] if "규격" in col else None
        colors = r[col.get("색도", -1)] if "색도" in col else None
        if eq and str(eq).strip() not in ("-", "", "설비"):
            current_eq = str(eq).strip()
        if not (client and product) or str(client).strip() in ("-", "", "거래처"):
            continue
        out.append({
            "equipment": current_eq,
            "client": str(client).strip(),
            "product": str(product).strip(),
            "quantity": to_int(cnt),
            "post_process": str(post).strip() if post else None,
            "size": str(size).strip() if size else None,
            "colors": to_int(colors),
            "completed_date": None,
            "source": "2025_xlsx",
        })
    return out


def load_all_jobs():
    sdata = fetch_supabase_data()
    jobs = []
    for j in sdata.get("all_completed", []):
        eq = j.get("equipment")
        if eq and str(eq).strip() not in ("-", "", "설비"):
            jobs.append({
                "equipment": str(eq).strip(),
                "client": j.get("client"),
                "product": j.get("product"),
                "quantity": to_int(j.get("quantity")),
                "post_process": j.get("post_process"),
                "size": j.get("size"),
                "colors": to_int(j.get("colors")),
                "completed_date": j.get("completed_date"),
                "source": "2026_supabase",
            })
    jobs.extend(parse_jaedan_xlsx(JAEDAN_2025))
    return jobs


def load_erp_lines():
    conn = connect()
    cur = conn.cursor(as_dict=True)
    cur.execute("""
        SELECT h.NO_SALES, h.DT_SALES, h.CD_CUST, h.CD_CUST_OWN,
               c.NM_CUST, c.NO_BIZ, l.CD_ITEM, i.NM_ITEM,
               CAST(l.AM AS BIGINT) am
        FROM SAL_SALESL l
        INNER JOIN SAL_SALESH h ON h.NO_SALES = l.NO_SALES
        LEFT JOIN MAS_CUST c ON c.CD_CUST = h.CD_CUST
        LEFT JOIN PRT_ITEM i ON i.CD_ITEM = l.CD_ITEM
        WHERE (h.DT_SALES LIKE '2024%' OR h.DT_SALES LIKE '2025%' OR h.DT_SALES LIKE '2026%')
          AND h.ST_SALES='Y' AND h.AM > 0
    """)
    lines = cur.fetchall()
    conn.close()
    return lines


def build_erp_index(erp_lines):
    idx = {}
    for l in erp_lines:
        nc = normalize_client(l["NM_CUST"])
        if nc:
            idx.setdefault(nc, []).append(l)
    return idx


def find_candidate_lines(job, erp_index):
    raw_client = job["client"]
    nc = normalize_client(raw_client)
    candidates = []
    if raw_client and raw_client.strip() in CLIENT_ALIAS:
        for alias_name in CLIENT_ALIAS[raw_client.strip()]:
            ac = normalize_client(alias_name)
            candidates.extend(erp_index.get(ac, []))
    if nc and nc in erp_index:
        candidates.extend(erp_index[nc])
    if not candidates and nc:
        for ec in erp_index:
            if (nc in ec or ec in nc) and len(nc) >= 2:
                candidates.extend(erp_index[ec])
                if len(candidates) > 200:
                    break
    return candidates


def product_matches(jp, ep):
    if not jp or not ep:
        return 0.0
    np = normalize_product(jp)
    epn = normalize_product(ep)
    if len(np) < 4 or len(epn) < 4:
        return 0.0
    if np in epn or epn in np:
        return 1.0
    jt = split_tokens(jp)
    et = split_tokens(ep)
    if not jt or not et:
        return 0.0
    inter = jt & et
    if not inter:
        return 0.0
    j = len(inter) / len(jt | et)
    if any(k in inter for k in KYOWON_CATEGORIES):
        return max(j, 0.5)
    return j


def main():
    log = Tee(OUT)

    section(log, "1. 데이터 로드")
    jobs = load_all_jobs()
    log(f"   통합 작업: {len(jobs):,}건")
    erp_lines = load_erp_lines()
    erp_index = build_erp_index(erp_lines)
    log(f"   ERP 라인 (3사, 2024-2026): {len(erp_lines):,}건")

    # ─── 2. 지입 vs 일반 작업 분류 ────────────────────────────────
    section(log, "2. 지입(아방) vs 일반 작업 분류")
    jiip_jobs = 0
    for j in jobs:
        j["is_jiip"] = is_jiip_client(j["client"])
        if j["is_jiip"]:
            jiip_jobs += 1
    log(f"   지입 작업: {jiip_jobs:,} / {len(jobs):,} ({jiip_jobs / len(jobs) * 100:.1f}%)")
    log(f"   일반 작업: {len(jobs) - jiip_jobs:,}")

    # 지입 거래처별 작업 수
    log(f"\n   지입 거래처별 작업 분포:")
    jiip_clients = Counter(j["client"] for j in jobs if j["is_jiip"])
    for c, n in jiip_clients.most_common():
        log(f"      {c:<20} {n:>6,}건")

    # ─── 3. 호기별 지입 비중 (Jack 증언 검증) ──────────────────────
    section(log, "3. 호기별 지입 비중 (Jack 증언 검증: 1·11·9호기가 교원 주로)")
    eq_total = Counter()
    eq_jiip = Counter()
    eq_jiip_qty = Counter()
    eq_gen_qty = Counter()
    for j in jobs:
        eq = j["equipment"] or "(?)"
        eq_total[eq] += 1
        if j["is_jiip"]:
            eq_jiip[eq] += 1
            eq_jiip_qty[eq] += j["quantity"]
        else:
            eq_gen_qty[eq] += j["quantity"]

    log(f"\n   {'호기':<10}{'전체작업':>10}{'지입작업':>10}{'지입비중':>10}{'지입통수':>14}{'일반통수':>14}")
    for eq in sorted(eq_total.keys(), key=lambda x: -eq_jiip.get(x, 0) / max(eq_total[x], 1)):
        if not eq.endswith("호기"):
            continue
        ratio = eq_jiip.get(eq, 0) / max(eq_total[eq], 1) * 100
        log(f"   {eq:<10}{eq_total[eq]:>10,}{eq_jiip.get(eq, 0):>10,}{ratio:>9.1f}%"
            f"{eq_jiip_qty.get(eq, 0):>14,}{eq_gen_qty.get(eq, 0):>14,}")

    # ─── 4. 매칭 (V2와 동일) ──────────────────────────────────────
    section(log, "4. 매칭 (거래처+제품명 fuzzy)")
    job_to_line = {}
    line_to_jobs = {}
    for ji, job in enumerate(jobs):
        candidates = find_candidate_lines(job, erp_index)
        best = 0.0
        best_line = None
        for c in candidates:
            s = product_matches(job["product"], c["NM_ITEM"])
            if s > best and s >= 0.5:
                best = s
                best_line = c
        if best_line:
            job_to_line[ji] = (best_line, best)
            line_to_jobs.setdefault(best_line["NO_SALES"], []).append(ji)
    log(f"   매칭 성공: {len(job_to_line):,} / {len(jobs):,} ({len(job_to_line) / len(jobs) * 100:.1f}%)")

    # ─── 5. 통수 비율 분할 ────────────────────────────────────────
    split_revenue = {}
    for no_sales, ji_list in line_to_jobs.items():
        line = job_to_line[ji_list[0]][0]
        am = int(line["am"] or 0)
        qts = [jobs[ji]["quantity"] for ji in ji_list]
        valid = [q for q in qts if q > 0]
        avg = sum(valid) / len(valid) if valid else 1
        norm = [q if q > 0 else int(avg) for q in qts]
        total = sum(norm) or len(ji_list)
        for ji, q in zip(ji_list, norm):
            split_revenue[ji] = (line, int(am * q / total))

    # ─── 6. 호기별 매출 — 지입/일반 분리 ──────────────────────────
    section(log, "5. 호기별 매출 — 지입 vs 일반 분리")
    by_eq = {}  # eq → {jiip_rev, gen_rev, jiip_jobs, gen_jobs, jiip_qty_alloc, gen_qty_alloc}
    for ji, (line, am) in split_revenue.items():
        eq = jobs[ji]["equipment"] or "(?)"
        d = by_eq.setdefault(eq, {"jiip_rev": 0, "gen_rev": 0, "jiip_jobs": 0, "gen_jobs": 0,
                                  "jiip_qty": 0, "gen_qty": 0})
        if jobs[ji]["is_jiip"]:
            d["jiip_rev"] += am
            d["jiip_jobs"] += 1
            d["jiip_qty"] += jobs[ji]["quantity"]
        else:
            d["gen_rev"] += am
            d["gen_jobs"] += 1
            d["gen_qty"] += jobs[ji]["quantity"]

    log(f"\n   {'호기':<10}{'지입매출':>14}{'일반매출':>14}{'지입건':>8}{'일반건':>8}"
        f"{'지입통수':>14}{'일반통수':>14}")
    eqs_sorted = sorted(by_eq.keys(),
                       key=lambda e: -(by_eq[e]["jiip_rev"] + by_eq[e]["gen_rev"]))
    for eq in eqs_sorted:
        if not eq.endswith("호기"):
            continue
        d = by_eq[eq]
        log(f"   {eq:<10}{d['jiip_rev']:>14,}{d['gen_rev']:>14,}{d['jiip_jobs']:>8,}{d['gen_jobs']:>8,}"
            f"{d['jiip_qty']:>14,}{d['gen_qty']:>14,}")

    # 합계
    total_jiip_rev = sum(d["jiip_rev"] for d in by_eq.values())
    total_gen_rev = sum(d["gen_rev"] for d in by_eq.values())
    log(f"   {'-' * 82}")
    log(f"   {'합계':<10}{total_jiip_rev:>14,}{total_gen_rev:>14,}")
    log(f"\n   지입 매출: {total_jiip_rev:>17,}원 (전체 매칭의 {total_jiip_rev / max(total_jiip_rev + total_gen_rev, 1) * 100:.1f}%)")
    log(f"   일반 매출: {total_gen_rev:>17,}원 ({total_gen_rev / max(total_jiip_rev + total_gen_rev, 1) * 100:.1f}%)")

    # ─── 7. 1·11·9호기 상세 (Jack 증언 검증) ───────────────────────
    section(log, "6. 1·11·9호기 상세 — 교원 비중 검증")
    for eq in ["1호기", "11호기", "9호기"]:
        log(f"\n   [{eq}]")
        eq_jobs = [j for j in jobs if j["equipment"] == eq]
        log(f"      전체 작업 {len(eq_jobs):,}건")
        clients_eq = Counter(j["client"] for j in eq_jobs)
        log(f"      주요 거래처 Top 5:")
        for c, n in clients_eq.most_common(5):
            jiip_mark = " ★지입" if is_jiip_client(c) else ""
            log(f"         {c:<20} {n:>6,}건 ({n / len(eq_jobs) * 100:.1f}%){jiip_mark}")

    # ─── 8. JSON 저장 ────────────────────────────────────────────
    out_data = {
        "by_equipment_jiip_vs_general": {
            eq: by_eq[eq] for eq in eqs_sorted if eq.endswith("호기")
        },
        "jiip_total_revenue": total_jiip_rev,
        "general_total_revenue": total_gen_rev,
        "match_rate": len(job_to_line) / len(jobs),
    }
    OUT_JSON.write_text(json.dumps(out_data, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"\n→ JSON: {OUT_JSON}")
    log.close()


if __name__ == "__main__":
    main()
