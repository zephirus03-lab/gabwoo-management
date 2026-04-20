"""
인쇄기별 매출 시뮬레이션 V4 — 후가공 외주비 차감.

V3 → V4 개선:
  1. 후가공 단가표(01_매입 단가표(출판).xlsx, '후가공' 시트) 파싱
  2. 재단지시서 각 작업의 규격·통수 → 연(R) 환산 → 단가 적용
  3. 호기별 외주비 합산 → 매출 - 외주비 = 공헌이익 추정

한계(중요):
  - 재단지시서에 후가공 공정 유형(유광/무광/박/실크/코팅) 정보 없음
    → 유광 대표 단가 적용 (실제 박·실크·에폭시는 단가 2~4배 → 외주비 과소 추정)
  - 통수/판수 해석 가정: 통수 = 총 인쇄 장수, 1연(R) = 500장
  - 후가공 컬럼에 외주처가 적힌 작업만 외주비 부착. 내부 후가공/생략은 0.
  - 교원TK 전용 단가는 일반 후가공 단가보다 약간 높을 수 있으나 여기서는 일반 단가 동일 적용.
"""
from pathlib import Path
import json
import re
import urllib.request
import openpyxl
import pymssql
from collections import Counter

ENV = Path("/Users/jack/dev/gabwoo/견적계산기/.env.local")
OUT = Path(__file__).parent / "output" / "machine_revenue_v4.txt"
OUT_JSON = Path(__file__).parent / "output" / "machine_revenue_v4.json"
JAEDAN_2025 = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/재단지시서_2025년.xlsx")
PRICING_XLSX = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/01_매입 단가표(출판).xlsx")

SUPABASE_URL = "https://btbqzbrtsmwoolurpqgx.supabase.co"
SUPABASE_SR = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJ0YnF6YnJ0c213b29sdXJwcWd4Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTQ0MzgwNSwiZXhwIjoyMDkxMDE5ODA1fQ.M6H48_EfhK_GbNB8LC557VukouFaYdXXTjSD8S5azqw"
DATA_LOCAL = Path("/tmp/gw_check/data.json")

# 출판 스코프 (동행복권=패키지 제외)
JIIP_NAME_KW = ["교원", "이투스", "에듀윌"]
CLIENT_ALIAS = {
    "교원": ["(주)교원구몬", "(주)교원프라퍼티 인천공장", "(주)교원"],
    "교원구몬": ["(주)교원구몬"],
    "교원턴키": ["(주)교원구몬", "(주)교원"],
    "이투스에듀": ["(주)이투스에듀"],
    "한국조폐공사": ["한국조폐공사"],
    "에듀윌": ["(주)에듀윌"],
}
KYOWON_CATEGORIES = ["정교재", "전집", "구몬", "성장노트", "한자", "수학", "일어", "중국어"]

# 1R = 500장 (업계 표준, 검증 필요)
SHEETS_PER_REAM = 500

# 후가공 유광 대표 단가 (5R~15R 미만 구간) — 규격별 원/연
# 출처: 01_매입 단가표(출판).xlsx '후가공' 시트 row 5
OUTSOURCE_UNIT = {
    "8x6": 28420,
    "국전": 30380,
    "4x6": 39200,
}


def load_env():
    env = {}
    for line in ENV.read_text().splitlines():
        line = line.strip()
        if line and "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def connect():
    env = load_env()
    return pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=10,
    )


class Tee:
    def __init__(self, path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.f = open(path, "w", encoding="utf-8")

    def __call__(self, *args):
        msg = " ".join(str(a) for a in args)
        print(msg)
        self.f.write(msg + "\n")

    def close(self):
        self.f.close()


def section(log, title):
    log("\n" + "=" * 80)
    log(title)
    log("=" * 80)


def normalize_client(s):
    if not s:
        return ""
    s = re.sub(r"\(주\)|주식회사|㈜|\(유\)|유한회사", "", str(s))
    s = re.sub(r"\s+", "", s)
    return s.lower()


def normalize_product(s):
    if not s:
        return ""
    return re.sub(r"[\s/_,()\[\]·\-]+", "", str(s)).lower()


def split_tokens(s):
    if not s:
        return set()
    return set(t for t in re.split(r"[\s/_,()\[\]·\-]+", str(s)) if len(t) >= 2)


def to_int(v):
    if v is None:
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def is_jiip_client(client_name):
    if not client_name:
        return False
    return any(kw in client_name for kw in JIIP_NAME_KW)


def classify_size(size_str):
    """규격 '636*890' → '국전' 등으로 분류.
    출판 업계 통상:
      636*890, 650*940 → 국전
      545*790, 545*788 → 4*6
      800*545, 800*460 등 → 8*6 (일부 4*6 변형)
    """
    if not size_str:
        return "국전"  # 기본값
    s = str(size_str)
    # 숫자 2개 추출
    nums = re.findall(r"\d+", s)
    if len(nums) < 2:
        return "국전"
    try:
        w, h = sorted([int(nums[0]), int(nums[1])])
    except ValueError:
        return "국전"
    # 4*6전: 788*1091 전지. 국전: 636*939 전지.
    # 절수 컷된 크기로도 분류 가능하지만 단순화
    if w >= 600 and h >= 800:  # 국전~ 4*6
        # 국전 636*890 근처
        if 600 <= w <= 680 and 850 <= h <= 950:
            return "국전"
        # 4*6 788*1091 근처
        if 780 <= w <= 820 and 1050 <= h <= 1120:
            return "4x6"
    # 그 외는 국전 기본
    return "국전"


def estimate_outsource_cost(quantity, size_str, has_post):
    """작업별 후가공 외주비 추정.
      - 후가공 컬럼이 비어있으면(has_post=False) 0원
      - quantity/500 = 연수(R)
      - 규격 분류 → 단가 적용
    """
    if not has_post or quantity <= 0:
        return 0
    reams = quantity / SHEETS_PER_REAM
    size_cls = classify_size(size_str)
    unit = OUTSOURCE_UNIT.get(size_cls, OUTSOURCE_UNIT["국전"])
    return int(reams * unit)


def fetch_supabase_data():
    DATA_LOCAL.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(
        f"{SUPABASE_URL}/storage/v1/object/gabwoo-data/data.json",
        headers={"apikey": SUPABASE_SR, "Authorization": f"Bearer {SUPABASE_SR}"},
    )
    with urllib.request.urlopen(req) as r:
        DATA_LOCAL.write_bytes(r.read())
    return json.loads(DATA_LOCAL.read_text())


def parse_jaedan_xlsx(path):
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if "재단지시서" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["재단지시서"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header_idx = next((i for i, r in enumerate(rows[:5])
                       if r and any(c == "설비" for c in r if c)), None)
    if header_idx is None:
        return []
    header = rows[header_idx]
    col = {n: i for i, n in enumerate(header) if n}
    out = []
    current_eq = None
    for r in rows[header_idx + 1:]:
        if not r or len(r) <= max(col.values()):
            continue
        eq = r[col["설비"]]
        client = r[col.get("거래처", -1)] if "거래처" in col else None
        product = r[col.get("제품명", -1)] if "제품명" in col else None
        cnt = r[col.get("통수", -1)] if "통수" in col else None
        post = r[col.get("후가공", -1)] if "후가공" in col else None
        size = r[col.get("규격", -1)] if "규격" in col else None
        if eq and str(eq).strip() not in ("-", "", "설비"):
            current_eq = str(eq).strip()
        if not (client and product) or str(client).strip() in ("-", "", "거래처"):
            continue
        out.append({
            "equipment": current_eq,
            "client": str(client).strip(),
            "product": str(product).strip(),
            "quantity": to_int(cnt),
            "post_process": str(post).strip() if post else None,
            "size": str(size).strip() if size else None,
            "source": "2025_xlsx",
        })
    return out


def load_all_jobs():
    sdata = fetch_supabase_data()
    jobs = []
    for j in sdata.get("all_completed", []):
        eq = j.get("equipment")
        if eq and str(eq).strip() not in ("-", "", "설비"):
            jobs.append({
                "equipment": str(eq).strip(),
                "client": j.get("client"),
                "product": j.get("product"),
                "quantity": to_int(j.get("quantity")),
                "post_process": j.get("post_process"),
                "size": j.get("size"),
                "source": "2026_supabase",
            })
    jobs.extend(parse_jaedan_xlsx(JAEDAN_2025))
    return jobs


def load_erp_lines():
    conn = connect()
    cur = conn.cursor(as_dict=True)
    cur.execute("""
        SELECT h.NO_SALES, h.DT_SALES, h.CD_CUST, h.CD_CUST_OWN,
               c.NM_CUST, l.CD_ITEM, i.NM_ITEM,
               CAST(l.AM AS BIGINT) am
        FROM SAL_SALESL l
        INNER JOIN SAL_SALESH h ON h.NO_SALES = l.NO_SALES
        LEFT JOIN MAS_CUST c ON c.CD_CUST = h.CD_CUST
        LEFT JOIN PRT_ITEM i ON i.CD_ITEM = l.CD_ITEM
        WHERE (h.DT_SALES LIKE '2024%' OR h.DT_SALES LIKE '2025%' OR h.DT_SALES LIKE '2026%')
          AND h.ST_SALES='Y' AND h.AM > 0
    """)
    lines = cur.fetchall()
    conn.close()
    return lines


def build_erp_index(erp_lines):
    idx = {}
    for l in erp_lines:
        nc = normalize_client(l["NM_CUST"])
        if nc:
            idx.setdefault(nc, []).append(l)
    return idx


def find_candidate_lines(job, erp_index):
    nc = normalize_client(job["client"])
    candidates = []
    raw = job["client"]
    if raw and raw.strip() in CLIENT_ALIAS:
        for alias_name in CLIENT_ALIAS[raw.strip()]:
            candidates.extend(erp_index.get(normalize_client(alias_name), []))
    if nc and nc in erp_index:
        candidates.extend(erp_index[nc])
    if not candidates and nc:
        for ec in erp_index:
            if (nc in ec or ec in nc) and len(nc) >= 2:
                candidates.extend(erp_index[ec])
                if len(candidates) > 200:
                    break
    return candidates


def product_matches(jp, ep):
    if not jp or not ep:
        return 0.0
    np = normalize_product(jp)
    epn = normalize_product(ep)
    if len(np) < 4 or len(epn) < 4:
        return 0.0
    if np in epn or epn in np:
        return 1.0
    jt = split_tokens(jp)
    et = split_tokens(ep)
    if not jt or not et:
        return 0.0
    inter = jt & et
    if not inter:
        return 0.0
    j = len(inter) / len(jt | et)
    if any(k in inter for k in KYOWON_CATEGORIES):
        return max(j, 0.5)
    return j


def main():
    log = Tee(OUT)

    section(log, "1. 데이터 로드")
    jobs = load_all_jobs()
    log(f"   통합 작업: {len(jobs):,}건")
    erp_lines = load_erp_lines()
    erp_index = build_erp_index(erp_lines)
    log(f"   ERP 라인: {len(erp_lines):,}건")

    # 지입 분류 + 외주비 추정
    for j in jobs:
        j["is_jiip"] = is_jiip_client(j["client"])
        has_post = j["post_process"] and str(j["post_process"]).strip() not in ("", "-")
        j["outsource_cost"] = estimate_outsource_cost(j["quantity"], j["size"], has_post)

    # ─── 2. 후가공 외주 커버리지 ──────────────────────────────────
    section(log, "2. 후가공 외주 발생 작업 비중")
    has_post_n = sum(1 for j in jobs if j["outsource_cost"] > 0)
    log(f"   외주 발생 작업: {has_post_n:,} / {len(jobs):,} ({has_post_n / len(jobs) * 100:.1f}%)")
    total_out_cost = sum(j["outsource_cost"] for j in jobs)
    log(f"   전체 후가공 외주비 추정: {total_out_cost:>17,}원")
    log(f"   평균 외주비/작업: {total_out_cost / max(has_post_n, 1):>12,}원")

    # ─── 3. 매칭 + 통수 비율 분할 (V3와 동일) ──────────────────────
    section(log, "3. 매출 매칭 + 통수 비율 분할")
    job_to_line = {}
    line_to_jobs = {}
    for ji, job in enumerate(jobs):
        candidates = find_candidate_lines(job, erp_index)
        best = 0.0
        best_line = None
        for c in candidates:
            s = product_matches(job["product"], c["NM_ITEM"])
            if s > best and s >= 0.5:
                best = s
                best_line = c
        if best_line:
            job_to_line[ji] = (best_line, best)
            line_to_jobs.setdefault(best_line["NO_SALES"], []).append(ji)
    split_revenue = {}
    for no_sales, ji_list in line_to_jobs.items():
        line = job_to_line[ji_list[0]][0]
        am = int(line["am"] or 0)
        qts = [jobs[ji]["quantity"] for ji in ji_list]
        valid = [q for q in qts if q > 0]
        avg = sum(valid) / len(valid) if valid else 1
        norm = [q if q > 0 else int(avg) for q in qts]
        total = sum(norm) or len(ji_list)
        for ji, q in zip(ji_list, norm):
            split_revenue[ji] = (line, int(am * q / total))
    log(f"   매칭: {len(job_to_line):,} / {len(jobs):,} ({len(job_to_line) / len(jobs) * 100:.1f}%)")

    # ─── 4. 호기별 매출 - 외주비 = 공헌이익 (지입/일반 분리) ────────
    section(log, "4. 호기별 매출 - 외주비 = 공헌이익 (V4, 추정)")
    by_eq = {}
    for ji, (line, am) in split_revenue.items():
        eq = jobs[ji]["equipment"] or "(?)"
        d = by_eq.setdefault(eq, {
            "jiip_rev": 0, "gen_rev": 0,
            "jiip_out": 0, "gen_out": 0,
            "jiip_jobs": 0, "gen_jobs": 0,
        })
        if jobs[ji]["is_jiip"]:
            d["jiip_rev"] += am
            d["jiip_out"] += jobs[ji]["outsource_cost"]
            d["jiip_jobs"] += 1
        else:
            d["gen_rev"] += am
            d["gen_out"] += jobs[ji]["outsource_cost"]
            d["gen_jobs"] += 1

    log(f"\n   {'호기':<8}{'지입매출':>13}{'지입외주':>13}{'지입공헌':>13}"
        f"{'일반매출':>13}{'일반외주':>13}{'일반공헌':>13}")
    eqs_sorted = sorted(
        [e for e in by_eq if e.endswith("호기")],
        key=lambda e: -(by_eq[e]["jiip_rev"] + by_eq[e]["gen_rev"])
    )
    for eq in eqs_sorted:
        d = by_eq[eq]
        jc = d["jiip_rev"] - d["jiip_out"]
        gc = d["gen_rev"] - d["gen_out"]
        log(f"   {eq:<8}"
            f"{d['jiip_rev']:>13,}{d['jiip_out']:>13,}{jc:>13,}"
            f"{d['gen_rev']:>13,}{d['gen_out']:>13,}{gc:>13,}")

    total_j_rev = sum(d["jiip_rev"] for d in by_eq.values())
    total_j_out = sum(d["jiip_out"] for d in by_eq.values())
    total_g_rev = sum(d["gen_rev"] for d in by_eq.values())
    total_g_out = sum(d["gen_out"] for d in by_eq.values())
    log(f"   {'-' * 92}")
    log(f"   {'합계':<8}"
        f"{total_j_rev:>13,}{total_j_out:>13,}{total_j_rev - total_j_out:>13,}"
        f"{total_g_rev:>13,}{total_g_out:>13,}{total_g_rev - total_g_out:>13,}")

    log(f"\n   ⚠️ 외주비는 유광 대표 단가 적용한 러프 추정치.")
    log(f"   박·실크·에폭시 등 고가 공정 있는 작업은 과소추정.")

    # ─── 5. 외주 공헌이익률 ───────────────────────────────────────
    section(log, "5. 호기별 공헌이익률 (매출-외주비)/매출")
    log(f"\n   {'호기':<8}{'지입 공헌이익률':>16}{'일반 공헌이익률':>16}")
    for eq in eqs_sorted:
        d = by_eq[eq]
        jr = (d["jiip_rev"] - d["jiip_out"]) / d["jiip_rev"] if d["jiip_rev"] > 0 else 0
        gr = (d["gen_rev"] - d["gen_out"]) / d["gen_rev"] if d["gen_rev"] > 0 else 0
        log(f"   {eq:<8}{jr * 100:>15.1f}%{gr * 100:>15.1f}%")

    # 음수(적자) 표시
    log(f"\n   ⚠️ 음수 = 외주비가 매출 초과 (지입 특성상 흔함 — 매출=임가공비인데 외주비가 더 큼)")

    # JSON 저장
    out_data = {
        "by_equipment_v4": {eq: by_eq[eq] for eq in eqs_sorted},
        "totals": {
            "jiip_revenue": total_j_rev, "jiip_outsource": total_j_out,
            "general_revenue": total_g_rev, "general_outsource": total_g_out,
        },
        "limits": {
            "outsource_unit_source": "유광 5R~15R 미만 구간 대표값",
            "missing_process_types": "유광/무광/박/실크/에폭시 구분 불가",
            "reams_per_sheet": SHEETS_PER_REAM,
            "size_classification": "규격에서 국전/4*6만 이분법적 분류",
        },
    }
    OUT_JSON.write_text(json.dumps(out_data, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"\n→ JSON: {OUT_JSON}")
    log.close()


if __name__ == "__main__":
    main()
