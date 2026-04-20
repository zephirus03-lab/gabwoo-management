"""
인쇄기별 매출 추정 가능성 탐구.

3가지 데이터원 교차 검증:
  A. 패키지 수주관리(엑셀) — 관리번호·거래처·품명·공정·진행 7단계
  B. 패키지 인쇄종합(엑셀) — 인쇄일정관리 시트 (수주번호·매출처·공정상세·인쇄기?)
  C. SNOTES ERP — SAL_SALESH/L 매출 라인

매칭 후보:
  1. 관리번호(`260402-0005`) ↔ ERP NO_EST/NO_SO/NO_SALES 직접 매칭
  2. 거래처+품명 fuzzy 매칭
  3. 인쇄종합에 실제 인쇄기(설비) 컬럼이 있는지 확인

각도 5가지로 점검하고 매칭률 측정.
"""
from pathlib import Path
import re
import openpyxl
import pymssql

PKG_DIR = Path("/Users/jack/dev/gabwoo/패키지_생산진행현황/생산진행현황")
ENV = Path("/Users/jack/dev/gabwoo/견적계산기/.env.local")
OUT = Path("/Users/jack/dev/gabwoo/관리 대시보드/scripts/output/machine_revenue_feasibility.txt")

ORDERS_XLSX = PKG_DIR / "수주관리 2026년도.xlsx"


def load_env():
    env = {}
    for line in ENV.read_text().splitlines():
        line = line.strip()
        if line and "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def connect():
    env = load_env()
    return pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=10,
    )


class Tee:
    def __init__(self, path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.f = open(path, "w", encoding="utf-8")

    def __call__(self, *args):
        msg = " ".join(str(a) for a in args)
        print(msg)
        self.f.write(msg + "\n")

    def close(self):
        self.f.close()


def section(log, title):
    log("\n" + "=" * 80)
    log(title)
    log("=" * 80)


def main():
    log = Tee(OUT)

    # ─── A. 패키지 수주관리 엑셀 분석 ────────────────────────────────
    section(log, "A. 패키지 수주관리 엑셀 — 관리번호 형식·거래처·인쇄기 컬럼 확인")
    if not ORDERS_XLSX.exists():
        log(f"   ❌ 파일 없음: {ORDERS_XLSX}")
        return
    wb = openpyxl.load_workbook(str(ORDERS_XLSX), read_only=True, data_only=True)
    log(f"   파일: {ORDERS_XLSX.name}")
    log(f"   시트: {wb.sheetnames}")

    # 가장 최근 월 시트 1개에서 헤더 + 샘플 5행
    month_sheets = [s for s in wb.sheetnames if re.match(r"수주관리\s*2026", s)]
    if month_sheets:
        sn = month_sheets[-1]
        log(f"\n   분석 대상 시트: '{sn}'")
        ws = wb[sn]
        rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
        # 헤더 탐지
        header = None
        header_row_idx = None
        for i, r in enumerate(rows):
            if r and any(c and "관리번호" in str(c) for c in r):
                header = r
                header_row_idx = i
                break
        if header:
            log(f"   헤더 행 #{header_row_idx + 1}: {[str(c)[:15] if c else '-' for c in header]}")
            log(f"\n   샘플 데이터 5행 (관리번호/거래처/품명/공정 위주):")
            for r in rows[header_row_idx + 1:header_row_idx + 6]:
                if r and len(r) > 9:
                    log(f"   No={r[0]} 관리번호={r[3]} 거래처={r[4]} 품명={str(r[5])[:30]} 공정={str(r[9])[:40]}")
        # 인쇄기 컬럼 있는지
        if header:
            machine_cols = [c for c in header if c and any(k in str(c) for k in ["인쇄기", "설비", "기계", "호기"])]
            log(f"\n   인쇄기/설비 관련 컬럼: {machine_cols if machine_cols else '❌ 없음'}")

    # 관리번호 일부 수집 (최근 시트 50개)
    mgmt_nos = set()
    if month_sheets:
        for sn in month_sheets[-3:]:  # 최근 3개월
            ws = wb[sn]
            for r in ws.iter_rows(min_row=1, values_only=True):
                if r and len(r) > 3 and r[3]:
                    mn = str(r[3]).strip()
                    if re.match(r"\d{6}-\d{3,4}", mn):
                        mgmt_nos.add(mn)
    wb.close()
    log(f"\n   수집된 관리번호 샘플: {len(mgmt_nos)}개")
    sample_mns = list(mgmt_nos)[:5]
    log(f"   샘플: {sample_mns}")

    # ─── B. 인쇄종합 엑셀 분석 ───────────────────────────────────────
    section(log, "B. 인쇄종합 엑셀 — 인쇄일정관리 시트의 인쇄기 컬럼 확인")
    printing_xlsx = list(PKG_DIR.glob("인쇄종합_*.xlsx"))
    # 0409 폴더에도 사본이 있을 수 있음
    if not printing_xlsx:
        for sub in PKG_DIR.glob("0*"):
            printing_xlsx.extend(sub.glob("인쇄종합_*.xlsx"))
    log(f"   파일 후보: {[p.name for p in printing_xlsx]}")
    if printing_xlsx:
        path = printing_xlsx[0]
        log(f"   분석: {path}")
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        log(f"   시트: {wb.sheetnames[:10]}{'...' if len(wb.sheetnames) > 10 else ''}")
        if "인쇄일정관리" in wb.sheetnames:
            ws = wb["인쇄일정관리"]
            rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
            for i, r in enumerate(rows[:3]):
                log(f"   row{i + 1}: {[str(c)[:15] if c else '-' for c in r[:25]]}")
            # 헤더 추정
            header = rows[1] if len(rows) > 1 else None
            if header:
                machine_cols_idx = [(i, c) for i, c in enumerate(header) if c and any(k in str(c) for k in ["인쇄기", "설비", "기계", "호기", "machine"])]
                log(f"\n   인쇄기 관련 컬럼: {machine_cols_idx if machine_cols_idx else '❌ 없음'}")
                # 모든 헤더 표시
                log(f"\n   전체 헤더: {[c for c in header if c]}")
        wb.close()

    # ─── C. ERP에서 관리번호 직접 검색 ───────────────────────────────
    section(log, "C. SNOTES ERP에서 관리번호 직접 검색 (NO_EST/NO_SO/NO_SALES 등)")
    conn = connect()
    cur = conn.cursor(as_dict=True)

    target_cols = [
        ("SAL_SALESH", "NO_SALES"),
        ("SAL_SALESH", "NO_EST"),
        ("SAL_SALESH", "NO_SO"),
        ("PRT_SOH", "NO_SO"),
        ("PRT_SOH", "NO_EST"),
        ("PRT_ESTH", "NO_EST"),
        ("PRT_WO", "NO_WO"),
        ("PRT_WO", "NO_SO"),
    ]

    for mn in sample_mns[:3]:
        log(f"\n   관리번호 '{mn}' 검색:")
        for tbl, col in target_cols:
            try:
                cur.execute(f"SELECT COUNT(*) c FROM {tbl} WHERE {col} = %s OR {col} LIKE %s",
                            (mn, f"%{mn}%"))
                r = cur.fetchone()
                if r and r["c"] > 0:
                    log(f"      ✓ {tbl}.{col} 매칭: {r['c']}건")
            except Exception as e:
                pass  # 컬럼 없으면 skip

        # NO_* 컬럼이 일반적으로 형식 다를 수 있음 — 일부분 검색
        # 관리번호 끝 4자리만으로
        suffix = mn.split("-")[-1]
        for tbl, col in [("SAL_SALESH", "DC_REMARK"), ("PRT_SOH", "DC_REMARK")]:
            try:
                cur.execute(f"SELECT COUNT(*) c FROM {tbl} WHERE {col} LIKE %s", (f"%{mn}%",))
                r = cur.fetchone()
                if r and r["c"] > 0:
                    log(f"      ✓ {tbl}.{col}(비고) 매칭: {r['c']}건")
            except Exception:
                pass

    # ─── D. 패키지 거래처 → SAL_SALESH 매출 발생 여부 ─────────────────
    section(log, "D. 패키지 핵심 거래처의 ERP 매출 데이터 확인 (CD_CUST_OWN 별)")
    pkg_kw = ["코스맥스", "코스메카", "정샘물", "라샘", "매그니프", "콜마"]
    for kw in pkg_kw:
        cur.execute("""
            SELECT h.CD_CUST_OWN,
                   COUNT(DISTINCT h.NO_SALES) cnt,
                   SUM(CAST(h.AM AS BIGINT)) am
            FROM SAL_SALESH h
            INNER JOIN MAS_CUST c ON c.CD_CUST = h.CD_CUST
            WHERE h.DT_SALES LIKE '2026%'
              AND c.NM_CUST LIKE '%' + %s + '%'
              AND h.ST_SALES='Y' AND h.AM > 0
            GROUP BY h.CD_CUST_OWN
            ORDER BY h.CD_CUST_OWN
        """, (kw,))
        rows = cur.fetchall()
        if rows:
            log(f"\n   '{kw}' 매출 (2026년):")
            for r in rows:
                log(f"      CD_CUST_OWN={r['CD_CUST_OWN']} 건수={r['cnt']:>4} 매출={int(r['am'] or 0):>12,}")
        else:
            log(f"\n   '{kw}' 2026년 매출 0건")

    # ─── E. PRT_WO (작업지시) 패키지 거래처 행 존재 여부 ──────────────
    section(log, "E. PRT_WO(작업지시)에 패키지 거래처 데이터가 있는가?")
    for kw in ["코스맥스", "정샘물", "코스메카"]:
        try:
            cur.execute("""
                SELECT TOP 5 wo.NO_WO, wo.DT_REG, wo.NM_ITEM, c.NM_CUST
                FROM PRT_WO wo
                LEFT JOIN MAS_CUST c ON c.CD_CUST = wo.CD_CUST
                WHERE wo.DT_REG LIKE '2026%' AND c.NM_CUST LIKE %s
            """, (f"%{kw}%",))
            rows = cur.fetchall()
            if rows:
                log(f"\n   '{kw}' PRT_WO 샘플:")
                for r in rows:
                    log(f"      WO={r['NO_WO']} 등록={r['DT_REG']} 거래처={r['NM_CUST']} 품명={str(r.get('NM_ITEM'))[:40]}")
            else:
                log(f"\n   '{kw}' PRT_WO 0건 (2026)")
        except Exception as e:
            log(f"   PRT_WO 조회 실패: {e}")
            break

    # ─── F. PRT_WOPROC_EQUIP — 인쇄기 매핑 가능 여부 ─────────────────
    section(log, "F. PRT_WOPROC_EQUIP — 작업지시 ↔ 인쇄기 매핑")
    try:
        cur.execute("""
            SELECT TOP 1 * FROM PRT_WOPROC_EQUIP
        """)
        r = cur.fetchone()
        if r:
            log(f"   샘플 행 컬럼: {list(r.keys())}")
            for k, v in r.items():
                log(f"      {k}={v}")
    except Exception as e:
        log(f"   조회 실패: {e}")

    # 2026년 설비별 작업 횟수
    try:
        cur.execute("""
            SELECT TOP 20 we.CD_EQUIP, e.NM_EQUIP, COUNT(*) cnt
            FROM PRT_WOPROC_EQUIP we
            LEFT JOIN PRT_EQUIP e ON e.CD_EQUIP = we.CD_EQUIP
            INNER JOIN PRT_WO wo ON wo.NO_WO = we.NO_WO
            WHERE wo.DT_REG LIKE '2026%'
            GROUP BY we.CD_EQUIP, e.NM_EQUIP
            ORDER BY COUNT(*) DESC
        """)
        log("\n   2026년 설비별 작업 횟수:")
        for r in cur.fetchall():
            log(f"      {str(r.get('CD_EQUIP')):<15} {str(r.get('NM_EQUIP')):<30} {r['cnt']:>8,}")
    except Exception as e:
        log(f"   조회 실패: {e}")

    log.close()
    conn.close()


if __name__ == "__main__":
    main()
