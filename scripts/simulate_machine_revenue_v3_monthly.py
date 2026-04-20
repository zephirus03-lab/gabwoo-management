"""
V3.1 — 월별 호기별 매출 (2025.01 ~ 2026.03, 15개월).

V3와 동일 로직 + 2025 엑셀의 "블록 헤더 datetime"으로 일자 forward-fill.
- 엑셀 구조: 매일 업로드된 재단지시서가 블록 단위로 누적
  - row 시작부: col18 = datetime (블록 헤더 1번째 줄, col0="설비별 재단 지시서")
  - 다음 줄: col8 = datetime (블록 헤더 2번째 줄, col0="※ XL106 ...")
  - 그 다음부터 데이터 (다음 블록 헤더 전까지)
- 이 블록 날짜를 forward-fill해서 각 작업 행에 부착
- 월별 호기별 매출 + 2026 Supabase completed_date 합산
"""
from pathlib import Path
import json
import re
import urllib.request
import openpyxl
import pymssql
from datetime import datetime
from collections import Counter, defaultdict

ENV = Path("/Users/jack/dev/gabwoo/견적계산기/.env.local")
OUT = Path(__file__).parent / "output" / "machine_revenue_monthly.txt"
OUT_JSON = Path(__file__).parent / "output" / "machine_revenue_monthly.json"
JAEDAN_2025 = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/재단지시서_2025년.xlsx")

SUPABASE_URL = "https://btbqzbrtsmwoolurpqgx.supabase.co"
SUPABASE_SR = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJ0YnF6YnJ0c213b29sdXJwcWd4Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTQ0MzgwNSwiZXhwIjoyMDkxMDE5ODA1fQ.M6H48_EfhK_GbNB8LC557VukouFaYdXXTjSD8S5azqw"
DATA_LOCAL = Path("/tmp/gw_check/data.json")

JIIP_NAME_KW = ["교원", "이투스", "에듀윌"]  # 출판 스코프
CLIENT_ALIAS = {
    "교원": ["(주)교원구몬", "(주)교원프라퍼티 인천공장", "(주)교원"],
    "교원구몬": ["(주)교원구몬"],
    "교원턴키": ["(주)교원구몬", "(주)교원"],
    "이투스에듀": ["(주)이투스에듀"],
    "한국조폐공사": ["한국조폐공사"],
    "에듀윌": ["(주)에듀윌"],
}
KYOWON_CATEGORIES = ["정교재", "전집", "구몬", "성장노트", "한자", "수학", "일어", "중국어"]


def load_env():
    env = {}
    for line in ENV.read_text().splitlines():
        line = line.strip()
        if line and "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def connect():
    env = load_env()
    return pymssql.connect(
        server=env["ERP_HOST"], port=int(env.get("ERP_PORT", 1433)),
        user=env["ERP_USER"], password=env["ERP_PASSWORD"],
        database=env["ERP_DATABASE"], login_timeout=10,
    )


class Tee:
    def __init__(self, path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.f = open(path, "w", encoding="utf-8")

    def __call__(self, *args):
        msg = " ".join(str(a) for a in args)
        print(msg)
        self.f.write(msg + "\n")

    def close(self):
        self.f.close()


def normalize_client(s):
    if not s:
        return ""
    s = re.sub(r"\(주\)|주식회사|㈜|\(유\)|유한회사", "", str(s))
    s = re.sub(r"\s+", "", s)
    return s.lower()


def normalize_product(s):
    if not s:
        return ""
    return re.sub(r"[\s/_,()\[\]·\-]+", "", str(s)).lower()


def split_tokens(s):
    if not s:
        return set()
    return set(t for t in re.split(r"[\s/_,()\[\]·\-]+", str(s)) if len(t) >= 2)


def to_int(v):
    if v is None:
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def is_jiip_client(client_name):
    return bool(client_name and any(kw in client_name for kw in JIIP_NAME_KW))


def fetch_supabase_data():
    DATA_LOCAL.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(
        f"{SUPABASE_URL}/storage/v1/object/gabwoo-data/data.json",
        headers={"apikey": SUPABASE_SR, "Authorization": f"Bearer {SUPABASE_SR}"},
    )
    with urllib.request.urlopen(req) as r:
        DATA_LOCAL.write_bytes(r.read())
    return json.loads(DATA_LOCAL.read_text())


def parse_jaedan_2025_with_dates(path):
    """블록 헤더 datetime forward-fill.
    블록 헤더 식별:
      col18 datetime + col0 "설비별 재단 지시서" → 블록 시작 (1번째 줄)
      col8 datetime + col0 "※" 메모 → 블록 2번째 줄 (무시, 날짜만 재확인)
    """
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if "재단지시서" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["재단지시서"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # 컬럼 헤더 탐지
    header_idx = next((i for i, r in enumerate(rows[:5])
                       if r and any(c == "설비" for c in r if c)), None)
    if header_idx is None:
        return []
    header = rows[header_idx]
    col = {n: i for i, n in enumerate(header) if n}

    out = []
    current_eq = None
    current_date = None

    for idx, r in enumerate(rows):
        if not r:
            continue
        # 블록 헤더 감지: col[0] 또는 col[18] 또는 col[8]에 "설비별 재단 지시서" / "※" 등
        c0 = str(r[0])[:20] if r[0] else ""
        c18 = r[18] if len(r) > 18 else None
        c8 = r[8] if len(r) > 8 else None

        # 헤더 1줄째 (col18 datetime + col0 "설비별")
        if "설비별" in c0 and isinstance(c18, datetime):
            current_date = c18.date()
            continue
        # 헤더 2줄째 (col0 "※" + col8 datetime)
        if c0.startswith("※") and isinstance(c8, datetime):
            current_date = c8.date()
            continue
        # 컬럼명 행 (설비 등)
        if idx <= header_idx:
            continue
        # 데이터 처리
        if len(r) <= max(col.values()):
            continue
        eq = r[col["설비"]]
        client = r[col.get("거래처", -1)] if "거래처" in col else None
        product = r[col.get("제품명", -1)] if "제품명" in col else None
        cnt = r[col.get("통수", -1)] if "통수" in col else None
        post = r[col.get("후가공", -1)] if "후가공" in col else None
        size = r[col.get("규격", -1)] if "규격" in col else None
        if eq and str(eq).strip() not in ("-", "", "설비"):
            current_eq = str(eq).strip()
        if not (client and product) or str(client).strip() in ("-", "", "거래처"):
            continue
        out.append({
            "equipment": current_eq,
            "client": str(client).strip(),
            "product": str(product).strip(),
            "quantity": to_int(cnt),
            "post_process": str(post).strip() if post else None,
            "size": str(size).strip() if size else None,
            "completed_date": current_date.strftime("%Y-%m-%d") if current_date else None,
            "source": "2025_xlsx",
        })
    return out


def load_all_jobs(log):
    log("[a] Supabase data.json…")
    sdata = fetch_supabase_data()
    jobs = []
    for j in sdata.get("all_completed", []):
        eq = j.get("equipment")
        if eq and str(eq).strip() not in ("-", "", "설비"):
            jobs.append({
                "equipment": str(eq).strip(),
                "client": j.get("client"),
                "product": j.get("product"),
                "quantity": to_int(j.get("quantity")),
                "post_process": j.get("post_process"),
                "size": j.get("size"),
                "completed_date": j.get("completed_date"),
                "source": "2026_supabase",
            })
    log(f"   2026 Supabase: {len(jobs):,}건 (completed_date 보유)")

    log("[b] 2025 엑셀 블록 헤더 파싱…")
    jobs_2025 = parse_jaedan_2025_with_dates(JAEDAN_2025)
    log(f"   2025 xlsx: {len(jobs_2025):,}건")
    with_date = sum(1 for j in jobs_2025 if j["completed_date"])
    log(f"   일자 부착 성공: {with_date:,} / {len(jobs_2025):,} ({with_date / max(len(jobs_2025), 1) * 100:.1f}%)")

    jobs.extend(jobs_2025)
    return jobs


def load_erp_lines():
    conn = connect()
    cur = conn.cursor(as_dict=True)
    cur.execute("""
        SELECT h.NO_SALES, h.DT_SALES, h.CD_CUST, c.NM_CUST,
               l.CD_ITEM, i.NM_ITEM, CAST(l.AM AS BIGINT) am
        FROM SAL_SALESL l
        INNER JOIN SAL_SALESH h ON h.NO_SALES = l.NO_SALES
        LEFT JOIN MAS_CUST c ON c.CD_CUST = h.CD_CUST
        LEFT JOIN PRT_ITEM i ON i.CD_ITEM = l.CD_ITEM
        WHERE (h.DT_SALES LIKE '2024%' OR h.DT_SALES LIKE '2025%' OR h.DT_SALES LIKE '2026%')
          AND h.ST_SALES='Y' AND h.AM > 0
    """)
    lines = cur.fetchall()
    conn.close()
    return lines


def build_erp_index(erp_lines):
    idx = {}
    for l in erp_lines:
        nc = normalize_client(l["NM_CUST"])
        if nc:
            idx.setdefault(nc, []).append(l)
    return idx


def find_candidate_lines(job, erp_index):
    nc = normalize_client(job["client"])
    candidates = []
    raw = job["client"]
    if raw and raw.strip() in CLIENT_ALIAS:
        for alias_name in CLIENT_ALIAS[raw.strip()]:
            candidates.extend(erp_index.get(normalize_client(alias_name), []))
    if nc and nc in erp_index:
        candidates.extend(erp_index[nc])
    if not candidates and nc:
        for ec in erp_index:
            if (nc in ec or ec in nc) and len(nc) >= 2:
                candidates.extend(erp_index[ec])
                if len(candidates) > 200:
                    break
    return candidates


def product_matches(jp, ep):
    if not jp or not ep:
        return 0.0
    np = normalize_product(jp)
    epn = normalize_product(ep)
    if len(np) < 4 or len(epn) < 4:
        return 0.0
    if np in epn or epn in np:
        return 1.0
    jt = split_tokens(jp)
    et = split_tokens(ep)
    if not jt or not et:
        return 0.0
    inter = jt & et
    if not inter:
        return 0.0
    j = len(inter) / len(jt | et)
    if any(k in inter for k in KYOWON_CATEGORIES):
        return max(j, 0.5)
    return j


def main():
    log = Tee(OUT)

    log("=" * 80)
    log("1. 데이터 로드 (날짜 포함)")
    log("=" * 80)
    jobs = load_all_jobs(log)
    log(f"\n   통합 작업: {len(jobs):,}건")
    erp_lines = load_erp_lines()
    erp_index = build_erp_index(erp_lines)
    log(f"   ERP 라인: {len(erp_lines):,}건")

    # 지입 분류
    for j in jobs:
        j["is_jiip"] = is_jiip_client(j["client"])

    # 매칭
    log("\n2. 매출 매칭 + 통수 분할")
    job_to_line = {}
    line_to_jobs = {}
    for ji, job in enumerate(jobs):
        candidates = find_candidate_lines(job, erp_index)
        best = 0.0
        best_line = None
        for c in candidates:
            s = product_matches(job["product"], c["NM_ITEM"])
            if s > best and s >= 0.5:
                best = s
                best_line = c
        if best_line:
            job_to_line[ji] = (best_line, best)
            line_to_jobs.setdefault(best_line["NO_SALES"], []).append(ji)
    split_revenue = {}
    for no_sales, ji_list in line_to_jobs.items():
        line = job_to_line[ji_list[0]][0]
        am = int(line["am"] or 0)
        qts = [jobs[ji]["quantity"] for ji in ji_list]
        valid = [q for q in qts if q > 0]
        avg = sum(valid) / len(valid) if valid else 1
        norm = [q if q > 0 else int(avg) for q in qts]
        total = sum(norm) or len(ji_list)
        for ji, q in zip(ji_list, norm):
            split_revenue[ji] = (line, int(am * q / total))
    log(f"   매칭 성공: {len(job_to_line):,} / {len(jobs):,} ({len(job_to_line) / len(jobs) * 100:.1f}%)")

    # ─── 3. 월별 호기별 매출 (2025.01 ~ 2026.03) ────────────────────
    log("\n3. 월별 호기별 매출 (completed_date 기준)")
    # {ym: {eq: {jiip: n, general: n}}}
    matrix = defaultdict(lambda: defaultdict(lambda: {"jiip": 0, "general": 0}))
    no_date = 0
    for ji, (line, am) in split_revenue.items():
        cd = jobs[ji].get("completed_date")
        if not cd or len(str(cd)) < 7:
            no_date += 1
            continue
        ym = str(cd)[:7]  # "YYYY-MM"
        eq = jobs[ji]["equipment"] or "(?)"
        cat = "jiip" if jobs[ji]["is_jiip"] else "general"
        matrix[ym][eq][cat] += am

    log(f"   일자 없는 작업 (매칭된 것 중): {no_date:,}건 제외")

    # 2025-01 ~ 2026-03 (15개월)
    months = [f"2025-{m:02d}" for m in range(1, 13)] + [f"2026-{m:02d}" for m in range(1, 4)]
    all_eqs = set()
    for ym in matrix:
        all_eqs.update(matrix[ym].keys())
    eqs_sorted = sorted([e for e in all_eqs if e.endswith("호기")],
                        key=lambda e: sum(matrix[ym][e]["general"] + matrix[ym][e]["jiip"]
                                          for ym in months) * -1)

    # 일반 매출 월별 호기별 표
    log("\n" + "─" * 100)
    log("[일반 매출] 호기 × 월별 (단위: 백만원)")
    log("─" * 100)
    log(f"   {'호기':<8}" + "".join(f"{ym[2:]:>7}" for ym in months) + f"{'합계':>10}")
    for eq in eqs_sorted:
        row = f"   {eq:<8}"
        total = 0
        for ym in months:
            v = matrix[ym].get(eq, {}).get("general", 0)
            total += v
            row += f"{v / 1_000_000:>7.0f}" if v else f"{'.':>7}"
        row += f"{total / 1_000_000:>10,.0f}"
        log(row)

    # 지입 매출 월별
    log("\n" + "─" * 100)
    log("[지입 매출] 호기 × 월별 (단위: 백만원)")
    log("─" * 100)
    log(f"   {'호기':<8}" + "".join(f"{ym[2:]:>7}" for ym in months) + f"{'합계':>10}")
    for eq in eqs_sorted:
        row = f"   {eq:<8}"
        total = 0
        for ym in months:
            v = matrix[ym].get(eq, {}).get("jiip", 0)
            total += v
            row += f"{v / 1_000_000:>7.0f}" if v else f"{'.':>7}"
        row += f"{total / 1_000_000:>10,.0f}"
        log(row)

    # 합산 월별
    log("\n" + "─" * 100)
    log("[합계 매출] 호기 × 월별 (단위: 백만원)")
    log("─" * 100)
    log(f"   {'호기':<8}" + "".join(f"{ym[2:]:>7}" for ym in months) + f"{'합계':>10}")
    for eq in eqs_sorted:
        row = f"   {eq:<8}"
        total = 0
        for ym in months:
            v = matrix[ym].get(eq, {}).get("general", 0) + matrix[ym].get(eq, {}).get("jiip", 0)
            total += v
            row += f"{v / 1_000_000:>7.0f}" if v else f"{'.':>7}"
        row += f"{total / 1_000_000:>10,.0f}"
        log(row)

    # 월별 열 합계
    log(f"\n   {'월합':<8}" + "".join(
        f"{sum(matrix[ym][e]['general'] + matrix[ym][e]['jiip'] for e in eqs_sorted) / 1_000_000:>7.0f}"
        for ym in months
    ))

    # JSON
    out_data = {
        "months": months,
        "equipments": eqs_sorted,
        "monthly_matrix_won": {ym: dict(matrix[ym]) for ym in months},
    }
    OUT_JSON.write_text(json.dumps(out_data, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"\n→ JSON: {OUT_JSON}")
    log.close()


if __name__ == "__main__":
    main()
