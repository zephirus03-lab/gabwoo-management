"""
외주 단가표(.numbers) + 재단지시서 후가공 컬럼 분석.

목적: 호기별 매출에서 외주 비용을 빼서 호기별 마진 추정.

데이터원:
1. 견적 및 단가표 예시/01_매입(외주) 단가표(출판).numbers
2. 출판_생산 진행 현황/재단지시서_2025년.xlsx 의 '후가공' 컬럼 (외주처)
3. Supabase data.json
"""
from pathlib import Path
import json
import re
import openpyxl
from numbers_parser import Document
from collections import Counter

NUMBERS_PATH = Path("/Users/jack/dev/gabwoo/관리 대시보드/견적 및 단가표 예시/01_매입(외주) 단가표(출판).numbers")
JAEDAN_2025 = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/재단지시서_2025년.xlsx")
DATA_LOCAL = Path("/tmp/gw_check/data.json")
OUT = Path("/Users/jack/dev/gabwoo/관리 대시보드/scripts/output/outsource_pricing_explore.txt")


class Tee:
    def __init__(self, path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.f = open(path, "w", encoding="utf-8")

    def __call__(self, *args):
        msg = " ".join(str(a) for a in args)
        print(msg)
        self.f.write(msg + "\n")

    def close(self):
        self.f.close()


def section(log, title):
    log("\n" + "=" * 80)
    log(title)
    log("=" * 80)


def main():
    log = Tee(OUT)

    # ─── 1. .numbers 외주 단가표 파싱 ─────────────────────────────
    section(log, "1. 외주 단가표 .numbers 파싱")
    if not NUMBERS_PATH.exists():
        log(f"   ❌ 파일 없음: {NUMBERS_PATH}")
    else:
        doc = Document(str(NUMBERS_PATH))
        log(f"   파일: {NUMBERS_PATH.name}")
        log(f"   시트 개수: {len(doc.sheets)}")
        for sheet in doc.sheets:
            log(f"\n   ── 시트 '{sheet.name}' ──────────")
            for table in sheet.tables:
                log(f"   테이블 '{table.name}': {table.num_rows}행 x {table.num_cols}열")
                # 헤더 + 샘플 5행
                for row_idx in range(min(table.num_rows, 8)):
                    row_data = []
                    for col_idx in range(min(table.num_cols, 12)):
                        try:
                            cell = table.cell(row_idx, col_idx)
                            v = cell.value if cell else None
                            row_data.append(str(v)[:18] if v is not None else "-")
                        except Exception:
                            row_data.append("?")
                    log(f"      r{row_idx}: {row_data}")

    # ─── 2. 재단지시서 후가공 컬럼 분포 ────────────────────────────
    section(log, "2. 재단지시서 '후가공' 컬럼 — 외주처 분포")

    outsource_counts = Counter()
    eq_outsource = {}  # 호기 → Counter(외주처)

    # 2025 엑셀
    if JAEDAN_2025.exists():
        wb = openpyxl.load_workbook(str(JAEDAN_2025), read_only=True, data_only=True)
        ws = wb["재단지시서"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # 헤더
        header_idx = None
        for i, r in enumerate(rows[:5]):
            if r and any(c == "설비" for c in r if c):
                header_idx = i
                break
        if header_idx is not None:
            header = rows[header_idx]
            col = {n: i for i, n in enumerate(header) if n}
            outsrc_col = col.get("후가공")
            equip_col = col.get("설비")

            current_eq = None
            for r in rows[header_idx + 1:]:
                if not r:
                    continue
                if equip_col is not None and r[equip_col] and str(r[equip_col]).strip() not in ("-", "", "설비"):
                    current_eq = str(r[equip_col]).strip()
                if outsrc_col is not None and r[outsrc_col] and str(r[outsrc_col]).strip() not in ("-", ""):
                    o = str(r[outsrc_col]).strip()
                    outsource_counts[o] += 1
                    if current_eq:
                        eq_outsource.setdefault(current_eq, Counter())[o] += 1

    # Supabase data.json
    if DATA_LOCAL.exists():
        sdata = json.loads(DATA_LOCAL.read_text())
        for j in sdata.get("all_completed", []):
            o = j.get("post_process") or j.get("outsource")  # 키 미상
            # data.json 행에 어떤 키가 있는지 모르니 다 출력
            if not o:
                continue
            outsource_counts[o] += 1

    log(f"\n   외주처별 등장 빈도 (Top 20):")
    for o, n in outsource_counts.most_common(20):
        log(f"      {o:<25} {n:>6,}회")

    log(f"\n   호기별 주요 외주처 (Top 5씩):")
    for eq in sorted(eq_outsource.keys()):
        if not eq.endswith("호기"):
            continue
        log(f"\n   [{eq}]")
        for o, n in eq_outsource[eq].most_common(5):
            log(f"      {o:<25} {n:>5,}회")

    # ─── 3. data.json 행에 어떤 키가 있는지 다시 확인 ───────────────
    section(log, "3. Supabase data.json all_completed 행 키 (외주 정보 위치 찾기)")
    if DATA_LOCAL.exists():
        sdata = json.loads(DATA_LOCAL.read_text())
        sample = sdata.get("all_completed", [{}])[0]
        log(f"   샘플 1행 키: {list(sample.keys())}")
        log(f"   값:")
        for k, v in sample.items():
            log(f"      {k}={v}")

    log.close()


if __name__ == "__main__":
    main()
