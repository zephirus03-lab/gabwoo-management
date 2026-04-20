"""
재단지시서_2026년.xlsx 구조 탐색.

목표: 인쇄기별 매출 추정 가능성 점검
- 시트 목록 / 헤더 / 샘플
- 인쇄기·설비·호기 컬럼 유무
- 거래처·품명·관리번호 등 매출 매칭 키 유무
- 매일 업데이트 흔적 (날짜 컬럼)

⚠️ 읽기 전용 — 원본 절대 수정 금지 (루트 CLAUDE.md)
"""
from pathlib import Path
import openpyxl

PATH = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/재단지시서_2026년.xlsx")
OUT = Path("/Users/jack/dev/gabwoo/관리 대시보드/scripts/output/jaedan_jisi_structure.txt")


class Tee:
    def __init__(self, path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.f = open(path, "w", encoding="utf-8")

    def __call__(self, *args):
        msg = " ".join(str(a) for a in args)
        print(msg)
        self.f.write(msg + "\n")

    def close(self):
        self.f.close()


def main():
    log = Tee(OUT)

    if not PATH.exists():
        log(f"❌ 파일 없음: {PATH}")
        return

    log(f"파일: {PATH}")
    log(f"파일 크기: {PATH.stat().st_size / 1024:.0f} KB")
    log(f"수정 시각: {PATH.stat().st_mtime}")

    wb = openpyxl.load_workbook(str(PATH), read_only=True, data_only=True)
    log(f"\n시트 개수: {len(wb.sheetnames)}")
    log(f"시트 목록 (앞 30개): {wb.sheetnames[:30]}")
    if len(wb.sheetnames) > 30:
        log(f"  … +{len(wb.sheetnames) - 30}개 더")

    # 메인 시트 ('재단지시서') 분석
    target = "재단지시서" if "재단지시서" in wb.sheetnames else wb.sheetnames[0]
    log(f"\n{'=' * 80}")
    log(f"최근 시트 '{target}' 상세 분석")
    log(f"{'=' * 80}")
    ws = wb[target]
    rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
    log(f"\n앞 20행:")
    for i, r in enumerate(rows, 1):
        compact = [str(c)[:18] if c is not None else "-" for c in r[:30]]
        log(f"  row{i:>2}: {compact}")

    # 추가: 마지막 20행 (오늘 등록된 데이터 확인용)
    log(f"\n마지막 20행 (최신 업로드 확인):")
    last_rows = list(ws.iter_rows(min_row=15335, max_row=15354, values_only=True))
    for i, r in enumerate(last_rows, 15335):
        if r and any(c for c in r):
            compact = [str(c)[:18] if c is not None else "-" for c in r[:30]]
            log(f"  row{i:>5}: {compact}")

    # 헤더 후보 찾기 (관리번호/거래처/품명/인쇄기 등 키워드)
    log(f"\n헤더 키워드 탐색:")
    header_idx = None
    for i, r in enumerate(rows):
        if not r:
            continue
        text = " ".join(str(c) for c in r if c)
        if any(k in text for k in ["관리번호", "거래처", "품명", "인쇄기", "호기", "수주번호"]):
            header_idx = i
            log(f"  row{i + 1}이 헤더로 추정")
            log(f"  내용: {[c for c in r if c is not None]}")
            break

    if header_idx is not None:
        header = rows[header_idx]
        # 인쇄기/설비 관련 컬럼
        machine_cols = []
        for j, c in enumerate(header):
            if c and any(k in str(c) for k in ["인쇄기", "설비", "기계", "호기"]):
                machine_cols.append((j, str(c)))
        log(f"\n  → 인쇄기/설비 관련 컬럼: {machine_cols if machine_cols else '❌ 없음'}")

        match_cols = []
        for j, c in enumerate(header):
            if c and any(k in str(c) for k in ["관리번호", "수주번호", "견적", "지시번호", "재단지시"]):
                match_cols.append((j, str(c)))
        log(f"  → ERP 매칭 가능 키 컬럼: {match_cols if match_cols else '없음'}")

        # 데이터 샘플 5행
        log(f"\n데이터 샘플 5행:")
        sample_rows = list(ws.iter_rows(min_row=header_idx + 2, max_row=header_idx + 7, values_only=True))
        for r in sample_rows:
            if r:
                compact = [str(c)[:25] if c is not None else "-" for c in r[:25]]
                log(f"  {compact}")

    # 시트 이름 패턴 분석 (날짜인지)
    log(f"\n{'=' * 80}")
    log("시트 이름 패턴")
    log(f"{'=' * 80}")
    import re
    date_pattern_count = 0
    for s in wb.sheetnames:
        if re.match(r"\d{2,4}[-./]?\d{1,2}[-./]?\d{0,2}", s):
            date_pattern_count += 1
    log(f"   날짜 형식으로 보이는 시트: {date_pattern_count} / {len(wb.sheetnames)}")

    # 마지막 5개 시트 이름
    log(f"   마지막 5개 시트: {wb.sheetnames[-5:]}")

    # 시트별 행 수 (대략)
    log(f"\n시트별 데이터 규모 (마지막 5개):")
    for s in wb.sheetnames[-5:]:
        ws = wb[s]
        # max_row는 read_only에서 None일 수 있으니 iter로 카운트
        n = sum(1 for _ in ws.iter_rows(values_only=True))
        log(f"   {s}: {n}행")

    wb.close()
    log.close()


if __name__ == "__main__":
    main()
