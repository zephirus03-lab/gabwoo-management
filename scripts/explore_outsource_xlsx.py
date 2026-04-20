"""01_매입 단가표(출판).xlsx 구조 탐색."""
from pathlib import Path
import openpyxl

P = Path("/Users/jack/dev/gabwoo/출판_생산 진행 현황/01_매입 단가표(출판).xlsx")
wb = openpyxl.load_workbook(str(P), read_only=True, data_only=True)

print(f"파일: {P.name}")
print(f"시트 개수: {len(wb.sheetnames)}")
print(f"시트 목록: {wb.sheetnames}")
print()

for sn in wb.sheetnames:
    ws = wb[sn]
    n = sum(1 for _ in ws.iter_rows(values_only=True))
    print(f"── 시트 '{sn}' ({n}행) ───────────────")
    rows = list(ws.iter_rows(min_row=1, max_row=12, values_only=True))
    for i, r in enumerate(rows, 1):
        if r and any(c for c in r):
            compact = [str(c)[:20] if c is not None else "-" for c in r[:12]]
            print(f"  r{i:>2}: {compact}")
    print()
wb.close()
