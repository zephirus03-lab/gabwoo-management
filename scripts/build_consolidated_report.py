"""
경영진단 잠정 보고서 — 통합본

3개 분석 결과를 하나의 Excel로 합쳐 경영진 보고용으로 정리합니다.
  - 경영진단_가설검증_YYYYMMDD.xlsx  (4트랙)
  - 거래처패턴분석_YYYYMMDD.xlsx     (A/B/C/D/E/F)
  - 원가반영_점검_YYYYMMDD.xlsx      (반영충분/일부반영/역행)

출력: 잠정_경영진단보고서_YYYYMMDD.xlsx
"""
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent / "output"
TODAY = datetime.now().strftime("%Y%m%d")

SRC1 = OUT_DIR / f"경영진단_가설검증_{TODAY}.xlsx"
SRC2 = OUT_DIR / f"거래처패턴분석_{TODAY}.xlsx"
SRC3 = OUT_DIR / f"원가반영_점검_{TODAY}.xlsx"
OUT  = OUT_DIR / f"잠정_경영진단보고서_{TODAY}.xlsx"

# 색상
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
RED_FILL = PatternFill("solid", fgColor="FFE7E6")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
BORDER = Border(left=Side(style="thin", color="BFBFBF"),
                right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="thin", color="BFBFBF"))


def read(src, sheet):
    return pd.read_excel(src, sheet_name=sheet)


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
                ln = sum(2 if ord(c) > 127 else 1 for c in v)
                if ln > max_len:
                    max_len = ln
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def write_cover(ws):
    ws["A1"] = "갑우문화사 경영진단 — 잠정 보고서"
    ws["A1"].font = Font(bold=True, size=18, color="1F3864")
    ws.merge_cells("A1:F1")
    ws["A2"] = f"작성일: {datetime.now():%Y-%m-%d}  /  분석범위: 2021~2025 (ERP 직접 조회)  /  대상: 갑우문화사 (CD_FIRM=7000)"
    ws["A2"].font = Font(size=10, color="555555")
    ws.merge_cells("A2:F2")

    # 가설
    ws["A4"] = "[출발 질문]"
    ws["A4"].font = SECTION_FONT
    ws["A4"].fill = SECTION_FILL
    ws.merge_cells("A4:F4")
    questions = [
        "1. 매출은 줄어드는데 매입 원가 비중이 늘어나고 있지 않은가?",
        "2. 과거 100개 팔던 회사에 주던 할인을 지금 50개 팔면서도 그대로 유지하고 있지 않은가?",
        "3. 즉, 비싸게 사서 싸게 팔고 있지 않은가?",
    ]
    for i, q in enumerate(questions, start=5):
        ws[f"A{i}"] = q
        ws.merge_cells(f"A{i}:F{i}")

    # 결론
    ws["A9"] = "[결론 — 데이터로 확인된 사실]"
    ws["A9"].font = SECTION_FONT
    ws["A9"].fill = SECTION_FILL
    ws.merge_cells("A9:F9")
    findings = [
        ("✅", "용지 매입 단가가 5년간 +70.6% 폭등 (2021 27,553원 → 2025 46,991원). 수량은 -30%인데 지출은 +18%."),
        ("✅", "용지 외 전체 매출원가율도 상승: 2021 14.2% → 2025 18.5% → 2026 Q1 25.3%."),
        ("✅", "판매 단가(전체 가중평균)는 오히려 -10% 하락 (2023 피크 67.20 → 2025 60.45)."),
        ("✅", "원가 +22.6%(2023→2025) 기간 중 주요 거래처 24곳은 오히려 단가 인하. 91억 손실."),
        ("⚠️", "원래 가설(물량↓·단가 그대로)은 5.9억 / 4개사로 소액. 실제 출혈은 더 큰 두 패턴에 있음."),
        ("🔴", "최악 패턴 A (단가 깎아주고 물량도 잃음): 15개사 / 94.4억 손실 — 한국조폐공사·케이랩·지에스리테일 등"),
        ("🔴", "최악 패턴 B (단가 올렸다 고객 이탈): 16개사 / 62.4억 손실 — (주)코스알엑스 단독 34.6억"),
        ("⚫", "완전이탈 19개사 / 40.0억 손실. 이탈 사유 기록 없음."),
        ("💡", "반면 일부반영(단가 조금 인상) 8개사는 매출 오히려 +21.5억. '적정 인상 가능'을 증명."),
    ]
    for i, (icon, txt) in enumerate(findings, start=10):
        ws[f"A{i}"] = icon
        ws[f"B{i}"] = txt
        ws.merge_cells(f"B{i}:F{i}")

    # 시트 안내
    start = 10 + len(findings) + 2
    ws[f"A{start}"] = "[본 보고서 구성]"
    ws[f"A{start}"].font = SECTION_FONT
    ws[f"A{start}"].fill = SECTION_FILL
    ws.merge_cells(f"A{start}:F{start}")
    guide = [
        ("1. 매출·원가 추이", "연도별 매출·용지매입·외주매입·매출원가율"),
        ("2. 용지 단가", "연도별 가중평균 + 지종별 TOP 지종"),
        ("3. 거래처 패턴 요약", "A~F 패턴 분류 (2023 매출 1억 이상)"),
        ("4-A. 패턴 A 상세", "단가 인하 + 물량 감소 = 양보실패 (15개사)"),
        ("4-B. 패턴 B 상세", "단가 인상 → 고객 이탈 (16개사)"),
        ("4-C. 패턴 C 상세", "완전 이탈 (19개사)"),
        ("5. 원가 반영 판정", "용지 원가 +22.6% 대비 거래처별 단가 인상 여부"),
        ("5-1. 역행(단가인하) 24개사", "원가 올랐는데 단가 내린 거래처 — 가장 아픈 지점"),
        ("5-2. 일부반영 8개사", "단가 일부만 인상했으나 고객 유지·매출 증가 — 반박 증거"),
        ("5-3. 반영충분 20개사", "원가 상승률 이상 인상 성공"),
        ("6. 후속 제안", "다음 행동 후보"),
    ]
    for i, (sheet, desc) in enumerate(guide, start=start+1):
        ws[f"A{i}"] = sheet
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = desc
        ws.merge_cells(f"B{i}:F{i}")

    # 출처
    last = start + len(guide) + 2
    ws[f"A{last}"] = "[데이터 출처]"
    ws[f"A{last}"].font = SECTION_FONT
    ws[f"A{last}"].fill = SECTION_FILL
    ws.merge_cells(f"A{last}:F{last}")
    src_rows = [
        "SNOTES ERP 직접 조회 (SAL_SALESH/L, SCT_TAXH/L, PUR_ETCCLSH, viewGabwoo_마감, MAS_CUST)",
        "비교 기준: 갑우문화사 감사보고서 제28~33기 (FY2020~FY2025, DART 공시)",
        "제외: 비피앤피(CD_FIRM=8000) — SNOTES에 데이터 없음, 별도 시스템 필요",
    ]
    for i, s in enumerate(src_rows, start=last+1):
        ws[f"A{i}"] = s
        ws.merge_cells(f"A{i}:F{i}")

    ws.column_dimensions["A"].width = 4
    for col in "BCDEF":
        ws.column_dimensions[col].width = 20


def write_df_with_header(ws, df, title, note=None, color="header"):
    ws["A1"] = title
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 5))
    row_offset = 2
    if note:
        ws.cell(row=2, column=1, value=note).font = Font(italic=True, size=10, color="555555")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(df.columns), 5))
        row_offset = 3
    # 헤더
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=row_offset, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    # 본문
    for i, row in enumerate(df.itertuples(index=False), start=row_offset+1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = BORDER
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")
                if abs(val) >= 1000:
                    c.number_format = "#,##0"
                elif isinstance(val, float):
                    c.number_format = "#,##0.00"
    autosize(ws)


def main():
    print("📑 통합 보고서 생성 중...")

    # 데이터 로드
    paper_yearly = read(SRC1, "1a_용지매입_연도별")
    paper_type = read(SRC1, "1b_용지단가_지종×연도")
    pur_yearly = read(SRC1, "1c_외주매입_연도별")
    sales_yearly = read(SRC1, "2a_매출_연도별지표")
    cost_ratio = read(SRC1, "3b_원가율_연도별")

    pattern_summary = read(SRC2, "0_패턴요약")
    pat_A = read(SRC2, "A_단가↓물량↓")
    pat_B = read(SRC2, "B_단가↑이탈")
    pat_C = read(SRC2, "C_완전이탈")

    cost_summary = read(SRC3, "0_요약")
    cost_reverse = read(SRC3, "역행_단가인하")
    cost_partial = read(SRC3, "일부반영")
    cost_full = read(SRC3, "반영충분")

    # 매출·원가 추이 — 합쳐서 한 표로
    overview = cost_ratio.copy()
    overview["매출공급가(억)"] = (overview["매출공급가"]/1e8).round(1)
    overview["용지매입(억)"] = (overview["용지매입"]/1e8).round(1)
    overview["외주매입(억)"] = (overview["외주매입"]/1e8).round(1)
    overview["총매입(억)"] = (overview["총매입"]/1e8).round(1)
    overview["매입매출비(%)"] = overview["매입매출비(%)"].round(1)
    overview["용지매입비중(%)"] = overview["용지매입비중(%)"].round(1)
    overview = overview[["연도","매출공급가(억)","용지매입(억)","외주매입(억)","총매입(억)","매입매출비(%)","용지매입비중(%)"]]

    # 용지 단가 연도별 (간소화)
    paper_yearly_clean = paper_yearly[["연도","총수량","총매입액","단가가중평균","건수","제조사수"]].copy()
    paper_yearly_clean["총매입액(억)"] = (paper_yearly_clean["총매입액"]/1e8).round(1)
    paper_yearly_clean["단가가중평균"] = paper_yearly_clean["단가가중평균"].round(0)
    paper_yearly_clean = paper_yearly_clean[["연도","총수량","총매입액(억)","단가가중평균","건수","제조사수"]]

    # 외주매입 (외주 = PUR_ETCCLSH 기타마감)
    pur_yearly_clean = pur_yearly.copy()
    pur_yearly_clean["총매입액(억)"] = (pur_yearly_clean["총매입액"]/1e8).round(2)
    pur_yearly_clean = pur_yearly_clean[["연도","마감건수","총매입액(억)","거래처수"]]

    # 매출 (sales_yearly는 중복집계 영향, 건당매출만 사용 가능. 재구성)
    sales_clean = overview[["연도","매출공급가(억)"]].copy()

    # 판매 단가 가중평균 (line 기준, Track 2a의 단가가중평균 → 중복과 무관)
    sales_price = sales_yearly[["연도","매출건수","단가가중평균","건당매출(백만원)"]].copy()
    sales_price["단가가중평균"] = sales_price["단가가중평균"].round(2)
    sales_price["건당매출(백만원)"] = sales_price["건당매출(백만원)"].round(2)

    # openpyxl로 Excel 작성
    from openpyxl import Workbook
    wb = Workbook()

    # 0. 표지
    ws0 = wb.active
    ws0.title = "0_표지"
    write_cover(ws0)

    # 1. 매출·원가 추이
    ws1 = wb.create_sheet("1_매출·원가추이")
    write_df_with_header(ws1, overview,
        "1. 매출·원가 추이 (연도별, 갑우문화사 본체만)",
        "용지: viewGabwoo_마감 / 외주: PUR_ETCCLSH 기타마감 / 매출: SAL_SALESH AM(공급가)")

    # 1-2. 판매단가 추이
    ws1b = wb.create_sheet("1-2_판매단가_추이")
    write_df_with_header(ws1b, sales_price,
        "1-2. 판매 단가 가중평균 추이",
        "단가가중평균 = Σ(라인금액) / Σ(라인수량), 라인 기준")

    # 2. 용지 단가
    ws2 = wb.create_sheet("2_용지단가_추이")
    write_df_with_header(ws2, paper_yearly_clean,
        "2. 용지 매입 단가 추이 (5년 +70.6% 폭등)",
        "단가가중평균 = 연간 총 공급가액 / 총 수량. 단위: 원")

    # 2-2. 지종별
    ws2b = wb.create_sheet("2-2_용지_지종별")
    # 지종별에서 상위 10개만 + 칼럼 정리
    pt_top = paper_type.head(10).copy()
    for y in [2021, 2022, 2023, 2024, 2025]:
        if y in pt_top.columns:
            pt_top[y] = pt_top[y].round(0)
    if "단가변화율_21→25" in pt_top.columns:
        pt_top["단가변화율_21→25"] = pt_top["단가변화율_21→25"].round(1)
    if "총매입액5년" in pt_top.columns:
        pt_top["총매입액5년(억)"] = (pt_top["총매입액5년"]/1e8).round(1)
        pt_top = pt_top.drop(columns=["총매입액5년"])
    write_df_with_header(ws2b, pt_top,
        "2-2. 용지 지종별 단가 YoY (매입액 TOP 10)",
        "모든 주요 지종에서 5년간 20~30% 이상 인상")

    # 3. 외주매입
    ws3 = wb.create_sheet("3_외주매입")
    write_df_with_header(ws3, pur_yearly_clean,
        "3. 외주/부자재 매입 추이 (PUR_ETCCLSH 기타마감)",
        "용지 대비 규모 작음. 연간 1.7~2.5억 수준. 증가 추세 없음.")

    # 4. 거래처 패턴 요약
    ws4 = wb.create_sheet("4_거래처패턴_요약")
    # 포맷: 매출 억원 컬럼 이미 존재
    ps = pattern_summary.copy()
    for c in ["매출_2023(억)","매출_2025(억)","매출손실(억)"]:
        ps[c] = pd.to_numeric(ps[c], errors="coerce").round(1)
    write_df_with_header(ws4, ps,
        "4. 거래처 패턴 분류 (2023 매출 1억 이상, 2023→2025 비교)",
        "A+B+C 합계 = 196.8억 손실 (감사보고서 매출 감소 135억과 같은 궤도)")

    # 4-A/B/C 상세
    ws4a = wb.create_sheet("4-A_단가↓물량↓_15개사")
    write_df_with_header(ws4a, pat_A,
        "4-A. 단가 깎아주고도 물량 잃음 (양보실패) — 15개사 / 94.4억 손실",
        "한국조폐공사·지에스리테일·케이랩이 핵심 (이 셋만 74억)")

    ws4b = wb.create_sheet("4-B_단가↑이탈_16개사")
    write_df_with_header(ws4b, pat_B,
        "4-B. 단가 올렸다 고객 이탈 — 16개사 / 62.4억 손실",
        "(주)코스알엑스 단독 34.6억 손실 (단가 +72% 인상 후 물량 93% 이탈)")

    ws4c = wb.create_sheet("4-C_완전이탈_19개사")
    write_df_with_header(ws4c, pat_C,
        "4-C. 완전 이탈 — 19개사 / 40.0억 손실",
        "이탈 사유 ERP에 기록 없음. 한국미스미·오메가포인트·농민신문사 등.")

    # 5. 원가 반영 판정
    ws5 = wb.create_sheet("5_원가반영_요약")
    cs = cost_summary.copy()
    cs["평균단가Δ"] = pd.to_numeric(cs["평균단가Δ"], errors="coerce")
    # 이상치(수만~수백만%) 주의 메모만 추가
    write_df_with_header(ws5, cs,
        "5. 용지 원가 상승률(+22.6%, 2023→2025) 대비 거래처별 단가 인상 판정",
        "평균단가Δ에 일부 이상치(소량 고단가 튐) 섞여 있음 — 개별 표에서 확인")

    ws5r = wb.create_sheet("5-1_역행_24개사")
    write_df_with_header(ws5r, cost_reverse,
        "5-1. 원가 올랐는데 단가 내린 거래처 24개사 (가장 아픈 지점)",
        "한국조폐공사·지에스리테일·케이랩 등 대형 고객이 여기 몰림")

    ws5p = wb.create_sheet("5-2_일부반영_8개사")
    write_df_with_header(ws5p, cost_partial,
        "5-2. 단가 일부만 인상한 8개사 — 그럼에도 매출 +21.5억 증가 (반박 증거)",
        "지담디앤피·교원·동행복권 — '적정 인상 가능' 증명")

    ws5f = wb.create_sheet("5-3_반영충분_20개사")
    write_df_with_header(ws5f, cost_full,
        "5-3. 원가 상승률 이상 인상한 20개사",
        "교원구몬(+29.9%)·이투스에듀(+46%) 성공. 코스알엑스(+72%)는 과도 인상 후 이탈.")

    # 6. 후속 제안
    ws6 = wb.create_sheet("6_후속제안")
    ws6["A1"] = "6. 후속 제안"
    ws6["A1"].font = SECTION_FONT
    ws6["A1"].fill = SECTION_FILL
    ws6.merge_cells("A1:F1")
    proposals = [
        ("우선순위", "행동", "기대효과"),
        ("1순위", "역행 24개 거래처를 담당 영업자별로 분해 — 특정 영업자 쏠림 확인",
                  "조폐공사·지에스리테일·케이랩 74억 손실의 책임 추적"),
        ("1순위", "코스알엑스(B 1위, 단독 34.6억) 단가 이력 시계열 분석",
                  "단가 72% 인상 시점·사유·담당자 파악 → 재협상 가능성 점검"),
        ("2순위", "역행 24곳의 단가 이력과 용지 단가 월별 추이 매칭",
                  "원가 급등 시점(2022 Q2, 2024 Q3) 이후 단가 조정 여부 정량화"),
        ("2순위", "완전이탈 19곳의 이탈 사유 수집 체계 구축",
                  "다음 이탈 예방 — 40억 손실이 재발하지 않도록"),
        ("3순위", "단가표 암묵지 공용화 (산출기 프로젝트 '견적계산기/' 추진)",
                  "영업자별 단가 결정 편차 제거 — 역행 패턴 근본 차단"),
        ("3순위", "용지 매입 재협상 타겟 선정 (지종별 +30% 이상 인상된 제조사 리스트)",
                  "매입 원가 절감 → 원가율 18% → 15% 복귀 시 매출총이익 10억+"),
    ]
    for i, row in enumerate(proposals, start=3):
        for j, val in enumerate(row, start=1):
            c = ws6.cell(row=i, column=j, value=val)
            if i == 3:
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
            else:
                c.border = BORDER
                if j == 1:
                    c.font = Font(bold=True, color="C00000" if "1" in str(val) else "1F3864")
            c.alignment = Alignment(wrap_text=True, vertical="center")
    ws6.column_dimensions["A"].width = 10
    ws6.column_dimensions["B"].width = 55
    ws6.column_dimensions["C"].width = 55

    wb.save(OUT)
    print(f"✅ 완료: {OUT}")
    print(f"   시트 수: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"   - {s}")


if __name__ == "__main__":
    main()
